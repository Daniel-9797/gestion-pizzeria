"""
🍕 Pizzería - Sistema de Cierre de Caja Diario
Conecta con PostgreSQL para guardar la facturación diaria.
"""

import tkinter as tk
from tkinter import ttk, messagebox
import psycopg2
from psycopg2 import sql
from datetime import date, datetime, timedelta
import os
import sys
import locale
import unicodedata
import time as time_module
import threading
import requests
from bs4 import BeautifulSoup
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from tkinter import filedialog, simpledialog
import platform

# ─────────────────────────────────────────────
#  AJUSTES ESPECÍFICOS DE PLATAFORMA (Windows / Mac / Linux)
#  Mismo app.py para todos los sistemas — se adapta solo, no hace falta
#  mantener una versión distinta por máquina.
#
#  Sin esto, en Windows pasan dos cosas que hacen que la app "se vea vieja":
#  1) Falta avisar a Windows de que la app sabe de pantallas de alta densidad
#     (DPI). Si no se avisa, Windows la escala como si fuera una imagen
#     (bitmap stretching) y el texto sale borroso/pixelado.
#  2) Las fuentes "Avenir Next" y "Menlo" son exclusivas de macOS. En Windows
#     no existen, así que Tk cae a una fuente de repuesto genérica que no
#     pega con el diseño (tamaños, interlineado y grosor distintos).
# ─────────────────────────────────────────────
_SO = platform.system()   # "Windows", "Darwin" (Mac) o "Linux"

if _SO == "Windows":
    try:
        import ctypes
        ctypes.windll.shcore.SetProcessDpiAwareness(1)   # Windows 8.1+
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()    # Windows Vista/7 (fallback)
        except Exception:
            pass
    FONT_SANS = "Segoe UI"        # fuente nativa de Windows 10/11
    FONT_MONO = "Consolas"        # monoespaciada nativa de Windows
elif _SO == "Darwin":
    FONT_SANS = "Avenir Next"
    FONT_MONO = "Menlo"
else:
    FONT_SANS = "DejaVu Sans"
    FONT_MONO = "DejaVu Sans Mono"

# ─────────────────────────────────────────────
#  CONFIGURACIÓN DE BASE DE DATOS
#  Edita estos valores o crea un .env
# ─────────────────────────────────────────────
# ─────────────────────────────────────────────
#  CONFIGURACIÓN (desde variables de entorno / .env)
#  Los secretos (credenciales de BD, token de API, contraseñas) se leen de
#  variables de entorno. Copia .env.example a .env y rellena tus valores.
#  El archivo .env NO se sube al repositorio (está en .gitignore).
# ─────────────────────────────────────────────
import os
try:
    from dotenv import load_dotenv
    load_dotenv()  # carga el .env si existe
except Exception:
    pass  # si python-dotenv no está instalado, se usan las env vars del sistema

DB_CONFIG = {
    "host":     os.getenv("DB_HOST", "localhost"),
    "port":     os.getenv("DB_PORT", "5432"),
    "database": os.getenv("DB_NAME", "productos_tpv"),
    "user":     os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASSWORD", ""),
}

# ─────────────────────────────────────────────
#  COPIA EN LA NUBE (espejo de solo lectura del último cierre)
#  La app SIEMPRE trabaja en local contra DB_CONFIG. Al guardar un cierre,
#  además empuja una copia a esta base en la nube (p.ej. Neon). Si no hay
#  internet, el guardado local NO se ve afectado: solo falla la sync, con aviso.
#  Se desactiva solo si no hay host configurado.
# ─────────────────────────────────────────────
CLOUD_DB_CONFIG = {
    "host":     os.getenv("CLOUD_DB_HOST", ""),
    "port":     os.getenv("CLOUD_DB_PORT", "5432"),
    "database": os.getenv("CLOUD_DB_NAME", ""),
    "user":     os.getenv("CLOUD_DB_USER", ""),
    "password": os.getenv("CLOUD_DB_PASSWORD", ""),
    "sslmode":  os.getenv("CLOUD_DB_SSLMODE", "require"),
    "connect_timeout": 8,
}
SYNC_NUBE_ACTIVADO = bool(CLOUD_DB_CONFIG["host"])

# ─────────────────────────────────────────────
#  CONTRASEÑAS DE ACCESO A LA APP (desde variables de entorno)
#  - APP_PASSWORD_EMPLEADO: acceso limitado a las pestañas operativas.
#  - APP_PASSWORD_MASTER:   acceso completo.
#  Nota: el control de acceso por contraseña en una app de escritorio local es
#  una barrera básica (evitar cotilleo del personal), no una protección fuerte.
# ─────────────────────────────────────────────
PASSWORD_EMPLEADO = os.getenv("APP_PASSWORD_EMPLEADO", "caja123")
PASSWORD_MASTER    = os.getenv("APP_PASSWORD_MASTER", "cambiame")

# Nombres de atributo de las pestañas (deben coincidir con los que se crean
# en _build_notebook: self.tab_cierre, self.tab_arqueo, etc.)
# Pon aquí SOLO las que debe ver un empleado; el resto se ocultan.
TABS_VISIBLES_EMPLEADO = ["tab_cierre", "tab_arqueo", "tab_historico", "tab_prevision", "tab_mapa"]

_MESES_ES_GLOBAL = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                     "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

# ─────────────────────────────────────────────
#  COLORES Y ESTILOS  (paleta pizzería)
# ─────────────────────────────────────────────
C = {
    "bg":        "#15141A",   # fondo principal — gris-azulado casi negro, más moderno que el marrón
    "panel":     "#1F1E26",   # paneles secundarios
    "card":      "#262530",   # tarjetas — gris-azulado oscuro neutro
    "accent":    "#FF7F32",   # Pantone 158 — naranja Carpi Pizza
    "accent2":   "#FFC72C",   # Pantone 123 — amarillo Carpi Pizza
    "text":      "#F5F4F8",   # texto principal — casi blanco, frío
    "muted":     "#9A98A6",   # texto secundario — gris-violeta
    "success":   "#22C55E",   # verde positivo
    "danger":    "#D52B1E",   # Pantone 1795 — rojo Carpi Pizza, también usado para negativo/alerta
    "border":    "#34323E",   # bordes sutiles
    "entry_bg":  "#15141A",   # fondo inputs
    "entry_fg":  "#F5F4F8",
    "highlight": "#FF7F32",
}

FONT_TITLE  = (FONT_SANS, 22, "bold")
FONT_LABEL  = (FONT_SANS, 11)
FONT_ENTRY  = (FONT_MONO, 13)
FONT_SMALL  = (FONT_SANS, 9)
FONT_TOTAL  = (FONT_MONO, 16, "bold")
FONT_BUTTON = (FONT_SANS, 12, "bold")

# ─────────────────────────────────────────────
#  BASE DE DATOS
# ─────────────────────────────────────────────

def get_connection():
    return psycopg2.connect(**DB_CONFIG)

def get_cloud_connection():
    return psycopg2.connect(**CLOUD_DB_CONFIG)

# Tablas que se sincronizan a la nube.
#   pk        → clave primaria para el upsert (ON CONFLICT)
#   col_fecha → columna de fecha para sincronización incremental (solo el día
#               en curso). Si es None, la tabla se sincroniza ENTERA (son tablas
#               pequeñas o que no tienen una fecha de "día" clara).
_TABLAS_SYNC = {
    "cierres_caja": {"pk": "id",    "col_fecha": "fecha"},
    "arqueos":      {"pk": "id",    "col_fecha": None},     # id = YYYYMMDD, se filtra aparte
    "festivos":     {"pk": "fecha", "col_fecha": None},     # tabla pequeña, entera
    "eventos":      {"pk": "id",    "col_fecha": None},     # tabla pequeña, entera
    "contexto_dia": {"pk": "fecha", "col_fecha": "fecha"},
    "previsiones":  {"pk": "fecha", "col_fecha": "fecha"},
}

def sincronizar_con_nube(dias_recientes: int = 3, completo: bool = False) -> tuple:
    """
    Empuja una copia de las tablas (ver _TABLAS_SYNC) desde la BD LOCAL a la BD
    de la NUBE (CLOUD_DB_CONFIG). Solo lectura en local, solo escritura en la
    nube — la app nunca lee de la nube, así que no hay conflictos posibles.

    EFICIENCIA: por defecto solo sube los últimos `dias_recientes` días de las
    tablas con fecha (cierres_caja, contexto_dia, previsiones), en vez de todo el
    histórico cada vez. Se usa una ventana de varios días (no solo hoy) como red
    de seguridad: si un día se cerró sin internet, los cierres siguientes
    recuperan ese hueco automáticamente.
      · completo=True  → sube TODO el histórico (botón de respaldo manual).
      · arqueos/festivos/eventos → siempre enteras (son pequeñas).

    Devuelve (ok: bool, mensaje: str). No lanza excepciones — cualquier fallo
    (típicamente: sin internet) se captura y se informa en el mensaje, para que
    nunca pueda interferir con un guardado local que sí haya funcionado.
    """
    if not SYNC_NUBE_ACTIVADO:
        return False, "Sincronización con la nube desactivada (SYNC_NUBE_ACTIVADO = False)"
    if not CLOUD_DB_CONFIG.get("host"):
        return False, "Faltan los datos de CLOUD_DB_CONFIG (host vacío)"

    try:
        conn_local = get_connection()
        conn_nube = get_cloud_connection()
    except Exception as e:
        return False, f"No se pudo conectar (¿sin internet?): {e}"

    # Fecha de corte para la sincronización incremental
    fecha_corte = (date.today() - timedelta(days=max(0, dias_recientes - 1)))
    corte_str = fecha_corte.isoformat()
    corte_id = fecha_corte.strftime("%Y%m%d")   # para arqueos, cuyo id es YYYYMMDD

    try:
        init_db(conn=conn_nube)   # asegura que las tablas existen en la nube
        cur_local = conn_local.cursor()
        cur_nube = conn_nube.cursor()
        total_filas = 0

        for tabla, info in _TABLAS_SYNC.items():
            pk = info["pk"]
            col_fecha = info["col_fecha"]

            # Construir el SELECT (incremental o completo)
            if completo or col_fecha is None:
                if tabla == "arqueos" and not completo:
                    # arqueos: id = YYYYMMDD; filtrar por los últimos días
                    cur_local.execute(
                        f"SELECT * FROM {tabla} WHERE id >= %s", (corte_id,))
                else:
                    cur_local.execute(f"SELECT * FROM {tabla}")
            else:
                cur_local.execute(
                    f"SELECT * FROM {tabla} WHERE {col_fecha} >= %s", (corte_str,))

            filas = cur_local.fetchall()
            columnas = [d[0] for d in cur_local.description]
            if not filas:
                continue

            placeholders = ", ".join(["%s"] * len(columnas))
            cols_sql = ", ".join(columnas)
            updates = ", ".join(f"{c} = EXCLUDED.{c}" for c in columnas if c != pk)
            sql_upsert = f"""
                INSERT INTO {tabla} ({cols_sql}) VALUES ({placeholders})
                ON CONFLICT ({pk}) DO UPDATE SET {updates}
            """
            for fila in filas:
                cur_nube.execute(sql_upsert, fila)
            total_filas += len(filas)

        conn_nube.commit()
        modo = "completo" if completo else f"últimos {dias_recientes} días"
        return True, f"Sincronizado ✓ ({total_filas} filas, {modo})"
    except Exception as e:
        try:
            conn_nube.rollback()
        except Exception:
            pass
        return False, f"Error al sincronizar: {e}"
    finally:
        try:
            cur_local.close(); conn_local.close()
        except Exception:
            pass
        try:
            cur_nube.close(); conn_nube.close()
        except Exception:
            pass

def init_db(conn=None):
    """Crea las tablas si no existen. Si no se pasa `conn`, usa la BD local
    (DB_CONFIG); se puede pasar una conexión a otra BD (p.ej. la nube) para
    crear ahí las mismas tablas con idéntico esquema."""
    cerrar_al_final = conn is None
    if conn is None:
        conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cierres_caja (
            id                CHAR(8) PRIMARY KEY,
            fecha             DATE NOT NULL,
            facturacion       NUMERIC(10,2) DEFAULT 0,
            gastos            NUMERIC(10,2) DEFAULT 0,
            visa              NUMERIC(10,2) DEFAULT 0,
            internet          NUMERIC(10,2) DEFAULT 0,
            justeat           NUMERIC(10,2) DEFAULT 0,
            glovo             NUMERIC(10,2) DEFAULT 0,
            uber              NUMERIC(10,2) DEFAULT 0,
            ticket_restaurant NUMERIC(10,2) DEFAULT 0,
            ingreso_banco     NUMERIC(10,2) DEFAULT 0,
            z_caja            NUMERIC(10,2) DEFAULT 0,
            facturacion_total NUMERIC(10,2) DEFAULT 0,
            notas             TEXT,
            creado_en         TIMESTAMP DEFAULT NOW(),
            actualizado_en    TIMESTAMP DEFAULT NOW()
        )
    """)
    # Columna tipo_dia: normal / festivo / vispera / fm
    cur.execute("""
        ALTER TABLE cierres_caja
        ADD COLUMN IF NOT EXISTS tipo_dia VARCHAR(10) NOT NULL DEFAULT 'normal'
    """)
    # Tabla de festivos y vísperas (fecha exacta + tipo)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS festivos (
            fecha    DATE PRIMARY KEY,
            tipo     VARCHAR(10) NOT NULL,   -- 'festivo' | 'vispera' | 'fm'
            nombre   TEXT,                   -- ej. "Sant Jordi", "Fiesta Mayor"
            creado_en TIMESTAMP DEFAULT NOW()
        )
    """)
    # Tabla de eventos puntuales (no históricos, solo alerta)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS eventos (
            id        SERIAL PRIMARY KEY,
            fecha     DATE NOT NULL,
            tipo      VARCHAR(20) NOT NULL,  -- 'futbol' | 'motogp' | 'f1' | 'otro'
            descripcion TEXT NOT NULL,       -- ej. "Barça - Real Madrid"
            creado_en TIMESTAMP DEFAULT NOW()
        )
    """)
    # Tabla de CONTEXTO del día: una fila por día, una columna por característica.
    # Se rellena al hacer el Cierre Diario. Diseñada para crecer (ALTER TABLE).
    cur.execute("""
        CREATE TABLE IF NOT EXISTS contexto_dia (
            fecha       DATE PRIMARY KEY,
            lluvia_mm   NUMERIC(5,1),
            llueve      BOOLEAN,
            temp_media  NUMERIC(4,1),
            estacion    TEXT,
            cole        BOOLEAN,
            notas       TEXT,
            creado_en   TIMESTAMP DEFAULT NOW()
        )
    """)
    # Tabla de PREVISIONES: la previsión (media ponderada 70/30) que había para
    # cada día, congelada al hacer el Cierre Diario, para luego comparar con la
    # venta real y medir la efectividad.
    cur.execute("""
        CREATE TABLE IF NOT EXISTS previsiones (
            fecha           DATE PRIMARY KEY,
            valor_previsto  NUMERIC(10,2),
            media2          NUMERIC(10,2),
            tipo_dia        TEXT,
            clasificacion   TEXT,
            int_inferior    NUMERIC(10,2),
            int_superior    NUMERIC(10,2),
            desviacion      NUMERIC(10,2),
            generada_en     TIMESTAMP DEFAULT NOW()
        )
    """)
    # Tabla de arqueo de caja: dos tablas físicas independientes por día (apertura y
    # cierre), cada una con su propio desglose editable de billetes/monedas. Al cargar
    # un día nuevo, la apertura se autocompleta con los valores del cierre de ayer
    # (punto de partida), pero queda editable sin afectar al registro de ayer.
    cur.execute("""
        CREATE TABLE IF NOT EXISTS arqueos (
            id                CHAR(8) PRIMARY KEY,  -- YYYYMMDD, igual que cierres_caja
            fecha             DATE NOT NULL,

            -- APERTURA: desglose editable (normalmente copiado del cierre de ayer)
            ap_b50            INT DEFAULT 0,
            ap_b20            INT DEFAULT 0,
            ap_b10            INT DEFAULT 0,
            ap_b5             INT DEFAULT 0,
            ap_m2             INT DEFAULT 0,
            ap_m1             INT DEFAULT 0,
            ap_m050           INT DEFAULT 0,
            ap_m020           INT DEFAULT 0,
            ap_m010           INT DEFAULT 0,
            ap_m005           INT DEFAULT 0,
            ap_segunda_caja   NUMERIC(10,2) DEFAULT 0,
            ap_total          NUMERIC(10,2) DEFAULT 0,  -- calculado

            -- CIERRE: desglose editable, recuento real ANTES de retirar nada
            ci_b50            INT DEFAULT 0,
            ci_b20            INT DEFAULT 0,
            ci_b10            INT DEFAULT 0,
            ci_b5             INT DEFAULT 0,
            ci_m2             INT DEFAULT 0,
            ci_m1             INT DEFAULT 0,
            ci_m050           INT DEFAULT 0,
            ci_m020           INT DEFAULT 0,
            ci_m010           INT DEFAULT 0,
            ci_m005           INT DEFAULT 0,
            ci_segunda_caja   NUMERIC(10,2) DEFAULT 0,
            ci_total          NUMERIC(10,2) DEFAULT 0,  -- calculado

            -- DIFERENCIA DE CAJA = ci_total - ap_total (dinámica, baja al retirar billetes)
            diferencia_caja           NUMERIC(10,2) DEFAULT 0,
            -- Foto fija de la diferencia en el momento de pulsar "Fijar diferencia",
            -- ANTES de que el empleado empiece a retirar billetes del recuento
            diferencia_caja_inicial   NUMERIC(10,2) DEFAULT 0,

            -- TEÓRICO: Z + Ingreso Banco del cierre del día (autocompletado, editable)
            teorico_z         NUMERIC(10,2) DEFAULT 0,
            teorico_banco     NUMERIC(10,2) DEFAULT 0,

            -- REAL: lo que el empleado anota a mano al retirar el efectivo
            real_z            NUMERIC(10,2) DEFAULT 0,
            real_banco        NUMERIC(10,2) DEFAULT 0,

            notas             TEXT,
            -- FALSE = solo se ha guardado el conteo físico (teórico/real siguen
            -- siendo dinámicos y se releen del cierre diario). TRUE = caja cerrada
            -- definitivamente: todos los valores quedan congelados tal cual.
            cerrado           BOOLEAN DEFAULT FALSE,
            creado_en         TIMESTAMP DEFAULT NOW(),
            actualizado_en    TIMESTAMP DEFAULT NOW()
        )
    """)
    # Migración para bases de datos existentes
    cur.execute("""
        ALTER TABLE arqueos
        ADD COLUMN IF NOT EXISTS cerrado BOOLEAN DEFAULT FALSE
    """)
    conn.commit()
    cur.close()
    if cerrar_al_final:
        conn.close()

# ─────────────────────────────────────────────
#  FESTIVOS Y EVENTOS
# ─────────────────────────────────────────────

def get_tipo_dia(fecha: date) -> str:
    """Consulta la tabla festivos para saber el tipo de un día. 'normal' si no está."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT tipo FROM festivos WHERE fecha = %s", (fecha,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row[0] if row else "normal"

def get_tipo_y_nombre_dia(fecha: date):
    """Devuelve (tipo, nombre) del festivo/víspera/fm para una fecha, o ('normal', None)."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT tipo, nombre FROM festivos WHERE fecha = %s", (fecha,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if row:
        return row[0], row[1]
    return "normal", None

def save_festivo(fecha: date, tipo: str, nombre: str = ""):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO festivos (fecha, tipo, nombre)
        VALUES (%s, %s, %s)
        ON CONFLICT (fecha) DO UPDATE SET tipo = EXCLUDED.tipo, nombre = EXCLUDED.nombre
    """, (fecha, tipo, nombre))
    conn.commit()
    cur.close()
    conn.close()

def delete_festivo(fecha: date):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM festivos WHERE fecha = %s", (fecha,))
    conn.commit()
    cur.close()
    conn.close()

def load_festivos():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT fecha, tipo, nombre FROM festivos ORDER BY fecha")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows

def save_evento(fecha: date, tipo: str, descripcion: str):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO eventos (fecha, tipo, descripcion) VALUES (%s, %s, %s)
    """, (fecha, tipo, descripcion))
    conn.commit()
    cur.close()
    conn.close()

def delete_evento(evento_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM eventos WHERE id = %s", (evento_id,))
    conn.commit()
    cur.close()
    conn.close()

def load_eventos_fecha(fecha: date):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, tipo, descripcion FROM eventos WHERE fecha = %s", (fecha,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows

def load_eventos_rango(desde: date, hasta: date):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha, tipo, descripcion FROM eventos
        WHERE fecha BETWEEN %s AND %s ORDER BY fecha
    """, (desde, hasta))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows

# ─────────────────────────────────────────────
#  SCRAPER DE PARTIDOS DE FÚTBOL
# ─────────────────────────────────────────────
FUT_BASE = "https://www.futbolhoy.es/"
FUT_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; FHoyScraper/1.2)"}

FUT_TEAM_FILTER = ["barcelona", "espanyol", "España", "psv", "madrid", "psg", "paris saint-germain"]
FUT_COMP_FILTER = ["Champions League", "LaLiga", "Mundial 2026", "Liga EA Sports", "Supercopa de España"]
FUT_HORIZONTE_DIAS = 21
FUT_DELAY_S = 0.8

# ── football-data.org: token y clasificación de partidos por NIVEL 0-5 ──
# El análisis de ventas demostró que solo los partidos de NIVEL 4-5 entre semana
# mueven la facturación; el resto es ruido. El scraper captura solo nivel 4-5.
FD_API_TOKEN = os.getenv("FOOTBALL_DATA_API_TOKEN", "")

# ─────────────────────────────────────────────
#  CORREO DIARIO (resumen del cierre al email)
#  Reutiliza las MISMAS variables de entorno que tu script de backup:
#    EMAIL_USER = tucorreo@gmail.com
#    EMAIL_PASS = contraseña de aplicación de Gmail (16 caracteres)
#  Y el destinatario del resumen (puede ser distinto del de backup):
#    EMAIL_DESTINO_CIERRE = tucorreo@gmail.com  (si vacío, va a EMAIL_USER)
#  Si EMAIL_USER o EMAIL_PASS están vacíos, el envío se omite sin error.
# ─────────────────────────────────────────────
EMAIL_USER          = os.getenv("EMAIL_USER", "")
EMAIL_PASS          = os.getenv("EMAIL_PASS", "")
EMAIL_DESTINO_CIERRE = os.getenv("EMAIL_DESTINO_CIERRE", "tucorreo@gmail.com")
EMAIL_ACTIVADO      = bool(EMAIL_USER and EMAIL_PASS)


def enviar_correo_cierre(fecha: date) -> tuple:
    """Envía el resumen del cierre: venta, previsión, desvío, Z, banco y
    acumulado del mes. Devuelve (ok, mensaje)."""
    if not EMAIL_ACTIVADO:
        return False, "correo no configurado"
    import smtplib, ssl
    from email.message import EmailMessage
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""SELECT facturacion_total, z_caja, ingreso_banco
                       FROM cierres_caja WHERE fecha = %s""", (fecha,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return False, "sin cierre ese día"
        venta = float(row[0] or 0)
        zcaja = float(row[1] or 0)
        banco = float(row[2] or 0)

        cur.execute("SELECT valor_previsto FROM previsiones WHERE fecha = %s", (fecha,))
        rp = cur.fetchone()
        prev = float(rp[0]) if rp and rp[0] is not None else None

        primer_mes = fecha.replace(day=1)
        cur.execute("""SELECT COALESCE(SUM(facturacion_total),0)
                       FROM cierres_caja WHERE fecha BETWEEN %s AND %s""",
                    (primer_mes, fecha))
        acum_mes = float(cur.fetchone()[0] or 0)
        cur.close()
        conn.close()

        def es(n):
            return f"{n:,.2f}".replace(",", "·").replace(".", ",").replace("·", ".")

        if prev is not None and prev > 0:
            desvio = (venta - prev) / prev * 100
            signo = "+" if desvio >= 0 else ""
            linea_prev = (f"Previsto (base):  {es(prev)} €\n"
                          f"Desvío:          {signo}{es(desvio)} %\n")
        else:
            linea_prev = "Previsto (base):  — (sin previsión guardada)\n"

        f_txt = fecha.strftime("%d/%m/%Y")
        nombre_mes = fecha.strftime("%B").capitalize()
        cuerpo = (
            f"Cierre del {f_txt}\n"
            f"────────────────────────\n"
            f"Venta total:     {es(venta)} €\n"
            f"{linea_prev}"
            f"Z de caja:       {es(zcaja)} €\n"
            f"Dinero al banco: {es(banco)} €\n"
            f"────────────────────────\n"
            f"Acumulado {nombre_mes}: {es(acum_mes)} €\n"
        )

        msg = EmailMessage()
        msg["From"] = EMAIL_USER
        msg["To"] = EMAIL_DESTINO_CIERRE
        msg["Subject"] = f"Cierre Carpi {f_txt} · {es(venta)} €"
        msg.set_content(cuerpo)

        context = ssl.create_default_context()
        with smtplib.SMTP("smtp.gmail.com", 587, timeout=20) as srv:
            srv.ehlo()
            srv.starttls(context=context)
            srv.ehlo()
            srv.login(EMAIL_USER, EMAIL_PASS)
            srv.send_message(msg)
        return True, "enviado"
    except Exception as e:
        return False, str(e)

_FD_RIVALES_FUERTES = ["sevilla", "athletic", "betis", "valencia", "villarreal", "real sociedad"]
_FD_ATLETICO = ["atletico", "atlético"]

def _fd_norm(s):
    s = (s or "").lower()
    for a, b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ñ","n")]:
        s = s.replace(a, b)
    return s

def _fd_equipo_clave(nombre):
    n = _fd_norm(nombre)
    if "barcelona" in n: return "BARÇA"
    if "real madrid" in n: return "MADRID"
    if "spain" in n or "espana" in n: return "ESPAÑA"
    return None

def _nombre_fase(stage_up: str) -> str:
    """Traduce el stage de la API a un nombre legible en español."""
    if "FINAL" in stage_up and "SEMI" not in stage_up and "QUARTER" not in stage_up:
        return "Final"
    if "SEMI" in stage_up:
        return "Semifinales"
    if "QUARTER" in stage_up:
        return "Cuartos de final"
    if "LAST_16" in stage_up or "ROUND_OF_16" in stage_up:
        return "Octavos de final"
    return "Fase eliminatoria"

def _fd_nivel(local, visitante, comp, stage):
    """Asigna 0-5 según importancia (misma lógica que la recuperación histórica)."""
    el, ev = _fd_equipo_clave(local), _fd_equipo_clave(visitante)
    es_clasico = ({el, ev} == {"BARÇA", "MADRID"})
    stage = (stage or "").upper()
    final = "FINAL" in stage
    semi = "SEMI" in stage
    cuartos_oct = ("QUARTER" in stage or "LAST_16" in stage or "ROUND_OF_16" in stage)
    if "ESPAÑA" in (el, ev):
        if comp in ("Mundial", "Eurocopa"):
            if final or semi: return 5
            if cuartos_oct: return 4
            return 3
        return 3
    if es_clasico:
        return 5
    if comp == "Champions" and (el in ("BARÇA","MADRID") or ev in ("BARÇA","MADRID")):
        if final or semi: return 4
        return 3
    if el in ("BARÇA", "MADRID"):
        if any(k in _fd_norm(visitante) for k in _FD_ATLETICO): return 3
        if any(k in _fd_norm(visitante) for k in _FD_RIVALES_FUERTES): return 2
        return 1
    return 0

def _fut_text(el):
    return el.get_text(strip=True) if el else ""

def _fut_normalize(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join([c for c in s if not unicodedata.combining(c)])
    return s.lower().strip()

def _fut_any_keyword_in(s, keywords):
    s_norm = _fut_normalize(s)
    return any(_fut_normalize(kw) in s_norm for kw in keywords)

def _fut_comp_match_exact(comp, keywords_norm):
    comp_norm = _fut_normalize(comp)
    return any(comp_norm == _fut_normalize(kw) for kw in keywords_norm)

def _fut_normalize_channel(ch):
    ch = (ch or "").strip()
    if ch.startswith("M+ Liga de Campeones"):
        parts = ch.split()
        num = parts[-1] if parts[-1].isdigit() else ""
        base = "Movistar Liga de Campeones"
        return (base + (" " + num if num else "")).strip()
    return ch

def _fut_parse_match_card(card, fecha_ymd, competicion):
    fase = _fut_text(card.select_one(".m_phase span"))
    local = _fut_text(card.select_one(".m_title span:nth-of-type(1)"))
    visitante = _fut_text(card.select_one(".m_title span:nth-of-type(2)"))
    hora = _fut_text(card.select_one(".m_logos .m_time"))
    canales = [_fut_normalize_channel(_fut_text(ch)) for ch in card.select(".m_chan .channelLnk")]
    return {
        "fecha": fecha_ymd, "competicion": competicion, "fase": fase,
        "local": local, "visitante": visitante, "hora": hora,
        "canales": " | ".join(canales)
    }

def _fut_parse_list_page(url, fecha_ymd, team_kw_norm, comp_kw_norm):
    r = requests.get(url, headers=FUT_HEADERS, timeout=20)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    rows = []
    comp_headers = soup.select(".matchday .matchdayCompetitionHeader h3")

    for comp_h3 in comp_headers:
        competicion = _fut_text(comp_h3)
        comp_ok = _fut_comp_match_exact(competicion, comp_kw_norm) if comp_kw_norm else True
        if not comp_ok:
            continue

        cont = comp_h3.find_parent().find_next_sibling()

        def has_row_eq_height(el):
            if not el or not el.has_attr("class"):
                return False
            classes = el.get("class", [])
            return ("row" in classes) and ("row-eq-height" in classes)

        while cont and not has_row_eq_height(cont):
            cont = cont.find_next_sibling()
        if not cont:
            continue

        for col in cont.select(".fh-cardcol .match"):
            row = _fut_parse_match_card(col, fecha_ymd, competicion)
            if not (row["local"] and row["visitante"]):
                continue
            team_ok = (_fut_any_keyword_in(row["local"], team_kw_norm) or
                       _fut_any_keyword_in(row["visitante"], team_kw_norm)) if team_kw_norm else True
            if team_ok:
                rows.append(row)
    return rows

def _fut_rango_fechas(desde_ymd, hasta_ymd):
    d0 = datetime.strptime(desde_ymd, "%Y%m%d")
    d1 = datetime.strptime(hasta_ymd, "%Y%m%d")
    cur = d0
    while cur <= d1:
        yield cur.strftime("%Y%m%d")
        cur += timedelta(days=1)

def scrape_partidos_futbol(progress_callback=None):
    """
    Obtiene los próximos partidos relevantes desde football-data.org (misma
    fuente que la recuperación histórica) y guarda en `eventos` TODOS los de
    nivel 1-5 (de Barça/Madrid/España), cada uno etiquetado con su nivel. Así se
    muestran como alerta y se acumulan datos para análisis futuro. El AJUSTE de
    previsión se aplica solo a nivel 4-5 entre semana (eso se decide en la
    previsión, no aquí). El nivel va en la descripción como '... | nivel=N'.

    Mantiene el nombre y la firma para no romper el cierre diario.
    Devuelve la lista de partidos guardados.
    """
    import time as _t
    horizonte = FUT_HORIZONTE_DIAS
    hoy = date.today()
    hasta = hoy + timedelta(days=horizonte)

    def _cb(msg):
        if progress_callback:
            progress_callback(msg)

    headers = {"X-Auth-Token": FD_API_TOKEN}
    encontrados = []

    # Competiciones a consultar (las del plan gratuito que importan)
    comps = {"PD": "LaLiga", "CL": "Champions", "WC": "Mundial", "EC": "Eurocopa"}
    for code, nombre in comps.items():
        _cb(f"Consultando {nombre}…")
        try:
            r = requests.get(
                f"https://api.football-data.org/v4/competitions/{code}/matches",
                headers=headers,
                params={"dateFrom": hoy.isoformat(), "dateTo": hasta.isoformat()},
                timeout=30)
            if r.status_code == 429:
                _t.sleep(int(r.headers.get("X-RequestCounter-Reset", "60")) + 1)
                r = requests.get(
                    f"https://api.football-data.org/v4/competitions/{code}/matches",
                    headers=headers,
                    params={"dateFrom": hoy.isoformat(), "dateTo": hasta.isoformat()},
                    timeout=30)
            r.raise_for_status()
            data = r.json()
        except Exception:
            continue  # sin internet o competición no disponible: se omite

        for m in data.get("matches", []):
            local = m.get("homeTeam", {}).get("name", "") or ""
            visit = m.get("awayTeam", {}).get("name", "") or ""
            utc = m.get("utcDate") or ""
            fecha_iso = utc[:10]
            stage = m.get("stage", "")
            if not fecha_iso:
                continue
            f = date.fromisoformat(fecha_iso)

            # Hora: la API la da en UTC. La pasamos a hora española (CET/CEST)
            # sumando 2h en horario de verano (abr-oct) o 1h en invierno. Es una
            # aproximación suficiente para clasificar la franja de cena.
            hora = utc[11:16]
            hora_es = hora
            try:
                hh, mm = int(utc[11:13]), int(utc[14:16])
                offset = 2 if 3 <= f.month <= 10 else 1  # CEST verano / CET invierno
                hh_es = (hh + offset) % 24
                hora_es = f"{hh_es:02d}:{mm:02d}"
            except Exception:
                hora_es = hora or "NA"

            # FILTRO DE HORA (común): solo cuenta lo que empieza entre 17:00 y
            # 23:00 hora española. Fuera de esa franja no afecta a la cena.
            try:
                hh_es2, mm_es2 = int(hora_es.split(":")[0]), int(hora_es.split(":")[1])
                minutos = hh_es2 * 60 + mm_es2
                en_franja = (17 * 60 <= minutos <= 23 * 60)
            except Exception:
                en_franja = False
            if not en_franja:
                continue  # fuera de franja: ignorar siempre

            # ── CASO PROVISIONAL: fase importante SIN equipos definidos ──
            # (octavos, cuartos, semis, final aún por decidir). Se guarda como
            # ALERTA informativa "por definir", sin nivel y sin ajuste. En el
            # siguiente cierre, si ya hay equipos: o se vuelve partido real con
            # nivel, o (si no es importante) desaparece. nivel=0 lo marca.
            stage_up = (stage or "").upper()
            fase_imp = ("FINAL" in stage_up or "SEMI" in stage_up
                        or "QUARTER" in stage_up or "LAST_16" in stage_up
                        or "ROUND_OF_16" in stage_up)
            if (not local or not visit):
                if fase_imp:
                    fase_txt = _nombre_fase(stage_up)
                    desc = f"Por definir | {nombre} | {stage or 'NA'} | {hora_es} | nivel=0"
                    encontrados.append((f, desc, 0))
                continue  # sin equipos: provisional (si fase imp) o nada

            el = _fd_equipo_clave(local); ev = _fd_equipo_clave(visit)
            # liga: solo Barça/Madrid en casa; Champions/selección: cuenta igual
            if nombre == "LaLiga" and el not in ("BARÇA", "MADRID"):
                continue
            if nombre in ("Champions", "Mundial", "Eurocopa") and not (el or ev):
                continue
            nivel = _fd_nivel(local, visit, nombre, stage)
            if nivel < 1:
                continue  # nivel 0 = sin equipo clave; el resto (1-5) se guarda

            desc = f"{local} - {visit} | {nombre} | {stage or 'NA'} | {hora_es} | nivel={nivel}"
            encontrados.append((f, desc, nivel))
        _t.sleep(6)  # respetar el límite de la API gratuita

    # Guardar en eventos. Estrategia: para cada FECHA que hemos re-escaneado,
    # borramos los partidos de fútbol previos y reinsertamos los actuales. Así:
    #  - un provisional "Por definir" que ya tiene equipos → se reemplaza por el real
    #  - un provisional que sigue sin ser importante → desaparece (no se reinserta)
    #  - los partidos reales se mantienen actualizados (hora, nivel)
    # Solo tocamos fechas dentro del rango escaneado, nunca el histórico anterior.
    conn = get_connection()
    cur = conn.cursor()
    fechas_escaneadas = {f for f, _, _ in encontrados}
    # también limpiamos fechas del rango aunque ya no tengan partido (un provisional
    # que desaparece del todo): borramos todo fútbol entre hoy y 'hasta'.
    cur.execute("""DELETE FROM eventos
                   WHERE tipo='futbol' AND fecha BETWEEN %s AND %s""",
                (hoy, hasta))
    guardados = 0
    for f, desc, nivel in encontrados:
        cur.execute("INSERT INTO eventos (fecha, tipo, descripcion) VALUES (%s,'futbol',%s)",
                    (f, desc))
        guardados += 1
    conn.commit()
    cur.close()
    conn.close()

    _cb(f"Listo: {len(encontrados)} partidos relevantes, {guardados} nuevos guardados.")
    return encontrados

# ─────────────────────────────────────────────
#  PREVISIÓN
# ─────────────────────────────────────────────

def prever_dia_normal(fecha: date) -> dict:
    """
    Media ponderada de facturación (70% año pasado, 30% hace 2 años)
    para el mismo mes + mismo día de semana.
    Devuelve dict con valor y nº de muestras usadas por año.
    """
    conn = get_connection()
    cur = conn.cursor()
    mes = fecha.month
    dow_iso = fecha.isoweekday()  # 1=lunes..7=domingo (igual que DIASEM modo 2)
    anio_actual = fecha.year

    resultado = {}
    for offset, peso in [(1, 0.7), (2, 0.3)]:
        anio_ref = anio_actual - offset
        cur.execute("""
            SELECT AVG(facturacion_total), COUNT(*)
            FROM cierres_caja
            WHERE EXTRACT(MONTH FROM fecha) = %s
              AND EXTRACT(ISODOW FROM fecha) = %s
              AND EXTRACT(YEAR FROM fecha) = %s
              AND tipo_dia = 'normal'
        """, (mes, dow_iso, anio_ref))
        media, n = cur.fetchone()
        resultado[anio_ref] = {"media": float(media) if media else None, "n": n, "peso": peso}

    cur.close()
    conn.close()

    total_peso = sum(r["peso"] for r in resultado.values() if r["media"] is not None)
    if total_peso == 0:
        return {"valor": None, "detalle": resultado}

    valor = sum(r["media"] * r["peso"] for r in resultado.values() if r["media"] is not None) / total_peso
    return {"valor": valor, "detalle": resultado}

def prever_dia_normal_media2(fecha: date) -> float:
    """
    Media 2: promedia los dos días más cercanos del año pasado con el mismo
    día de la semana — uno justo antes y otro justo después de la fecha
    equivalente del año pasado.
    Equivale a la fórmula INDICE/COINCIDIR/MAX/MIN de Excel.
    """
    conn = get_connection()
    cur = conn.cursor()
    dow_iso = fecha.isoweekday()
    fecha_ref = fecha.replace(year=fecha.year - 1)

    # Día más cercano ANTES (o igual) de fecha_ref, mismo día de semana, año pasado
    cur.execute("""
        SELECT facturacion_total FROM cierres_caja
        WHERE EXTRACT(YEAR FROM fecha) = %s
          AND EXTRACT(ISODOW FROM fecha) = %s
          AND fecha <= %s
        ORDER BY fecha DESC
        LIMIT 1
    """, (fecha_ref.year, dow_iso, fecha_ref))
    row_antes = cur.fetchone()

    # Día más cercano DESPUÉS (o igual) de fecha_ref, mismo día de semana, año pasado
    cur.execute("""
        SELECT facturacion_total FROM cierres_caja
        WHERE EXTRACT(YEAR FROM fecha) = %s
          AND EXTRACT(ISODOW FROM fecha) = %s
          AND fecha >= %s
        ORDER BY fecha ASC
        LIMIT 1
    """, (fecha_ref.year, dow_iso, fecha_ref))
    row_despues = cur.fetchone()

    cur.close()
    conn.close()

    if not row_antes or not row_despues:
        return None
    return (float(row_antes[0]) + float(row_despues[0])) / 2

def calcular_desviacion(fecha: date) -> float:
    """
    Desviación estándar de TODOS los años para el mismo mes + día de semana.
    Equivale a: DESVEST(FILTRAR(K; (MES=mes)*(DIASEM=dow)))
    """
    import statistics
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT facturacion_total FROM cierres_caja
        WHERE EXTRACT(MONTH FROM fecha) = %s
          AND EXTRACT(ISODOW FROM fecha) = %s
    """, (fecha.month, fecha.isoweekday()))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    valores = [float(r[0]) for r in rows]
    if len(valores) < 2:
        return None
    return statistics.stdev(valores)  # DESVEST de Excel = muestral (n-1)

def calcular_percentil(valor: float) -> float:
    """
    Percentil del valor dentro de TODA la distribución histórica de facturación.
    Equivale a: CONTAR.SI(K; "<="&valor) / CONTARA(K)
    """
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM cierres_caja WHERE facturacion_total <= %s", (valor,))
    n_menor_igual = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM cierres_caja")
    n_total = cur.fetchone()[0]
    cur.close()
    conn.close()
    if n_total == 0:
        return None
    return n_menor_igual / n_total

def clasificar_dia(valor: float, percentil: float) -> str:
    """
    Equivale a:
    SI(valor > MAX histórico; "Record";
       SI(percentil <= 0.3; "Flojo";
          SI(percentil <= 0.9; "Normal"; "Potente")))
    """
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT MAX(facturacion_total) FROM cierres_caja")
    maximo = cur.fetchone()[0]
    cur.close()
    conn.close()

    if maximo is not None and valor > float(maximo):
        return "Record"
    if percentil <= 0.3:
        return "Flojo"
    if percentil <= 0.9:
        return "Normal"
    return "Potente"

def _enriquecer_con_estadisticas(r: dict) -> dict:
    """
    Añade intervalo (±5% de la media), percentil y clasificación a un resultado
    de previsión basado en una lista de muestras (festivo/vispera/fm).
    Equivale a: int_inf = media*0.95, int_sup = media*1.05
    """
    if r["valor"] is None:
        r["intervalo_inferior"] = None
        r["intervalo_superior"] = None
        r["percentil"] = None
        r["clasificacion"] = None
        return r

    r["intervalo_inferior"] = r["valor"] * 0.95
    r["intervalo_superior"] = r["valor"] * 1.05
    r["percentil"] = calcular_percentil(r["valor"])
    r["clasificacion"] = clasificar_dia(r["valor"], r["percentil"]) if r["percentil"] is not None else None
    return r

def prever_dia_festivo_vispera(fecha: date) -> dict:
    """
    Media de facturación de la misma fecha exacta (dd/mm) en los últimos 2 años,
    para festivos y vísperas (eventos que se repiten en fecha fija).
    """
    conn = get_connection()
    cur = conn.cursor()
    anio_actual = fecha.year
    cur.execute("""
        SELECT fecha, facturacion_total
        FROM cierres_caja
        WHERE EXTRACT(MONTH FROM fecha) = %s
          AND EXTRACT(DAY FROM fecha) = %s
          AND EXTRACT(YEAR FROM fecha) >= %s
          AND EXTRACT(YEAR FROM fecha) < %s
    """, (fecha.month, fecha.day, anio_actual - 2, anio_actual))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    if not rows:
        return {"valor": None, "muestras": []}
    valores = [float(r[1]) for r in rows]
    resultado = {"valor": sum(valores) / len(valores), "muestras": rows}
    return _enriquecer_con_estadisticas(resultado)

def prever_dia_fm(fecha: date) -> dict:
    """Media de todos los días históricos marcados como Fiesta Mayor."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT fecha, facturacion_total FROM cierres_caja WHERE tipo_dia = 'fm'
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    if not rows:
        return {"valor": None, "muestras": []}
    valores = [float(r[1]) for r in rows]
    resultado = {"valor": sum(valores) / len(valores), "muestras": rows}
    return _enriquecer_con_estadisticas(resultado)

# Factor de ajuste por partido GORDO (nivel 4-5) en día ENTRE SEMANA.
# El análisis de ventas demostró ~+58% en esos días; usamos +55% fijo.
# (En finde no se aplica: el finde ya vende solo y el efecto se diluye.)
AJUSTE_FUTBOL_TOCHO = 1.55

def _nivel_futbol_dia(fecha: date) -> int:
    """Mayor nivel de partido (1-5) registrado en eventos ese día, o 0 si no hay.
    Lee la etiqueta 'nivel=N' de la descripción. (La hora ya la filtra el
    scraper al guardar, así que aquí no hace falta volver a comprobarla.)"""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""SELECT descripcion FROM eventos
                       WHERE fecha=%s AND tipo='futbol' AND descripcion LIKE %s""",
                    (fecha, '%nivel=%'))
        rows = cur.fetchall()
        cur.close()
        conn.close()
        niveles = []
        for (desc,) in rows:
            try:
                niveles.append(int(desc.split("nivel=")[1].split()[0].strip()))
            except Exception:
                pass
        return max(niveles) if niveles else 0
    except Exception:
        return 0

def prever_facturacion(fecha: date) -> dict:
    """Determina el tipo de día y aplica la fórmula correspondiente."""
    tipo, nombre = get_tipo_y_nombre_dia(fecha)
    if tipo == "normal":
        r = prever_dia_normal(fecha)
        r["media2"] = prever_dia_normal_media2(fecha)
        if r["valor"] is not None:
            desv = calcular_desviacion(fecha)
            r["desviacion"] = desv
            if desv is not None:
                r["intervalo_inferior"] = r["valor"] - desv * 0.5
                r["intervalo_superior"] = r["valor"] + desv * 0.5
                r["porc_desviacion"] = desv / r["valor"] if r["valor"] else None
            else:
                r["intervalo_inferior"] = None
                r["intervalo_superior"] = None
                r["porc_desviacion"] = None
            r["percentil"] = calcular_percentil(r["valor"])
            r["clasificacion"] = clasificar_dia(r["valor"], r["percentil"]) if r["percentil"] is not None else None
    elif tipo in ("festivo", "vispera"):
        r = prever_dia_festivo_vispera(fecha)
    elif tipo == "fm":
        r = prever_dia_fm(fecha)
    else:
        r = {"valor": None}
    r["tipo_dia"] = tipo
    r["nombre_dia"] = nombre
    r["eventos"] = load_eventos_fecha(fecha)

    # AJUSTE por partido GORDO (nivel 4-5) en día ENTRE SEMANA.
    # No modifica r["valor"] (la previsión base se conserva siempre). Añade
    # campos extra para que la app pueda mostrar AMBAS cifras: la normal y la
    # ajustada. Solo se calcula en días normales con previsión válida.
    # (La hora ya la filtra el scraper: solo guarda partidos de 17:00 a 23:00.)
    r["futbol_nivel"] = _nivel_futbol_dia(fecha)
    r["futbol_ajuste"] = False
    r["valor_ajustado"] = None
    es_finde = fecha.isoweekday() in (5, 6, 7)  # vie, sáb, dom
    if (tipo == "normal" and r.get("valor") is not None
            and r["futbol_nivel"] >= 4 and not es_finde):
        r["futbol_ajuste"] = True
        r["valor_ajustado"] = round(r["valor"] * AJUSTE_FUTBOL_TOCHO, 2)

    return r

# ─────────────────────────────────────────────
#  CONTEXTO DEL DÍA (estación, cole, lluvia) — se rellena en el Cierre Diario
# ─────────────────────────────────────────────
# Coordenadas del local: Montornès del Vallès (08170)
_CTX_LAT = 41.542
_CTX_LON = 2.267
_CTX_UMBRAL_LLUVIA_MM = 1.0

def _estacion_de(f: date) -> str:
    m = f.month
    if m in (3, 4, 5):   return "primavera"
    if m in (6, 7, 8):   return "verano"
    if m in (9, 10, 11): return "otoño"
    return "invierno"

def _domingo_de_pascua(anio: int) -> date:
    """Algoritmo de Gauss/Meeus para la fecha de Pascua (domingo)."""
    a = anio % 19; b = anio // 100; c = anio % 100
    d = b // 4; e = b % 4; f = (b + 8) // 25; g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4; k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes = (h + l - 7 * m + 114) // 31
    dia = ((h + l - 7 * m + 114) % 31) + 1
    return date(anio, mes, dia)

def _es_vacaciones_escolares(f: date) -> bool:
    """Solo vacaciones largas: Navidad (22 dic→7 ene), Semana Santa, verano (22 jun→11 sep)."""
    y = f.year
    if f >= date(y, 12, 22) or f <= date(y, 1, 7):
        return True
    if date(y, 6, 22) <= f <= date(y, 9, 11):
        return True
    viernes_santo = _domingo_de_pascua(y) - timedelta(days=2)
    if (viernes_santo - timedelta(days=4)) <= f <= (viernes_santo + timedelta(days=3)):
        return True
    return False

def _descargar_lluvia(fecha: date):
    """Devuelve (lluvia_mm, temp_media) de Open-Meteo para esa fecha, o (None, None) si falla."""
    try:
        url = "https://archive-api.open-meteo.com/v1/archive"
        params = {
            "latitude": _CTX_LAT, "longitude": _CTX_LON,
            "start_date": fecha.isoformat(), "end_date": fecha.isoformat(),
            "daily": "precipitation_sum,temperature_2m_mean",
            "timezone": "Europe/Madrid",
        }
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        d = resp.json()["daily"]
        mm = d["precipitation_sum"][0]
        t = d["temperature_2m_mean"][0]
        return (round(mm, 1) if mm is not None else 0.0,
                round(t, 1) if t is not None else None)
    except Exception:
        return (None, None)

def guardar_contexto_dia(fecha: date) -> bool:
    """
    Calcula y guarda el contexto del día (estación, cole, lluvia) en contexto_dia.
    La lluvia se descarga de Open-Meteo; si no hay internet, se guarda el resto
    igual (lluvia queda NULL y se puede rellenar después). Nunca lanza excepción.
    """
    try:
        estacion = _estacion_de(fecha)
        cole = not _es_vacaciones_escolares(fecha)
        lluvia_mm, temp = _descargar_lluvia(fecha)
        llueve = (lluvia_mm >= _CTX_UMBRAL_LLUVIA_MM) if lluvia_mm is not None else None

        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO contexto_dia (fecha, lluvia_mm, llueve, temp_media, estacion, cole)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (fecha) DO UPDATE SET
                lluvia_mm  = EXCLUDED.lluvia_mm,
                llueve     = EXCLUDED.llueve,
                temp_media = EXCLUDED.temp_media,
                estacion   = EXCLUDED.estacion,
                cole       = EXCLUDED.cole
        """, (fecha, lluvia_mm, llueve, temp, estacion, cole))
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception:
        return False

def registrar_prevision_dia(fecha: date) -> bool:
    """
    Congela en la tabla previsiones la previsión (media ponderada 70/30) que
    había para `fecha`, calculada con datos de años anteriores. Nunca lanza.
    """
    try:
        r = prever_facturacion(fecha)
        if r.get("valor") is None:
            return False
        def _r(x): return round(float(x), 2) if x is not None else None
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO previsiones
                (fecha, valor_previsto, media2, tipo_dia, clasificacion,
                 int_inferior, int_superior, desviacion)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (fecha) DO UPDATE SET
                valor_previsto = EXCLUDED.valor_previsto,
                media2         = EXCLUDED.media2,
                tipo_dia       = EXCLUDED.tipo_dia,
                clasificacion  = EXCLUDED.clasificacion,
                int_inferior   = EXCLUDED.int_inferior,
                int_superior   = EXCLUDED.int_superior,
                desviacion     = EXCLUDED.desviacion,
                generada_en    = NOW()
        """, (fecha, _r(r.get("valor")), _r(r.get("media2")), r.get("tipo_dia"),
              r.get("clasificacion"), _r(r.get("intervalo_inferior")),
              _r(r.get("intervalo_superior")), _r(r.get("desviacion"))))
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception:
        return False

# ─────────────────────────────────────────────
#  PASO 2 — FACTOR DE CORRECCIÓN RECIENTE (últimas 8 semanas)
# ─────────────────────────────────────────────
def calcular_factor_correccion(semanas: int = 8) -> dict:
    """
    Compara previsión vs venta real en las últimas `semanas` semanas para detectar
    el nivel ACTUAL del negocio. Si últimamente vendes por debajo de lo previsto
    (p.ej. por la bajada general de ventas), el factor será < 1 y sirve para
    ajustar las previsiones futuras al nivel real de hoy.

    Devuelve dict: {factor, n, real, previsto, desvio_pct}. factor=1.0 si no hay
    datos suficientes (no corregir).
    """
    try:
        desde = date.today() - timedelta(weeks=semanas)
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT c.facturacion_total, p.valor_previsto
            FROM cierres_caja c
            JOIN previsiones p ON p.fecha = c.fecha
            WHERE c.fecha >= %s
              AND p.valor_previsto IS NOT NULL
              AND c.facturacion_total IS NOT NULL
        """, (desde,))
        filas = cur.fetchall()
        cur.close()
        conn.close()

        if len(filas) < 5:
            return {"factor": 1.0, "n": len(filas), "real": 0.0,
                    "previsto": 0.0, "desvio_pct": 0.0}

        suma_real = sum(float(r) for r, _ in filas)
        suma_prev = sum(float(p) for _, p in filas)
        factor = suma_real / suma_prev if suma_prev else 1.0
        desvio = 100 * (factor - 1.0)
        return {"factor": factor, "n": len(filas), "real": suma_real,
                "previsto": suma_prev, "desvio_pct": desvio}
    except Exception:
        return {"factor": 1.0, "n": 0, "real": 0.0, "previsto": 0.0, "desvio_pct": 0.0}


def datos_mes_acumulado(anio: int, mes: int, factor: float = 1.0,
                        completar_al_vuelo: bool = True) -> dict:
    """
    Para un mes, devuelve listas acumuladas día a día:
      dias, prev_acum (toda la previsión del mes), prev_corr_acum (× factor),
      real_acum (solo hasta hoy/último cierre).
    El resumen compara real vs previsión corregida HASTA EL MISMO DÍA (justo).
    """
    import calendar as _cal
    n_dias = _cal.monthrange(anio, mes)[1]
    primer = date(anio, mes, 1)
    ultimo = date(anio, mes, n_dias)

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""SELECT fecha, valor_previsto FROM previsiones
                   WHERE fecha BETWEEN %s AND %s""", (primer, ultimo))
    prev = {f: (float(v) if v is not None else None) for f, v in cur.fetchall()}
    cur.execute("""SELECT fecha, facturacion_total FROM cierres_caja
                   WHERE fecha BETWEEN %s AND %s""", (primer, ultimo))
    real = {f: (float(v) if v is not None else None) for f, v in cur.fetchall()}
    cur.close()
    conn.close()

    # Completar la previsión de CUALQUIER día del mes que no la tenga en la
    # tabla (días pasados sin backfill o días futuros), calculándola al vuelo,
    # para que la línea de previsión no se aplane. Esto se hace en la vista
    # mensual (máx. 31 días, asumible). En la vista anual se desactiva con
    # completar_al_vuelo=False, porque 12×31 cálculos serían lentísimos.
    if completar_al_vuelo:
        for d in range(1, n_dias + 1):
            f = date(anio, mes, d)
            if f not in prev or prev[f] is None:
                try:
                    r = prever_facturacion(f)
                    prev[f] = r.get("valor")
                except Exception:
                    prev[f] = None

    hoy = date.today()
    dias, prev_acum, prev_corr_acum, real_acum = [], [], [], []
    ap = acp = ar = 0.0
    hay_prev = len(prev) > 0
    for d in range(1, n_dias + 1):
        f = date(anio, mes, d)
        dias.append(f)
        pv = prev.get(f)
        if pv is not None:
            ap += pv
            acp += pv * factor
        prev_acum.append(ap if hay_prev else None)
        prev_corr_acum.append(acp if hay_prev else None)
        # REAL: solo se acumula y dibuja en días que TIENEN cierre registrado.
        # Un día sin cierre (futuro, o pasado pero aún no cerrado) → None, para
        # que la línea naranja se CORTE ahí en vez de quedarse plana.
        rv = real.get(f)
        if rv is not None:
            ar += rv
            real_acum.append(ar)
        else:
            real_acum.append(None)

    # Resumen JUSTO: comparar real vs previsión corregida HASTA EL MISMO DÍA.
    idx = [i for i, v in enumerate(real_acum) if v is not None]
    resumen = {"prev_hoy": None, "prev_corr_hoy": None, "real_hoy": None}
    if idx:
        i = idx[-1]  # último día con real (=hoy o último cierre)
        resumen = {"prev_hoy": prev_acum[i],
                   "prev_corr_hoy": prev_corr_acum[i],   # previsión hasta ESE día
                   "real_hoy": real_acum[i]}
    return {"dias": dias, "prev_acum": prev_acum, "prev_corr_acum": prev_corr_acum,
            "real_acum": real_acum, "resumen": resumen}


def datos_calendario_mes(anio: int, mes: int) -> dict:
    """
    Facturación real por día de un mes, organizada como calendario.
    Devuelve {semanas: [[(dia, valor|None), ...x7], ...], min, max, total}
    donde cada semana es una fila de 7 (lunes→domingo); None = día fuera del mes.
    """
    import calendar as _cal
    n_dias = _cal.monthrange(anio, mes)[1]
    primer = date(anio, mes, 1)
    ultimo = date(anio, mes, n_dias)

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""SELECT fecha, facturacion_total FROM cierres_caja
                   WHERE fecha BETWEEN %s AND %s""", (primer, ultimo))
    real = {f: (float(v) if v is not None else None) for f, v in cur.fetchall()}
    cur.close()
    conn.close()

    valores = [v for v in real.values() if v is not None]
    # Construir matriz de semanas (lunes=0 ... domingo=6)
    cal = _cal.Calendar(firstweekday=0)  # lunes
    semanas = []
    for semana in cal.monthdayscalendar(anio, mes):  # listas de día (0 = fuera)
        fila = []
        for d in semana:
            if d == 0:
                fila.append((None, None))
            else:
                f = date(anio, mes, d)
                fila.append((d, real.get(f)))
        semanas.append(fila)

    return {"semanas": semanas,
            "min": min(valores) if valores else 0.0,
            "max": max(valores) if valores else 0.0,
            "total": sum(valores) if valores else 0.0}


def datos_heatmap_mensual() -> dict:
    """
    Devuelve la facturación REAL total por (año, mes) para pintar un mapa de calor.
    {anios: [...], matriz: {anio: {mes: total}}, min, max}
    """
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT EXTRACT(YEAR FROM fecha)::int  AS anio,
               EXTRACT(MONTH FROM fecha)::int AS mes,
               SUM(facturacion_total)         AS total
        FROM cierres_caja
        GROUP BY anio, mes
        ORDER BY anio, mes
    """)
    filas = cur.fetchall()
    cur.close()
    conn.close()

    matriz = {}
    valores = []
    anios = set()
    for anio, mes, total in filas:
        t = float(total) if total is not None else 0.0
        matriz.setdefault(anio, {})[mes] = t
        valores.append(t)
        anios.add(anio)
    return {"anios": sorted(anios), "matriz": matriz,
            "min": min(valores) if valores else 0.0,
            "max": max(valores) if valores else 0.0}


def datos_calendario_mes(anio: int, mes: int) -> dict:
    """
    Facturación real por día organizada como CALENDARIO (filas = semanas,
    columnas = lunes..domingo), para pintarlo como un calendario coloreado.
    Devuelve {semanas: [[(dia, valor|None), ...7], ...], min, max, total}.
    Los huecos antes/después del mes son (None, None).
    """
    import calendar as _cal
    n_dias = _cal.monthrange(anio, mes)[1]
    primer = date(anio, mes, 1)
    ultimo = date(anio, mes, n_dias)
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""SELECT fecha, facturacion_total FROM cierres_caja
                   WHERE fecha BETWEEN %s AND %s""", (primer, ultimo))
    filas = cur.fetchall()
    cur.close()
    conn.close()
    por_dia = {f.day: float(v) for f, v in filas if v is not None}
    valores = list(por_dia.values())

    # Construir la rejilla del calendario (semanas empezando en lunes)
    cal = _cal.Calendar(firstweekday=0)  # 0 = lunes
    semanas = []
    for semana in cal.monthdayscalendar(anio, mes):  # 0 = día fuera del mes
        fila = []
        for dia in semana:
            if dia == 0:
                fila.append((None, None))
            else:
                fila.append((dia, por_dia.get(dia)))
        semanas.append(fila)

    return {"semanas": semanas,
            "min": min(valores) if valores else 0.0,
            "max": max(valores) if valores else 0.0,
            "total": sum(valores) if valores else 0.0}


def save_cierre(data: dict) -> bool:
    """Inserta o actualiza el cierre del día."""
    conn = get_connection()
    cur = conn.cursor()
    if "tipo_dia" not in data:
        data["tipo_dia"] = "normal"
    cur.execute("""
        INSERT INTO cierres_caja
            (id, fecha, facturacion, gastos, visa, internet,
             justeat, glovo, uber, ticket_restaurant,
             ingreso_banco, z_caja, facturacion_total, notas, tipo_dia, actualizado_en)
        VALUES
            (%(id)s, %(fecha)s, %(facturacion)s, %(gastos)s, %(visa)s, %(internet)s,
             %(justeat)s, %(glovo)s, %(uber)s, %(ticket_restaurant)s,
             %(ingreso_banco)s, %(z_caja)s, %(facturacion_total)s, %(notas)s, %(tipo_dia)s, NOW())
        ON CONFLICT (id) DO UPDATE SET
            facturacion       = EXCLUDED.facturacion,
            gastos            = EXCLUDED.gastos,
            visa              = EXCLUDED.visa,
            internet          = EXCLUDED.internet,
            justeat           = EXCLUDED.justeat,
            glovo             = EXCLUDED.glovo,
            uber              = EXCLUDED.uber,
            ticket_restaurant = EXCLUDED.ticket_restaurant,
            ingreso_banco     = EXCLUDED.ingreso_banco,
            z_caja            = EXCLUDED.z_caja,
            facturacion_total = EXCLUDED.facturacion_total,
            notas             = EXCLUDED.notas,
            tipo_dia          = EXCLUDED.tipo_dia,
            actualizado_en    = NOW()
    """, data)
    conn.commit()
    cur.close()
    conn.close()
    return True

def load_cierre(fecha: date):
    """Carga el cierre de una fecha dada, o None si no existe."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT facturacion, z_caja, facturacion_total,
               visa, internet, justeat, glovo, uber,
               ticket_restaurant, gastos, ingreso_banco, notas, tipo_dia
        FROM cierres_caja WHERE fecha = %s
    """, (fecha,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row

def get_teorico_efectivo(fecha: date):
    """
    Devuelve (z_caja, ingreso_banco, suma) del cierre del día, para comparar
    contra la diferencia real de caja en el arqueo. None si no hay cierre.
    """
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT z_caja, ingreso_banco FROM cierres_caja WHERE fecha = %s
    """, (fecha,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return None
    z = float(row[0] or 0)
    banco = float(row[1] or 0)
    return z, banco, z + banco

def get_efectivo_tpv(fecha: date):
    """
    Suma del efectivo cobrado por el TPV (tabla pagaments) para una fecha dada.
    Devuelve None si no hay ningún pago en efectivo ese día.
    """
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT SUM(import) FROM pagaments
        WHERE metode = 'efectiu' AND data::date = %s
    """, (fecha,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return float(row[0]) if row and row[0] is not None else None

# ─────────────────────────────────────────────
#  ARQUEO DE CAJA
# ─────────────────────────────────────────────
DENOMINACIONES = [
    ("b50",  50.00,  "Billete 50€"),
    ("b20",  20.00,  "Billete 20€"),
    ("b10",  10.00,  "Billete 10€"),
    ("b5",    5.00,  "Billete 5€"),
    ("m2",    2.00,  "Moneda 2€"),
    ("m1",    1.00,  "Moneda 1€"),
    ("m050",  0.50,  "Moneda 0.50€"),
    ("m020",  0.20,  "Moneda 0.20€"),
    ("m010",  0.10,  "Moneda 0.10€"),
    ("m005",  0.05,  "Moneda 0.05€"),
]

def calcular_total_denominaciones(cantidades: dict, segunda_caja: float) -> float:
    """cantidades: dict {'b50': 3, 'b20': 5, ...} -> total en €, incluyendo segunda caja."""
    total = segunda_caja
    for clave, valor, _ in DENOMINACIONES:
        total += cantidades.get(clave, 0) * valor
    return total

def get_cierre_arqueo_anterior(fecha: date):
    """
    Devuelve el desglose completo del CIERRE del último arqueo guardado
    antes de 'fecha' (para usarlo como punto de partida de la apertura de hoy).
    None si no hay ningún arqueo previo.
    """
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT ci_b50, ci_b20, ci_b10, ci_b5, ci_m2, ci_m1, ci_m050, ci_m020, ci_m010, ci_m005,
               ci_segunda_caja
        FROM arqueos
        WHERE fecha < %s
        ORDER BY fecha DESC
        LIMIT 1
    """, (fecha,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return None
    claves = [c for c, _, _ in DENOMINACIONES]
    resultado = dict(zip(claves, row[:10]))
    resultado["segunda_caja"] = float(row[10] or 0)
    return resultado

def save_arqueo(fecha: date, apertura: dict, cierre: dict,
                 diferencia_caja_inicial: float,
                 teorico_z: float, teorico_banco: float,
                 real_z: float, real_banco: float,
                 notas: str = "", cerrado: bool = False) -> bool:
    """
    apertura/cierre: dicts con claves 'b50','b20',...,'m005' (cantidades int)
    y 'segunda_caja' (importe €). Cada fecha guarda su propio registro
    independiente; modificar la apertura de hoy nunca toca el cierre de ayer.

    cerrado=False -> guardado parcial: solo importa el conteo físico, el resto
    de valores (teórico/real) se vuelven a leer dinámicamente al recargar.
    cerrado=True  -> cierre definitivo: se congelan todos los valores.
    """
    id_val = fecha.strftime("%Y%m%d")
    total_ap = calcular_total_denominaciones(apertura, apertura.get("segunda_caja", 0))
    total_ci = calcular_total_denominaciones(cierre, cierre.get("segunda_caja", 0))
    diferencia_caja = total_ci - total_ap

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO arqueos (
            id, fecha,
            ap_b50, ap_b20, ap_b10, ap_b5, ap_m2, ap_m1, ap_m050, ap_m020, ap_m010, ap_m005,
            ap_segunda_caja, ap_total,
            ci_b50, ci_b20, ci_b10, ci_b5, ci_m2, ci_m1, ci_m050, ci_m020, ci_m010, ci_m005,
            ci_segunda_caja, ci_total,
            diferencia_caja, diferencia_caja_inicial,
            teorico_z, teorico_banco, real_z, real_banco,
            notas, cerrado, actualizado_en
        ) VALUES (
            %(id)s, %(fecha)s,
            %(ap_b50)s, %(ap_b20)s, %(ap_b10)s, %(ap_b5)s, %(ap_m2)s, %(ap_m1)s,
            %(ap_m050)s, %(ap_m020)s, %(ap_m010)s, %(ap_m005)s,
            %(ap_segunda_caja)s, %(ap_total)s,
            %(ci_b50)s, %(ci_b20)s, %(ci_b10)s, %(ci_b5)s, %(ci_m2)s, %(ci_m1)s,
            %(ci_m050)s, %(ci_m020)s, %(ci_m010)s, %(ci_m005)s,
            %(ci_segunda_caja)s, %(ci_total)s,
            %(diferencia_caja)s, %(diferencia_caja_inicial)s,
            %(teorico_z)s, %(teorico_banco)s, %(real_z)s, %(real_banco)s,
            %(notas)s, %(cerrado)s, NOW()
        )
        ON CONFLICT (id) DO UPDATE SET
            ap_b50=EXCLUDED.ap_b50, ap_b20=EXCLUDED.ap_b20, ap_b10=EXCLUDED.ap_b10, ap_b5=EXCLUDED.ap_b5,
            ap_m2=EXCLUDED.ap_m2, ap_m1=EXCLUDED.ap_m1, ap_m050=EXCLUDED.ap_m050,
            ap_m020=EXCLUDED.ap_m020, ap_m010=EXCLUDED.ap_m010, ap_m005=EXCLUDED.ap_m005,
            ap_segunda_caja=EXCLUDED.ap_segunda_caja, ap_total=EXCLUDED.ap_total,
            ci_b50=EXCLUDED.ci_b50, ci_b20=EXCLUDED.ci_b20, ci_b10=EXCLUDED.ci_b10, ci_b5=EXCLUDED.ci_b5,
            ci_m2=EXCLUDED.ci_m2, ci_m1=EXCLUDED.ci_m1, ci_m050=EXCLUDED.ci_m050,
            ci_m020=EXCLUDED.ci_m020, ci_m010=EXCLUDED.ci_m010, ci_m005=EXCLUDED.ci_m005,
            ci_segunda_caja=EXCLUDED.ci_segunda_caja, ci_total=EXCLUDED.ci_total,
            diferencia_caja=EXCLUDED.diferencia_caja,
            diferencia_caja_inicial=EXCLUDED.diferencia_caja_inicial,
            teorico_z=EXCLUDED.teorico_z, teorico_banco=EXCLUDED.teorico_banco,
            real_z=EXCLUDED.real_z, real_banco=EXCLUDED.real_banco,
            notas=EXCLUDED.notas, cerrado=EXCLUDED.cerrado, actualizado_en=NOW()
    """, {
        "id": id_val, "fecha": fecha,
        "ap_b50": apertura.get("b50", 0), "ap_b20": apertura.get("b20", 0),
        "ap_b10": apertura.get("b10", 0), "ap_b5": apertura.get("b5", 0),
        "ap_m2": apertura.get("m2", 0), "ap_m1": apertura.get("m1", 0),
        "ap_m050": apertura.get("m050", 0), "ap_m020": apertura.get("m020", 0),
        "ap_m010": apertura.get("m010", 0), "ap_m005": apertura.get("m005", 0),
        "ap_segunda_caja": apertura.get("segunda_caja", 0), "ap_total": total_ap,
        "ci_b50": cierre.get("b50", 0), "ci_b20": cierre.get("b20", 0),
        "ci_b10": cierre.get("b10", 0), "ci_b5": cierre.get("b5", 0),
        "ci_m2": cierre.get("m2", 0), "ci_m1": cierre.get("m1", 0),
        "ci_m050": cierre.get("m050", 0), "ci_m020": cierre.get("m020", 0),
        "ci_m010": cierre.get("m010", 0), "ci_m005": cierre.get("m005", 0),
        "ci_segunda_caja": cierre.get("segunda_caja", 0), "ci_total": total_ci,
        "diferencia_caja": diferencia_caja, "diferencia_caja_inicial": diferencia_caja_inicial,
        "teorico_z": teorico_z, "teorico_banco": teorico_banco,
        "real_z": real_z, "real_banco": real_banco,
        "notas": notas, "cerrado": cerrado,
    })
    conn.commit()
    cur.close()
    conn.close()
    return True

def load_arqueo(fecha: date):
    """Devuelve dict con apertura, cierre y todos los datos del arqueo, o None si no existe."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT ap_b50, ap_b20, ap_b10, ap_b5, ap_m2, ap_m1, ap_m050, ap_m020, ap_m010, ap_m005,
               ap_segunda_caja, ap_total,
               ci_b50, ci_b20, ci_b10, ci_b5, ci_m2, ci_m1, ci_m050, ci_m020, ci_m010, ci_m005,
               ci_segunda_caja, ci_total,
               diferencia_caja, diferencia_caja_inicial,
               teorico_z, teorico_banco, real_z, real_banco, notas, cerrado
        FROM arqueos WHERE fecha = %s
    """, (fecha,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return None
    return {
        "apertura": {
            "b50": row[0], "b20": row[1], "b10": row[2], "b5": row[3],
            "m2": row[4], "m1": row[5], "m050": row[6], "m020": row[7],
            "m010": row[8], "m005": row[9], "segunda_caja": float(row[10] or 0),
            "total": float(row[11] or 0),
        },
        "cierre": {
            "b50": row[12], "b20": row[13], "b10": row[14], "b5": row[15],
            "m2": row[16], "m1": row[17], "m050": row[18], "m020": row[19],
            "m010": row[20], "m005": row[21], "segunda_caja": float(row[22] or 0),
            "total": float(row[23] or 0),
        },
        "diferencia_caja": float(row[24] or 0),
        "diferencia_caja_inicial": float(row[25] or 0),
        "teorico_z": float(row[26] or 0),
        "teorico_banco": float(row[27] or 0),
        "real_z": float(row[28] or 0),
        "real_banco": float(row[29] or 0),
        "notas": row[30] or "",
        "cerrado": bool(row[31]),
    }

def get_meses_disponibles():
    """Lista de (año, mes) que tienen datos, más reciente primero."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT EXTRACT(YEAR FROM fecha)::int, EXTRACT(MONTH FROM fecha)::int
        FROM cierres_caja
        ORDER BY 1 DESC, 2 DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows

def load_historico(anio=None, mes=None):
    """Cierres filtrados por año/mes. Si mes es None pero anio no, filtra el año completo. Si ambos son None, trae todo."""
    conn = get_connection()
    cur = conn.cursor()
    where = ""
    params = []
    if anio is not None and mes is not None:
        where = "WHERE EXTRACT(YEAR FROM fecha) = %s AND EXTRACT(MONTH FROM fecha) = %s"
        params = [anio, mes]
    elif anio is not None:
        where = "WHERE EXTRACT(YEAR FROM fecha) = %s"
        params = [anio]
    cur.execute(f"""
        SELECT fecha, facturacion, gastos, visa, internet, justeat, glovo, uber,
               ticket_restaurant, ingreso_banco, z_caja, facturacion_total
        FROM cierres_caja
        {where}
        ORDER BY fecha DESC
    """, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows

def get_facturacion_por_dia_semana(anio=None, mes=None):
    """
    Suma facturacion_total agrupada por día de la semana (0=lunes...6=domingo),
    con el mismo filtro de año/mes que load_historico (None = sin filtrar, es
    decir todo el histórico). Devuelve (lista_de_7_totales, total_general).
    """
    rows = load_historico(anio, mes)
    totales = [0.0] * 7
    for row in rows:
        fecha = row[0]
        total = float(row[11] or 0)   # facturacion_total es la última columna del SELECT
        totales[fecha.weekday()] += total
    return totales, sum(totales)

def load_evolucion_mensual(anio_desde=None, anio_hasta=None):
    """
    Datos agregados por mes para los gráficos:
    - facturacion_total (suma del mes)
    - justeat, glovo, uber (suma del mes)
    Excluye el mes en curso (incompleto) para no distorsionar la tendencia.
    Si se indican anio_desde/anio_hasta, filtra ese rango de años (inclusive).
    Devuelve filas ordenadas cronológicamente.
    """
    conn = get_connection()
    cur = conn.cursor()
    where = ["DATE_TRUNC('month', fecha) < DATE_TRUNC('month', CURRENT_DATE)"]
    params = []
    if anio_desde is not None:
        where.append("EXTRACT(YEAR FROM fecha) >= %s")
        params.append(anio_desde)
    if anio_hasta is not None:
        where.append("EXTRACT(YEAR FROM fecha) <= %s")
        params.append(anio_hasta)
    where_sql = " AND ".join(where)
    cur.execute(f"""
        SELECT
            DATE_TRUNC('month', fecha)::date AS mes,
            SUM(facturacion_total) AS total,
            SUM(justeat) AS justeat,
            SUM(glovo) AS glovo,
            SUM(uber) AS uber
        FROM cierres_caja
        WHERE {where_sql}
        GROUP BY 1
        ORDER BY 1
    """, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows

def get_rango_anios_disponibles():
    """Devuelve (anio_min, anio_max) de los datos guardados."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT MIN(EXTRACT(YEAR FROM fecha))::int, MAX(EXTRACT(YEAR FROM fecha))::int FROM cierres_caja")
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row if row and row[0] else (date.today().year, date.today().year)

def get_ranking_meses(limit=10):
    """
    Top N meses por facturación total, excluyendo el mes en curso (incompleto).
    """
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            DATE_TRUNC('month', fecha)::date AS mes,
            SUM(facturacion_total) AS total
        FROM cierres_caja
        WHERE DATE_TRUNC('month', fecha) < DATE_TRUNC('month', CURRENT_DATE)
        GROUP BY 1
        ORDER BY 2 DESC
        LIMIT %s
    """, (limit,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows

def get_count_meses_disponibles() -> int:
    """Número total de meses cerrados (excluyendo el mes en curso) con datos."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(DISTINCT DATE_TRUNC('month', fecha))
        FROM cierres_caja
        WHERE DATE_TRUNC('month', fecha) < DATE_TRUNC('month', CURRENT_DATE)
    """)
    total = cur.fetchone()[0]
    cur.close()
    conn.close()
    return int(total or 0)

# ─────────────────────────────────────────────
#  KPIs COMPARATIVOS
# ─────────────────────────────────────────────

def _suma_periodo(desde: date, hasta: date) -> float:
    """Suma de facturacion_total entre dos fechas (inclusive)."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT COALESCE(SUM(facturacion_total), 0) FROM cierres_caja
        WHERE fecha BETWEEN %s AND %s
    """, (desde, hasta))
    total = cur.fetchone()[0]
    cur.close()
    conn.close()
    return float(total)

def get_ultima_fecha_disponible() -> date:
    """Devuelve la fecha del último cierre guardado, o hoy si no hay ninguno."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT MAX(fecha) FROM cierres_caja")
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row[0] if row and row[0] else date.today()

def _valor_dia(fecha: date) -> float:
    """facturacion_total de un día concreto, o None si no hay cierre."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT facturacion_total FROM cierres_caja WHERE fecha = %s", (fecha,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return float(row[0]) if row else None

def _variacion(actual, anterior):
    """Devuelve (diferencia_absoluta, porcentaje) o (None, None) si no se puede calcular."""
    if actual is None or anterior is None or anterior == 0:
        return None, None
    diff = actual - anterior
    pct = (diff / anterior) * 100
    return diff, pct

def calcular_kpis_generales() -> dict:
    """
    Calcula un conjunto de comparativas clave, usando como referencia el
    último día con cierre guardado (no 'hoy', ya que el cierre del día
    actual se introduce por la noche y aún no existe en la base de datos):
    - Último día disponible vs mismo día semana pasada
    - Último día disponible vs mismo día año pasado
    - Mes en curso (hasta ese día) vs mismo periodo mes pasado
    - Mes en curso (hasta ese día) vs mismo periodo mismo mes año pasado
    - Año en curso (hasta ese día) vs mismo periodo año pasado
    """
    hoy = get_ultima_fecha_disponible()
    kpis = {}

    # Último día disponible vs mismo día de la semana pasada
    hoy_val = _valor_dia(hoy)
    semana_pasada = hoy - timedelta(days=7)
    semana_pasada_val = _valor_dia(semana_pasada)
    diff, pct = _variacion(hoy_val, semana_pasada_val)
    kpis["hoy_vs_semana_pasada"] = {
        "label": f"{hoy.strftime('%d/%m')} vs. mismo día semana pasada",
        "actual": hoy_val, "anterior": semana_pasada_val,
        "fecha_actual": hoy, "fecha_anterior": semana_pasada,
        "diff": diff, "pct": pct,
    }

    # Último día disponible vs mismo día año pasado
    try:
        mismo_dia_anio_pasado = hoy.replace(year=hoy.year - 1)
    except ValueError:
        mismo_dia_anio_pasado = hoy.replace(year=hoy.year - 1, day=28)
    anio_pasado_val = _valor_dia(mismo_dia_anio_pasado)
    diff, pct = _variacion(hoy_val, anio_pasado_val)
    kpis["hoy_vs_anio_pasado"] = {
        "label": f"{hoy.strftime('%d/%m')} vs. mismo día año pasado",
        "actual": hoy_val, "anterior": anio_pasado_val,
        "fecha_actual": hoy, "fecha_anterior": mismo_dia_anio_pasado,
        "diff": diff, "pct": pct,
    }

    # Mes en curso (1 al día de hoy) vs mismo tramo del mes anterior
    inicio_mes_actual = hoy.replace(day=1)
    dias_transcurridos = (hoy - inicio_mes_actual).days  # 0-indexado

    if inicio_mes_actual.month == 1:
        inicio_mes_anterior = date(inicio_mes_actual.year - 1, 12, 1)
    else:
        inicio_mes_anterior = date(inicio_mes_actual.year, inicio_mes_actual.month - 1, 1)
    fin_tramo_mes_anterior = inicio_mes_anterior + timedelta(days=dias_transcurridos)

    actual_mes = _suma_periodo(inicio_mes_actual, hoy)
    anterior_mes = _suma_periodo(inicio_mes_anterior, fin_tramo_mes_anterior)
    diff, pct = _variacion(actual_mes, anterior_mes)
    kpis["mes_vs_mes_anterior"] = {
        "label": f"Mes en curso (hasta {hoy.strftime('%d/%m')}) vs. mes anterior (mismo nº de días)",
        "actual": actual_mes, "anterior": anterior_mes,
        "fecha_actual": (inicio_mes_actual, hoy), "fecha_anterior": (inicio_mes_anterior, fin_tramo_mes_anterior),
        "diff": diff, "pct": pct,
    }

    # Mes en curso vs mismo mes año pasado (mismo tramo de días)
    try:
        inicio_mismo_mes_anio_pasado = inicio_mes_actual.replace(year=inicio_mes_actual.year - 1)
    except ValueError:
        inicio_mismo_mes_anio_pasado = inicio_mes_actual.replace(year=inicio_mes_actual.year - 1, day=28)
    fin_tramo_anio_pasado = inicio_mismo_mes_anio_pasado + timedelta(days=dias_transcurridos)

    anterior_anio = _suma_periodo(inicio_mismo_mes_anio_pasado, fin_tramo_anio_pasado)
    diff, pct = _variacion(actual_mes, anterior_anio)
    kpis["mes_vs_mismo_mes_anio_pasado"] = {
        "label": f"Mes en curso (hasta {hoy.strftime('%d/%m')}) vs. mismo mes año pasado",
        "actual": actual_mes, "anterior": anterior_anio,
        "fecha_actual": (inicio_mes_actual, hoy),
        "fecha_anterior": (inicio_mismo_mes_anio_pasado, fin_tramo_anio_pasado),
        "diff": diff, "pct": pct,
    }

    # ── Mes anterior COMPLETO (último mes cerrado) vs mismo mes año pasado, ambos completos ──
    if inicio_mes_anterior.month == 12:
        fin_mes_anterior_completo = date(inicio_mes_anterior.year, 12, 31)
    else:
        fin_mes_anterior_completo = date(inicio_mes_anterior.year, inicio_mes_anterior.month + 1, 1) - timedelta(days=1)

    try:
        inicio_mismo_mes_anterior_anio_pasado = inicio_mes_anterior.replace(year=inicio_mes_anterior.year - 1)
    except ValueError:
        inicio_mismo_mes_anterior_anio_pasado = inicio_mes_anterior.replace(year=inicio_mes_anterior.year - 1, day=28)
    if inicio_mismo_mes_anterior_anio_pasado.month == 12:
        fin_mismo_mes_anterior_anio_pasado = date(inicio_mismo_mes_anterior_anio_pasado.year, 12, 31)
    else:
        fin_mismo_mes_anterior_anio_pasado = date(
            inicio_mismo_mes_anterior_anio_pasado.year,
            inicio_mismo_mes_anterior_anio_pasado.month + 1, 1
        ) - timedelta(days=1)

    actual_mes_anterior_completo = _suma_periodo(inicio_mes_anterior, fin_mes_anterior_completo)
    anterior_mismo_mes_anio_pasado_completo = _suma_periodo(
        inicio_mismo_mes_anterior_anio_pasado, fin_mismo_mes_anterior_anio_pasado
    )
    diff, pct = _variacion(actual_mes_anterior_completo, anterior_mismo_mes_anio_pasado_completo)
    kpis["mes_anterior_completo_vs_mismo_mes_anio_pasado"] = {
        "label": f"Último mes completo ({_MESES_ES_GLOBAL[inicio_mes_anterior.month-1]} {inicio_mes_anterior.year}) vs. mismo mes año pasado (ambos completos)",
        "actual": actual_mes_anterior_completo, "anterior": anterior_mismo_mes_anio_pasado_completo,
        "fecha_actual": (inicio_mes_anterior, fin_mes_anterior_completo),
        "fecha_anterior": (inicio_mismo_mes_anterior_anio_pasado, fin_mismo_mes_anterior_anio_pasado),
        "diff": diff, "pct": pct,
    }

    # Año en curso (1 enero - hoy) vs mismo tramo año pasado
    inicio_anio_actual = date(hoy.year, 1, 1)
    dias_transcurridos_anio = (hoy - inicio_anio_actual).days
    inicio_anio_pasado = date(hoy.year - 1, 1, 1)
    fin_tramo_anio_pasado_completo = inicio_anio_pasado + timedelta(days=dias_transcurridos_anio)

    actual_anio = _suma_periodo(inicio_anio_actual, hoy)
    anterior_anio_completo = _suma_periodo(inicio_anio_pasado, fin_tramo_anio_pasado_completo)
    diff, pct = _variacion(actual_anio, anterior_anio_completo)
    kpis["anio_vs_anio_pasado"] = {
        "label": f"Año en curso (hasta {hoy.strftime('%d/%m')}) vs. año pasado (mismo tramo)",
        "actual": actual_anio, "anterior": anterior_anio_completo,
        "fecha_actual": (inicio_anio_actual, hoy),
        "fecha_anterior": (inicio_anio_pasado, fin_tramo_anio_pasado_completo),
        "diff": diff, "pct": pct,
    }

    # ── Última semana completa (lunes-domingo) vs la semana anterior a esa ──
    lunes_semana_de_hoy = hoy - timedelta(days=hoy.weekday())
    if hoy.weekday() == 6:  # si el último día disponible es domingo, esa semana ya está completa
        lunes_actual = lunes_semana_de_hoy
    else:
        lunes_actual = lunes_semana_de_hoy - timedelta(days=7)
    domingo_actual = lunes_actual + timedelta(days=6)
    lunes_pasado = lunes_actual - timedelta(days=7)
    domingo_pasado = lunes_pasado + timedelta(days=6)

    actual_semana = _suma_periodo(lunes_actual, domingo_actual)
    anterior_semana = _suma_periodo(lunes_pasado, domingo_pasado)
    diff, pct = _variacion(actual_semana, anterior_semana)
    kpis["semana_vs_semana_pasada"] = {
        "label": "Última semana completa vs. la anterior",
        "actual": actual_semana, "anterior": anterior_semana,
        "fecha_actual": (lunes_actual, domingo_actual),
        "fecha_anterior": (lunes_pasado, domingo_pasado),
        "diff": diff, "pct": pct,
    }

    # ── Mejor y peor día del mes en curso ──
    mejor_dia, peor_dia = _mejor_peor_dia_mes(inicio_mes_actual, hoy)
    kpis["mejor_peor_dia_mes"] = {"mejor": mejor_dia, "peor": peor_dia}

    # ── Media diaria mes actual vs media diaria mes anterior ──
    n_dias_actual = dias_transcurridos + 1  # incluye hoy
    n_dias_anterior = dias_transcurridos + 1
    media_actual = actual_mes / n_dias_actual if n_dias_actual else None
    media_anterior = anterior_mes / n_dias_anterior if n_dias_anterior else None
    diff, pct = _variacion(media_actual, media_anterior)
    kpis["media_diaria_mes"] = {
        "label": "Media diaria: mes en curso vs. mes anterior",
        "actual": media_actual, "anterior": media_anterior,
        "fecha_actual": (inicio_mes_actual, hoy),
        "fecha_anterior": (inicio_mes_anterior, fin_tramo_mes_anterior),
        "diff": diff, "pct": pct,
    }

    # ── Cuota por plataforma (mes en curso) + cuál crece/cae más vs mes anterior ──
    kpis["cuota_plataformas"] = _calcular_cuota_plataformas(inicio_mes_actual, hoy,
                                                              inicio_mes_anterior, fin_tramo_mes_anterior)

    return kpis

def _mejor_peor_dia_mes(desde: date, hasta: date):
    """Devuelve (mejor_dia, peor_dia) como tuplas (fecha, valor) del rango dado."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT fecha, facturacion_total FROM cierres_caja
        WHERE fecha BETWEEN %s AND %s
        ORDER BY facturacion_total DESC
    """, (desde, hasta))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    if not rows:
        return None, None
    mejor = (rows[0][0], float(rows[0][1]))
    peor = (rows[-1][0], float(rows[-1][1]))
    return mejor, peor

def _suma_plataforma_periodo(plataforma: str, desde: date, hasta: date) -> float:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT COALESCE(SUM({plataforma}), 0) FROM cierres_caja
        WHERE fecha BETWEEN %s AND %s
    """, (desde, hasta))
    total = cur.fetchone()[0]
    cur.close()
    conn.close()
    return float(total)

def _calcular_cuota_plataformas(desde_actual, hasta_actual, desde_anterior, hasta_anterior):
    """
    Cuota de cada plataforma (justeat/glovo/uber) sobre el total de plataformas
    en el mes actual, y comparación de crecimiento vs mes anterior.
    """
    plataformas = ["justeat", "glovo", "uber"]
    datos = {}
    total_actual = 0.0
    total_anterior = 0.0

    for p in plataformas:
        actual = _suma_plataforma_periodo(p, desde_actual, hasta_actual)
        anterior = _suma_plataforma_periodo(p, desde_anterior, hasta_anterior)
        datos[p] = {"actual": actual, "anterior": anterior}
        total_actual += actual
        total_anterior += anterior

    resultado = {"total_actual": total_actual, "total_anterior": total_anterior, "plataformas": {}}
    mejor_crecimiento = None  # (plataforma, pct)

    for p in plataformas:
        actual = datos[p]["actual"]
        anterior = datos[p]["anterior"]
        cuota = (actual / total_actual * 100) if total_actual else 0.0
        diff, pct = _variacion(actual, anterior)
        resultado["plataformas"][p] = {
            "actual": actual, "anterior": anterior, "cuota": cuota, "diff": diff, "pct": pct
        }
        if pct is not None:
            if mejor_crecimiento is None or pct > mejor_crecimiento[1]:
                mejor_crecimiento = (p, pct)

    resultado["mejor_crecimiento"] = mejor_crecimiento
    return resultado


# ─────────────────────────────────────────────
#  VENTANA PRINCIPAL
# ─────────────────────────────────────────────

class PizzeriaApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.withdraw()  # oculta la ventana mientras se pide la contraseña

        # Calibra la escala real de Tk a la densidad de píxeles del monitor.
        # Junto con el DPI-awareness avisado a Windows más arriba, esto evita
        # que el texto salga diminuto o desproporcionado en pantallas de alta
        # resolución (muy comunes en portátiles con Windows).
        try:
            self.tk.call("tk", "scaling", self.winfo_fpixels("1i") / 72.0)
        except Exception:
            pass

        self.title("🍕 CarPizzzeta — Cierre de Caja")
        self.modo_acceso = self._pedir_login()   # 'master' o 'empleado'
        if self.modo_acceso == "empleado":
            self.title("🍕 CarPizzzeta — Cierre de Caja  [Acceso empleado]")

        self.deiconify()  # ya autenticado: mostramos la ventana
        self.configure(bg=C["bg"])
        self.resizable(True, True)
        self.geometry("900x780")
        self.minsize(760, 680)

        # Estado del cierre actual: ¿se ha guardado ya, hay cambios sin guardar?
        self.cierre_guardado = True
        self.fecha_desbloqueada = None  # fecha (no-hoy) ya autorizada con contraseña en esta sesión de fecha

        self._build_header()
        self._build_status()
        self._build_notebook()
        if self.modo_acceso == "empleado":
            self._aplicar_restriccion_empleado()
        self.bind("<Button-1>", lambda e: e.widget.focus_set() if isinstance(e.widget, (tk.Entry, tk.Text)) else self.focus_set())        # Init DB on start
        self.protocol("WM_DELETE_WINDOW", self._on_close_app)

        try:
            init_db()
            self._load_today()
            self._populate_meses_combo()
            if self.modo_acceso == "empleado":
                self._restringir_historico_anio_actual()
            self._refresh_historico()
            self._populate_dow_combo()
            if self.modo_acceso == "empleado":
                self._restringir_dow_anio_actual()
            self._refresh_dow_tabla()
            self._calcular_prevision()
            self.set_status("Conectado a PostgreSQL ✓", C["success"])
        except Exception as e:
            self.set_status(f"Error de conexión: {e}", C["danger"])
            messagebox.showerror("Error de conexión",
                f"No se pudo conectar a PostgreSQL.\n\n{e}\n\n"
                "Edita DB_CONFIG en app.py o configura las variables de entorno.")

    # ── LOGIN AL ARRANCAR (doble contraseña: empleado / master) ──
    def _pedir_login(self) -> str:
        """
        Pide contraseña al arrancar la app. Devuelve 'master' o 'empleado'
        según cuál se haya introducido. Si se cancela o se falla demasiadas
        veces, cierra la aplicación directamente.
        """
        intentos = 0
        max_intentos = 3
        while intentos < max_intentos:
            pwd = simpledialog.askstring(
                "🍕 CarPizzzeta — Acceso",
                "Introduce la contraseña:",
                show="*", parent=self
            )
            if pwd is None:   # se pulsó "Cancelar"
                self.destroy()
                sys.exit(0)
            if pwd == PASSWORD_MASTER:
                return "master"
            if pwd == PASSWORD_EMPLEADO:
                return "empleado"
            intentos += 1
            restantes = max_intentos - intentos
            if restantes > 0:
                messagebox.showerror(
                    "Contraseña incorrecta",
                    f"Contraseña incorrecta. Te quedan {restantes} intento(s)."
                )
        messagebox.showerror("Acceso denegado",
                              "Demasiados intentos fallidos. Cerrando la aplicación.")
        self.destroy()
        sys.exit(0)

    def _aplicar_restriccion_empleado(self):
        """
        En modo empleado, oculta del Notebook las pestañas que no estén en
        TABS_VISIBLES_EMPLEADO. nb.forget() solo las desvincula del Notebook
        (no las destruye), así que el resto de la app sigue funcionando igual.
        """
        todas = ["tab_cierre", "tab_arqueo", "tab_historico", "tab_prevision",
                 "tab_seguimiento",
                 "tab_mapa", "tab_calendario", "tab_graficos", "tab_kpis", "tab_ranking"]
        for nombre_attr in todas:
            if nombre_attr in TABS_VISIBLES_EMPLEADO:
                continue
            tab = getattr(self, nombre_attr, None)
            if tab is not None:
                try:
                    self.nb.forget(tab)
                except Exception:
                    pass

    def _restringir_historico_anio_actual(self):
        """
        En modo empleado, la pestaña Histórico solo puede mostrar el año en
        curso — se quita "Todo el histórico" y los años/meses anteriores de
        los selectores, para que no se pueda consultar la facturación de
        años pasados.
        """
        anio_actual = date.today().year

        # Combo de meses: solo los meses que pertenecen al año en curso
        meses_anio_actual = [lbl for lbl, (a, m) in self._meses_map.items() if a == anio_actual]
        orden_original = [v for v in self.mes_combo["values"] if v in meses_anio_actual]
        self.mes_combo["values"] = orden_original
        self._meses_map = {k: v for k, v in self._meses_map.items() if k in orden_original}

        # Combo de años: solo el año en curso (sin "—" = todo, ni años anteriores)
        self.anio_combo["values"] = [str(anio_actual)]
        self._anios_map = {str(anio_actual): anio_actual}

        # Selección por defecto: el mes en curso si hay datos; si no, el año completo
        hoy = date.today()
        label_mes_actual = f"{self.MESES_ES[hoy.month-1]} {hoy.year}"
        if label_mes_actual in self._meses_map:
            self.mes_var.set(label_mes_actual)
        elif orden_original:
            self.mes_var.set(orden_original[0])
        else:
            self.mes_var.set(str(anio_actual))
            self.anio_var.set(str(anio_actual))



    # ── UNDO/REDO PARA ENTRY (tk.Entry no lo soporta de forma nativa) ──
    def _bind_undo_redo(self, entry, var):
        """
        A diferencia de tk.Text, tk.Entry no tiene ningún mecanismo de undo/redo
        en Tk — por eso Ctrl+Z no hacía nada en los campos numéricos. Aquí se
        implementa a mano: guardamos un historial de valores de `var` y
        Ctrl+Z / Cmd+Z (Mac) retrocede, Ctrl+Shift+Z / Cmd+Shift+Z avanza.
        """
        hist = {"stack": [], "pos": 0, "lock": False}

        def _snapshot():
            try:
                return str(var.get())
            except Exception:
                return ""

        hist["stack"].append(_snapshot())

        def _on_change(*_args):
            if hist["lock"]:
                return
            valor = _snapshot()
            if valor != hist["stack"][hist["pos"]]:
                hist["stack"] = hist["stack"][:hist["pos"] + 1]
                hist["stack"].append(valor)
                hist["pos"] = len(hist["stack"]) - 1

        def _restore(valor):
            hist["lock"] = True
            try:
                var.set(valor)
            except Exception:
                pass
            hist["lock"] = False

        def _undo(_event=None):
            if hist["pos"] > 0:
                hist["pos"] -= 1
                _restore(hist["stack"][hist["pos"]])
            return "break"

        def _redo(_event=None):
            if hist["pos"] < len(hist["stack"]) - 1:
                hist["pos"] += 1
                _restore(hist["stack"][hist["pos"]])
            return "break"

        var.trace_add("write", _on_change)
        for seq in ("<Control-z>", "<Control-Z>", "<Command-z>"):
            entry.bind(seq, _undo)
        for seq in ("<Control-Shift-Z>", "<Control-Shift-z>", "<Command-Shift-Z>", "<Command-Shift-z>"):
            entry.bind(seq, _redo)

    # ── SCROLL CON RUEDA DEL RATÓN (multiplataforma) ──
    def _bind_canvas_scroll(self, canvas):
        """
        Activa el scroll con la rueda del ratón sobre `canvas` mientras el
        cursor está encima (Windows, macOS y Linux/X11 usan eventos distintos):
          - Windows/Mac: <MouseWheel>, con event.delta. En Windows es múltiplo
            de 120; en Mac suele ser un entero pequeño (1, 2, 3...), por eso NO
            se puede dividir por 120 en ambos casos — solo importa el signo.
          - Linux/X11: no dispara <MouseWheel>; usa <Button-4> (arriba) y
            <Button-5> (abajo).
        """
        def _on_wheel(event):
            if getattr(event, "num", None) == 4:
                canvas.yview_scroll(-1, "units")
            elif getattr(event, "num", None) == 5:
                canvas.yview_scroll(1, "units")
            elif getattr(event, "delta", 0):
                canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")

        def _activar(_e=None):
            canvas.bind_all("<MouseWheel>", _on_wheel)
            canvas.bind_all("<Button-4>", _on_wheel)
            canvas.bind_all("<Button-5>", _on_wheel)

        def _desactivar(_e=None):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")

        canvas.bind("<Enter>", _activar)
        canvas.bind("<Leave>", _desactivar)

    # ── BOTÓN (Label clicable, fiable en Mac) ──
    def make_button(self, parent, text, command, bg, fg="white", font=None,
                     padx=10, pady=6, side=None, padx_pack=0, pady_pack=0, **pack_kwargs):
        """
        tk.Button ignora 'bg' en macOS (lo pinta el sistema). Usamos un Label
        clicable con el mismo aspecto, que sí respeta los colores en todas
        las plataformas.
        """
        font = font or FONT_SMALL
        btn = tk.Label(parent, text=text, font=font, bg=bg, fg=fg,
                        padx=padx, pady=pady, cursor="hand2")
        btn.bind("<Button-1>", lambda ev: command())
        bg_hover = bg
        btn.bind("<Enter>", lambda ev: btn.config(bg=self._aclarar_color(bg_hover)))
        btn.bind("<Leave>", lambda ev: btn.config(bg=bg_hover))
        if side is not None:
            pack_kwargs["side"] = side
        if padx_pack:
            pack_kwargs["padx"] = padx_pack
        if pady_pack:
            pack_kwargs["pady"] = pady_pack
        btn.pack(**pack_kwargs)
        return btn

    @staticmethod
    def _aclarar_color(hexcolor, factor=1.15):
        """Aclara ligeramente un color hexadecimal para el efecto hover."""
        try:
            hexcolor = hexcolor.lstrip("#")
            r, g, b = int(hexcolor[0:2], 16), int(hexcolor[2:4], 16), int(hexcolor[4:6], 16)
            r = min(255, int(r * factor))
            g = min(255, int(g * factor))
            b = min(255, int(b * factor))
            return f"#{r:02x}{g:02x}{b:02x}"
        except Exception:
            return hexcolor

    # ── HEADER ───────────────────────────────
    def _build_header(self):
        hdr = tk.Frame(self, bg=C["card"], pady=14)
        hdr.pack(fill="x")

        tk.Label(hdr, text="🍕", font=("Helvetica", 28),
                 bg=C["card"], fg=C["accent"]).pack(side="left", padx=(20,8))
        tk.Label(hdr, text="CarPizzzeta", font=(FONT_SANS, 20, "bold"),
                 bg=C["card"], fg=C["text"]).pack(side="left")
        tk.Label(hdr, text="cierre de caja diario", font=(FONT_SANS, 10),
                 bg=C["card"], fg=C["muted"]).pack(side="left", padx=(8,0), pady=(6,0))


        try:
            locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
        except:
            locale.setlocale(locale.LC_TIME, "")
        today_str = date.today().strftime("%A, %d de %B de %Y").capitalize()
        tk.Label(hdr, text=today_str, font=FONT_SMALL,
                 bg=C["card"], fg=C["accent2"]).pack(side="right", padx=20)

    # ── NOTEBOOK (pestañas) ───────────────────
    def _build_notebook(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TNotebook",
            background=C["bg"], borderwidth=0)
        style.configure("TNotebook.Tab",
            background=C["panel"], foreground=C["muted"],
            padding=[18, 8], font=FONT_LABEL)
        style.map("TNotebook.Tab",
            background=[("selected", C["card"])],
            foreground=[("selected", C["accent"])])

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=12, pady=(8,0))

        self.tab_cierre    = tk.Frame(self.nb, bg=C["bg"])
        self.tab_arqueo    = tk.Frame(self.nb, bg=C["bg"])
        self.tab_historico = tk.Frame(self.nb, bg=C["bg"])
        self.tab_prevision = tk.Frame(self.nb, bg=C["bg"])
        self.tab_seguimiento = tk.Frame(self.nb, bg=C["bg"])
        self.tab_mapa      = tk.Frame(self.nb, bg=C["bg"])
        self.tab_calendario = tk.Frame(self.nb, bg=C["bg"])
        self.tab_graficos  = tk.Frame(self.nb, bg=C["bg"])
        self.tab_kpis      = tk.Frame(self.nb, bg=C["bg"])
        self.tab_ranking   = tk.Frame(self.nb, bg=C["bg"])

        self.nb.add(self.tab_cierre,    text=" 📋  Cierre del Día ")
        self.nb.add(self.tab_arqueo,    text=" 💰  Arqueo ")
        self.nb.add(self.tab_historico, text=" 📊  Histórico ")
        self.nb.add(self.tab_prevision, text=" 🔮  Previsión ")
        self.nb.add(self.tab_seguimiento, text=" 🎯  Seguimiento ")
        self.nb.add(self.tab_mapa,      text=" 🗓️  Calendario Semanal ")
        self.nb.add(self.tab_calendario, text=" ⚽  Eventos ")
        self.nb.add(self.tab_graficos,  text=" 📈  Gráficos ")
        self.nb.add(self.tab_kpis,      text=" 📌  KPIs ")
        self.nb.add(self.tab_ranking,   text=" 🏆  Ranking de Meses ")

        self._build_tab_cierre()
        self._build_tab_arqueo()
        self._build_tab_historico()
        self._build_tab_prevision()
        self._build_tab_seguimiento()
        self._build_tab_mapa()
        self._build_tab_calendario()
        self._build_tab_graficos()
        self._build_tab_kpis()
        self._build_tab_ranking()

        self.nb.bind("<<NotebookTabChanged>>", self._on_tab_changed)

    # ── TAB: CIERRE DEL DÍA ──────────────────
    def _build_tab_cierre(self):
        parent = self.tab_cierre

        # Date row
        date_row = tk.Frame(parent, bg=C["bg"])
        date_row.pack(fill="x", padx=20, pady=(14, 2))
        tk.Label(date_row, text="Fecha:", font=FONT_LABEL,
                 bg=C["bg"], fg=C["muted"]).pack(side="left")
        self.fecha_var = tk.StringVar(value=date.today().isoformat())
        date_entry = tk.Entry(date_row, textvariable=self.fecha_var,
                              font=FONT_ENTRY, width=12,
                              bg=C["entry_bg"], fg=C["accent2"],
                              insertbackground=C["accent2"],
                              relief="flat", bd=4)
        date_entry.pack(side="left", padx=8)
        self._bind_undo_redo(date_entry, self.fecha_var)
        self.make_button(date_row, "Cargar", self._load_by_date,
                          bg=C["accent"], font=FONT_SMALL, side="left")

        tk.Label(date_row, text="Tipo de día:", font=FONT_LABEL,
                 bg=C["bg"], fg=C["muted"]).pack(side="left", padx=(20,6))
        self.tipo_dia_var = tk.StringVar(value="normal")
        tipo_combo = ttk.Combobox(date_row, textvariable=self.tipo_dia_var,
                                   state="readonly", width=10, font=FONT_SMALL,
                                   values=["normal", "festivo", "vispera", "fm"])
        tipo_combo.pack(side="left")

        # Scrollable canvas
        canvas = tk.Canvas(parent, bg=C["bg"], highlightthickness=0)
        scrollbar = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        self.scroll_frame = tk.Frame(canvas, bg=C["bg"])
        self.scroll_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=self.scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(20,0), pady=8)
        scrollbar.pack(side="right", fill="y", pady=8)
        self._bind_canvas_scroll(canvas)

        sf = self.scroll_frame
        self.entries = {}

        def section(title, show_status=False):
            f = tk.Frame(sf, bg=C["panel"], bd=0)
            f.pack(fill="x", pady=(10,2), padx=4)
            header_row = tk.Frame(f, bg=C["panel"])
            header_row.pack(fill="x", padx=12, pady=6)
            tk.Label(header_row, text=title, font=(FONT_SANS, 10, "bold"),
                     bg=C["panel"], fg=C["accent"]).pack(side="left")
            if show_status:
                self.estado_cierre_lbl = tk.Label(header_row, text="", font=(FONT_SANS, 10, "bold"),
                                                   bg=C["panel"], fg=C["danger"])
                self.estado_cierre_lbl.pack(side="left", padx=(16,0))
            return f
        
        def row(parent_frame, label, key, is_auto=False):
            row_f = tk.Frame(parent_frame, bg=C["card"])
            row_f.pack(fill="x", padx=2, pady=1)
            tk.Label(row_f, text=label, font=FONT_LABEL, width=22, anchor="w",
                     bg=C["card"], fg=C["text"] if not is_auto else C["muted"],
                     padx=10).pack(side="left")
            tk.Label(row_f, text="€", font=FONT_ENTRY,
                     bg=C["card"], fg=C["accent2"]).pack(side="left")
            var = tk.StringVar(value="0.00")
            self.entries[key] = var
            state = "readonly" if is_auto else "normal"

            def solo_numeros(P):
                return P == "" or P.replace(".", "").replace(",", "").isdigit()
            vcmd = (self.register(solo_numeros), "%P")

            e = tk.Entry(row_f, textvariable=var, font=FONT_ENTRY,
                         width=14, bg=C["entry_bg"] if not is_auto else C["panel"],
                         fg=C["accent2"] if not is_auto else C["muted"],
                         insertbackground=C["accent2"],
                         relief="flat", bd=6, state=state,
                         **({} if is_auto else {"validate": "key", "validatecommand": vcmd}))
            e.pack(side="left", padx=6)
            if not is_auto:
                e.bind("<FocusIn>",  lambda ev, en=e: en.select_range(0, "end"))
                e.bind("<FocusOut>", lambda ev, v=var: v.set(f"{float(v.get().replace(',','.') or 0):.2f}"))
                e.bind("<KeyRelease>", lambda ev: (self._recalculate(), self._marcar_sin_guardar()))
                self._bind_undo_redo(e, var)
            return var

# ── Sección 1: Facturación y canales ──
        s1 = section("💰  Facturación y Canales", show_status=True)
        row(s1, "Facturación", "facturacion")
        row(s1, "Gastos del día", "gastos")
        row(s1, "Visa / TPV", "visa")
        row(s1, "Internet", "internet")
        row(s1, "Just Eat", "justeat")
        row(s1, "Glovo", "glovo")
        row(s1, "Uber Eats", "uber")
        row(s1, "Ticket Restaurant", "ticket_restaurant")
        row(s1, "Ingreso banco", "ingreso_banco", is_auto=True)
        row(s1, "Z ", "z_caja")
        row(s1, "TOTAL Facturación", "facturacion_total", is_auto=True)

        # ── Notas ──
        notes_frame = tk.Frame(sf, bg=C["card"])
        notes_frame.pack(fill="x", padx=4, pady=(10,2))
        tk.Label(notes_frame, text="📝  Notas", font=(FONT_SANS, 10, "bold"),
                 bg=C["card"], fg=C["accent"], padx=10, pady=6).pack(anchor="w")
        self.notas_text = tk.Text(notes_frame, height=3, font=FONT_LABEL,
                                  bg=C["entry_bg"], fg=C["text"],
                                  insertbackground=C["text"],
                                  relief="flat", bd=6, wrap="word",
                                  undo=True, autoseparators=True, maxundo=-1)
        self.notas_text.pack(fill="x", padx=10, pady=(0,8))

        # ── Botones ──
        btn_row = tk.Frame(sf, bg=C["bg"])
        btn_row.pack(pady=16)
        self.make_button(btn_row, "💾  Guardar Cierre", self._save,
                          bg=C["accent"], font=FONT_BUTTON, padx=28, pady=10,
                          side="left", padx_pack=8)
        self.make_button(btn_row, "🗑  Limpiar", self._clear,
                          bg=C["panel"], fg=C["muted"], font=FONT_BUTTON, padx=20, pady=10,
                          side="left", padx_pack=8)

    # ── TAB: ARQUEO DE CAJA ────────────────────
    def _build_tab_arqueo(self):
        parent = self.tab_arqueo

        # Selector de fecha
        date_row = tk.Frame(parent, bg=C["bg"])
        date_row.pack(fill="x", padx=20, pady=(14, 8))
        tk.Label(date_row, text="Fecha:", font=FONT_LABEL,
                 bg=C["bg"], fg=C["muted"]).pack(side="left")
        self.arq_fecha_var = tk.StringVar(value=date.today().isoformat())
        arq_fecha_entry = tk.Entry(date_row, textvariable=self.arq_fecha_var, font=FONT_ENTRY, width=12,
                 bg=C["entry_bg"], fg=C["accent2"], insertbackground=C["accent2"],
                 relief="flat", bd=4)
        arq_fecha_entry.pack(side="left", padx=8)
        self._bind_undo_redo(arq_fecha_entry, self.arq_fecha_var)
        self.make_button(date_row, "Cargar", self._load_arqueo_by_date,
                          bg=C["accent"], font=FONT_SMALL, side="left")
        self.make_button(date_row, "💾 Guardar conteo", self._save_conteo_arqueo,
                          bg=C["panel"], fg=C["accent2"], font=FONT_SMALL, side="left", padx_pack=(8,0))
        self.arq_estado_lbl = tk.Label(date_row, text="", font=FONT_SMALL,
                                        bg=C["bg"], fg=C["muted"])
        self.arq_estado_lbl.pack(side="left", padx=12)

        # Contenedor con scroll: todas las secciones (apertura/cierre, diferencia,
        # teórico, real, comparación, notas, botones) viven dentro de este canvas
        # para que nunca queden ocultas por falta de espacio vertical.
        arq_canvas = tk.Canvas(parent, bg=C["bg"], highlightthickness=0)
        arq_scrollbar = tk.Scrollbar(parent, orient="vertical", command=arq_canvas.yview)
        arq_scroll_frame = tk.Frame(arq_canvas, bg=C["bg"])
        arq_scroll_frame.bind("<Configure>",
            lambda e: arq_canvas.configure(scrollregion=arq_canvas.bbox("all")))
        arq_canvas_window = arq_canvas.create_window((0,0), window=arq_scroll_frame, anchor="nw")
        arq_canvas.configure(yscrollcommand=arq_scrollbar.set)
        arq_canvas.pack(side="left", fill="both", expand=True, padx=(20,0), pady=8)
        arq_scrollbar.pack(side="right", fill="y", pady=8)

        def _arq_on_canvas_configure(event):
            arq_canvas.itemconfig(arq_canvas_window, width=event.width)
        arq_canvas.bind("<Configure>", _arq_on_canvas_configure)

        self._bind_canvas_scroll(arq_canvas)

        parent = arq_scroll_frame  # todo lo siguiente se construye dentro del área con scroll

        # Contenedor de dos columnas: Apertura | Cierre (ambas editables)
        cols_frame = tk.Frame(parent, bg=C["bg"])
        cols_frame.pack(fill="both", expand=True, padx=20, pady=(4,8))
        cols_frame.grid_columnconfigure(0, weight=1)
        cols_frame.grid_columnconfigure(1, weight=1)

        self.arq_entries = {"apertura": {}, "cierre": {}}

        def solo_enteros(P):
            return P == "" or P.isdigit()
        vcmd_int = (self.register(solo_enteros), "%P")

        def solo_numeros(P):
            return P == "" or P.replace(".", "").replace(",", "").isdigit()
        vcmd_num = (self.register(solo_numeros), "%P")

        def build_columna(parent_col, titulo, prefijo, subtitulo):
            box = tk.Frame(parent_col, bg=C["panel"])
            box.pack(fill="both", expand=True, padx=(0,6) if prefijo == "apertura" else (6,0))
            tk.Label(box, text=titulo, font=(FONT_SANS, 11, "bold"),
                     bg=C["panel"], fg=C["accent"], padx=10).pack(anchor="w", pady=(8,0))
            tk.Label(box, text=subtitulo, font=FONT_SMALL,
                     bg=C["panel"], fg=C["muted"], padx=10).pack(anchor="w", pady=(0,8))

            for clave, valor, etiqueta in DENOMINACIONES:
                row_f = tk.Frame(box, bg=C["card"])
                row_f.pack(fill="x", padx=4, pady=1)
                tk.Label(row_f, text=etiqueta, font=FONT_LABEL, width=14, anchor="w",
                         bg=C["card"], fg=C["text"], padx=8).pack(side="left")
                tk.Label(row_f, text="×", font=FONT_SMALL,
                         bg=C["card"], fg=C["muted"]).pack(side="left")
                var = tk.IntVar(value=0)
                e = tk.Entry(row_f, textvariable=var, font=FONT_ENTRY, width=6,
                             bg=C["entry_bg"], fg=C["accent2"], insertbackground=C["accent2"],
                             relief="flat", bd=4, validate="key", validatecommand=vcmd_int)
                e.pack(side="left", padx=4)
                e.bind("<FocusIn>", lambda ev, en=e: en.select_range(0, "end"))
                e.bind("<KeyRelease>", lambda ev: self._recalcular_arqueo())
                self._bind_undo_redo(e, var)
                subtotal_lbl = tk.Label(row_f, text="€0.00", font=(FONT_MONO, 11),
                                         bg=C["card"], fg=C["muted"], width=10, anchor="e")
                subtotal_lbl.pack(side="left", padx=(8,4))
                self.arq_entries[prefijo][clave] = {"var": var, "valor": valor, "lbl": subtotal_lbl}

            sc_row = tk.Frame(box, bg=C["card"])
            sc_row.pack(fill="x", padx=4, pady=(6,1))
            tk.Label(sc_row, text="2ª caja (caja fuerte)", font=FONT_LABEL, width=18, anchor="w",
                     bg=C["card"], fg=C["accent2"], padx=8).pack(side="left")
            sc_var = tk.StringVar(value="0.00")
            sc_entry = tk.Entry(sc_row, textvariable=sc_var, font=FONT_ENTRY, width=10,
                                 bg=C["entry_bg"], fg=C["accent2"], insertbackground=C["accent2"],
                                 relief="flat", bd=4, validate="key", validatecommand=vcmd_num)
            sc_entry.pack(side="left", padx=4)
            sc_entry.bind("<FocusIn>", lambda ev, en=sc_entry: en.select_range(0, "end"))
            sc_entry.bind("<FocusOut>", lambda ev, v=sc_var: v.set(f"{float(v.get().replace(',','.') or 0):.2f}"))
            sc_entry.bind("<KeyRelease>", lambda ev: self._recalcular_arqueo())
            self._bind_undo_redo(sc_entry, sc_var)
            self.arq_entries[prefijo]["segunda_caja"] = {"var": sc_var}

            total_row = tk.Frame(box, bg=C["panel"])
            total_row.pack(fill="x", padx=4, pady=(8,8))
            tk.Label(total_row, text="TOTAL", font=(FONT_SANS, 11, "bold"),
                     bg=C["panel"], fg=C["text"], padx=8).pack(side="left")
            total_lbl = tk.Label(total_row, text="€0.00", font=(FONT_MONO, 16, "bold"),
                                  bg=C["panel"], fg=C["accent2"])
            total_lbl.pack(side="left", padx=8)
            self.arq_entries[prefijo]["total_lbl"] = total_lbl

        col_ap = tk.Frame(cols_frame, bg=C["bg"])
        col_ap.grid(row=0, column=0, sticky="nsew")
        col_ci = tk.Frame(cols_frame, bg=C["bg"])
        col_ci.grid(row=0, column=1, sticky="nsew")

        build_columna(col_ap, "🌅  Apertura", "apertura",
                       "Se autocompleta con el cierre de ayer; edítala si has metido o sacado efectivo")
        build_columna(col_ci, "🌙  Cierre", "cierre",
                       "Recuento real de caja, antes de retirar nada")

        # ── Diferencia dinámica + botón Fijar ──
        dif_frame = tk.Frame(parent, bg=C["card"], padx=16, pady=12)
        dif_frame.pack(fill="x", padx=20, pady=(0,8))
        fila_dif = tk.Frame(dif_frame, bg=C["card"])
        fila_dif.pack(fill="x")
        tk.Label(fila_dif, text="Diferencia de caja (Cierre contado − Apertura):", font=FONT_LABEL,
                 bg=C["card"], fg=C["muted"]).pack(side="left")
        self.arq_diferencia_lbl = tk.Label(fila_dif, text="€0.00", font=(FONT_MONO, 18, "bold"),
                                            bg=C["card"], fg=C["accent2"])
        self.arq_diferencia_lbl.pack(side="left", padx=10)
        self.make_button(fila_dif, "📌 Fijar diferencia", self._fijar_diferencia_arqueo,
                          bg=C["accent"], font=FONT_SMALL, side="left", padx_pack=(12,0))

        fila_dif_fija = tk.Frame(dif_frame, bg=C["card"])
        fila_dif_fija.pack(fill="x", pady=(8,0))
        tk.Label(fila_dif_fija, text="Diferencia inicial registrada (fija, antes de retirar billetes):",
                 font=FONT_LABEL, bg=C["card"], fg=C["muted"]).pack(side="left")
        self.arq_diferencia_inicial_lbl = tk.Label(fila_dif_fija, text="— sin fijar —",
                                                     font=(FONT_MONO, 16, "bold"),
                                                     bg=C["card"], fg=C["text"])
        self.arq_diferencia_inicial_lbl.pack(side="left", padx=10)

        # ── Sección TEÓRICO (automático, editable) ──
        teorico_frame = tk.Frame(parent, bg=C["panel"], padx=16, pady=10)
        teorico_frame.pack(fill="x", padx=20, pady=(0,8))
        tk.Label(teorico_frame, text="📐  Teórico (autocompletado desde el cierre del día, editable)",
                 font=(FONT_SANS, 10, "bold"), bg=C["panel"], fg=C["accent"]).pack(anchor="w")

        teorico_row = tk.Frame(teorico_frame, bg=C["panel"])
        teorico_row.pack(fill="x", pady=(8,0))

        def campo_decimal(parent_row, label, key, store):
            tk.Label(parent_row, text=label, font=FONT_LABEL, width=14, anchor="w",
                     bg=C["panel"], fg=C["text"]).pack(side="left", padx=(0,4))
            var = tk.StringVar(value="0.00")
            e = tk.Entry(parent_row, textvariable=var, font=FONT_ENTRY, width=10,
                         bg=C["entry_bg"], fg=C["accent2"], insertbackground=C["accent2"],
                         relief="flat", bd=4, validate="key", validatecommand=vcmd_num)
            e.pack(side="left", padx=(0,16))
            e.bind("<FocusIn>", lambda ev, en=e: en.select_range(0, "end"))
            e.bind("<FocusOut>", lambda ev, v=var: v.set(f"{float(v.get().replace(',','.') or 0):.2f}"))
            e.bind("<KeyRelease>", lambda ev: self._recalcular_arqueo())
            self._bind_undo_redo(e, var)
            store[key] = var

        self.arq_teorico_vars = {}
        campo_decimal(teorico_row, "Z teórico:", "z", self.arq_teorico_vars)
        campo_decimal(teorico_row, "Banco teórico:", "banco", self.arq_teorico_vars)

        teorico_total_row = tk.Frame(teorico_frame, bg=C["panel"])
        teorico_total_row.pack(fill="x", pady=(6,0))
        tk.Label(teorico_total_row, text="Total teórico:", font=(FONT_SANS, 10, "bold"),
                 bg=C["panel"], fg=C["text"]).pack(side="left")
        self.arq_teorico_total_lbl = tk.Label(teorico_total_row, text="€0.00",
                                               font=(FONT_MONO, 14, "bold"),
                                               bg=C["panel"], fg=C["text"])
        self.arq_teorico_total_lbl.pack(side="left", padx=8)

        # ── Sección REAL (manual, lo que el empleado retira) ──
        real_frame = tk.Frame(parent, bg=C["panel"], padx=16, pady=10)
        real_frame.pack(fill="x", padx=20, pady=(0,8))
        tk.Label(real_frame, text="✋  Real (lo que se ha retirado físicamente de caja, anotado a mano)",
                 font=(FONT_SANS, 10, "bold"), bg=C["panel"], fg=C["accent"]).pack(anchor="w")

        real_row = tk.Frame(real_frame, bg=C["panel"])
        real_row.pack(fill="x", pady=(8,0))

        self.arq_real_vars = {}
        campo_decimal(real_row, "Z real:", "z", self.arq_real_vars)
        campo_decimal(real_row, "Banco real:", "banco", self.arq_real_vars)

        real_total_row = tk.Frame(real_frame, bg=C["panel"])
        real_total_row.pack(fill="x", pady=(6,0))
        tk.Label(real_total_row, text="Efectivo real:", font=(FONT_SANS, 10, "bold"),
                 bg=C["panel"], fg=C["text"]).pack(side="left")
        self.arq_real_total_lbl = tk.Label(real_total_row, text="€0.00",
                                            font=(FONT_MONO, 14, "bold"),
                                            bg=C["panel"], fg=C["text"])
        self.arq_real_total_lbl.pack(side="left", padx=8)

        # ── Comparación final ──
        comp_frame = tk.Frame(parent, bg=C["card"], padx=16, pady=12)
        comp_frame.pack(fill="x", padx=20, pady=(0,8))
        tk.Label(comp_frame, text="Comparación: Diferencia inicial vs. Teórico vs. Real",
                 font=(FONT_SANS, 11, "bold"), bg=C["card"], fg=C["text"]).pack(anchor="w")
        self.arq_comparacion_lbl = tk.Label(comp_frame, text="—", font=(FONT_MONO, 13),
                                             bg=C["card"], fg=C["muted"], justify="left")
        self.arq_comparacion_lbl.pack(anchor="w", pady=(6,0))

        # Notas
        notas_frame = tk.Frame(parent, bg=C["card"])
        notas_frame.pack(fill="x", padx=20, pady=(0,8))
        tk.Label(notas_frame, text="📝  Notas del arqueo", font=(FONT_SANS, 10, "bold"),
                 bg=C["card"], fg=C["accent"], padx=10, pady=6).pack(anchor="w")
        self.arq_notas_text = tk.Text(notas_frame, height=2, font=FONT_LABEL,
                                       bg=C["entry_bg"], fg=C["text"],
                                       insertbackground=C["text"], relief="flat", bd=6, wrap="word",
                                       undo=True, autoseparators=True, maxundo=-1)
        self.arq_notas_text.pack(fill="x", padx=10, pady=(0,8))

        btn_row = tk.Frame(parent, bg=C["bg"])
        btn_row.pack(pady=10)
        self.make_button(btn_row, "✅  Cerrar caja (definitivo)", self._save_arqueo,
                          bg=C["accent"], font=FONT_BUTTON, padx=22, pady=10, side="left", padx_pack=8)
        self.make_button(btn_row, "🗑  Limpiar", self._clear_arqueo,
                          bg=C["panel"], fg=C["muted"], font=FONT_BUTTON, padx=20, pady=10, side="left", padx_pack=8)

        # Aclaración de los botones
        tk.Label(parent,
                 text=("«Guardar conteo» (arriba, junto a Cargar) almacena solo el recuento físico de billetes y "
                       "monedas; el teórico se recalcula al recargar leyendo el cierre diario actualizado.\n"
                       "«Cerrar caja» es el cierre definitivo: congela todos los valores cuando la caja ya está cuadrada."),
                 font=FONT_SMALL, bg=C["bg"], fg=C["muted"], justify="left", wraplength=900).pack(anchor="w", padx=20, pady=(0,10))

        self._load_arqueo_by_date()

    def _recalcular_arqueo(self):
        for prefijo in ("apertura", "cierre"):
            cantidades = {}
            for clave, valor, _ in DENOMINACIONES:
                entry = self.arq_entries[prefijo][clave]
                try:
                    cantidad = int(entry["var"].get() or 0)
                except (tk.TclError, ValueError):
                    cantidad = 0
                cantidades[clave] = cantidad
                subtotal = cantidad * valor
                entry["lbl"].config(text=f"€{subtotal:.2f}")
            try:
                segunda_caja = float(str(self.arq_entries[prefijo]["segunda_caja"]["var"].get()).replace(",", ".") or 0)
            except ValueError:
                segunda_caja = 0.0
            total = calcular_total_denominaciones(cantidades, segunda_caja)
            self.arq_entries[prefijo]["total_lbl"].config(text=f"€{total:.2f}")
            self.arq_entries[prefijo]["_cantidades"] = cantidades
            self.arq_entries[prefijo]["_segunda_caja"] = segunda_caja
            self.arq_entries[prefijo]["_total"] = total

        diferencia = self.arq_entries["cierre"]["_total"] - self.arq_entries["apertura"]["_total"]
        self.arq_diferencia_lbl.config(text=f"€{diferencia:,.2f}")
        self._arq_diferencia_dinamica = diferencia

        try:
            z_teo = float(self.arq_teorico_vars["z"].get().replace(",", ".") or 0)
        except ValueError:
            z_teo = 0.0
        try:
            banco_teo = float(self.arq_teorico_vars["banco"].get().replace(",", ".") or 0)
        except ValueError:
            banco_teo = 0.0
        total_teorico = z_teo + banco_teo
        self.arq_teorico_total_lbl.config(text=f"€{total_teorico:,.2f}")

        try:
            z_real = float(self.arq_real_vars["z"].get().replace(",", ".") or 0)
        except ValueError:
            z_real = 0.0
        try:
            banco_real = float(self.arq_real_vars["banco"].get().replace(",", ".") or 0)
        except ValueError:
            banco_real = 0.0
        total_real = z_real + banco_real
        self.arq_real_total_lbl.config(text=f"€{total_real:,.2f}")

        dif_inicial = getattr(self, "_arq_diferencia_inicial", None)
        if dif_inicial is not None:
            partes = [
                f"Contado físicamente (dif. inicial): €{dif_inicial:,.2f}",
                f"Teórico (Z+Banco):                  €{total_teorico:,.2f}",
                f"Real anotado a mano (Z+Banco):       €{total_real:,.2f}",
            ]
            # El descuadre se calcula SIEMPRE contra el teórico (lo que dice el
            # Cierre Diario que debería haber), nunca contra el "real" anotado a
            # mano — ese campo lo rellena el empleado y sería manipulable.
            # Positivo -> ha sobrado dinero en caja respecto al teórico.
            # Negativo -> ha faltado dinero en caja respecto al teórico.
            descuadre = dif_inicial - total_teorico
            z_real_correcta = z_teo + descuadre
            if abs(descuadre) < 0.01:
                partes.append("✅ Lo contado coincide con el teórico — caja cuadrada")
                color = C["success"]
            elif descuadre > 0:
                partes.append(f"⬆️ SOBRAN €{descuadre:,.2f} en caja respecto al teórico.")
                partes.append(
                    f"   → Suma €{descuadre:,.2f} a la Z (Z real correcta ≈ €{z_real_correcta:,.2f})."
                )
                color = C["danger"]
            else:
                falta = abs(descuadre)
                partes.append(f"⬇️ FALTAN €{falta:,.2f} en caja respecto al teórico.")
                partes.append(
                    f"   → Resta €{falta:,.2f} de la Z (Z real correcta ≈ €{z_real_correcta:,.2f})."
                )
                color = C["danger"]

            # Verificación anti-manipulación: lo que se anote a mano como Real
            # (Z+Banco) debería terminar coincidiendo con lo contado físicamente,
            # una vez aplicado el ajuste anterior. Si no coincide, alguien ha
            # escrito un importe distinto al que corresponde.
            if abs(total_real - dif_inicial) > 0.01 and (z_real != 0 or banco_real != 0):
                partes.append(
                    f"⚠️ El Real anotado (€{total_real:,.2f}) no coincide con lo contado "
                    f"(€{dif_inicial:,.2f}) tras el ajuste — revisa la Z real."
                )
                color = C["danger"]
            self.arq_comparacion_lbl.config(text="\n".join(partes), fg=color)
        else:
            self.arq_comparacion_lbl.config(
                text="Pulsa \"Fijar diferencia\" tras contar el cierre, antes de retirar billetes.",
                fg=C["muted"]
            )

    def _fijar_diferencia_arqueo(self):
        self._recalcular_arqueo()
        self._arq_diferencia_inicial = self._arq_diferencia_dinamica
        self.arq_diferencia_inicial_lbl.config(text=f"€{self._arq_diferencia_inicial:,.2f}")
        self._recalcular_arqueo()
        self.set_status("Diferencia inicial fijada — ya puedes retirar el efectivo", C["success"])

    def _clear_arqueo(self):
        for prefijo in ("apertura", "cierre"):
            for clave, _, _ in DENOMINACIONES:
                self.arq_entries[prefijo][clave]["var"].set(0)
            self.arq_entries[prefijo]["segunda_caja"]["var"].set("0.00")
        self.arq_teorico_vars["z"].set("0.00")
        self.arq_teorico_vars["banco"].set("0.00")
        self.arq_real_vars["z"].set("0.00")
        self.arq_real_vars["banco"].set("0.00")
        self.arq_notas_text.delete("1.0", "end")
        self._arq_diferencia_inicial = None
        self.arq_diferencia_inicial_lbl.config(text="— sin fijar —")
        if hasattr(self, "arq_estado_lbl"):
            self.arq_estado_lbl.config(text="", fg=C["muted"])
        self._recalcular_arqueo()

    def _load_arqueo_by_date(self):
        try:
            fecha = date.fromisoformat(self.arq_fecha_var.get().strip())
        except:
            messagebox.showerror("Fecha inválida", "Usa el formato AAAA-MM-DD")
            return

        if not self._pedir_password_fecha(fecha):
            self.arq_fecha_var.set(date.today().isoformat())
            return

        self._clear_arqueo()

        try:
            datos = load_arqueo(fecha)
        except Exception as e:
            self.set_status(f"Error al cargar arqueo: {e}", C["danger"])
            return

        if datos:
            # Ya existe un registro de arqueo para este día.
            cerrado = bool(datos.get("cerrado"))

            # CIERRE: siempre se carga tal cual se guardó — es el conteo físico
            # real que se ha hecho de los billetes/monedas al final del día.
            for clave, _, _ in DENOMINACIONES:
                self.arq_entries["cierre"][clave]["var"].set(datos["cierre"].get(clave, 0))
            self.arq_entries["cierre"]["segunda_caja"]["var"].set(
                f"{datos['cierre'].get('segunda_caja', 0):.2f}"
            )

            # APERTURA: NO es un dato que se introduzca a mano, es el cierre de
            # AYER heredado (cuando se sacaron los billetes y se dejó la caja a
            # cero). Por eso, mientras el día no esté cerrado definitivamente,
            # se relee siempre del último arqueo anterior — así si corriges el
            # cierre de ayer, hoy se actualiza solo en vez de quedarse pegado al
            # valor que se guardó la primera vez que se tocó esta fecha.
            # APERTURA: es el cierre de AYER heredado, PERO si hoy ya has guardado
            # un arqueo con la apertura modificada (p.ej. cambiaste billetes por
            # monedas para el turno, o metiste/sacaste cambio), se respeta ESA
            # apertura guardada. Solo si no hay apertura propia guardada se
            # autocompleta con el cierre del día anterior.
            if cerrado:
                for clave, _, _ in DENOMINACIONES:
                    self.arq_entries["apertura"][clave]["var"].set(datos["apertura"].get(clave, 0))
                self.arq_entries["apertura"]["segunda_caja"]["var"].set(
                    f"{datos['apertura'].get('segunda_caja', 0):.2f}"
                )
            else:
                # ¿la apertura guardada de hoy tiene algún valor? Si sí, la
                # respetamos (es la que el usuario dejó preparada para el turno).
                ap_guardada = datos.get("apertura", {}) or {}
                tiene_apertura_propia = (
                    calcular_total_denominaciones(ap_guardada, ap_guardada.get("segunda_caja", 0)) > 0
                )
                if tiene_apertura_propia:
                    origen_apertura = ap_guardada
                else:
                    try:
                        cierre_anterior = get_cierre_arqueo_anterior(fecha)
                    except Exception:
                        cierre_anterior = None
                    origen_apertura = cierre_anterior if cierre_anterior else ap_guardada
                for clave, _, _ in DENOMINACIONES:
                    self.arq_entries["apertura"][clave]["var"].set(origen_apertura.get(clave, 0))
                self.arq_entries["apertura"]["segunda_caja"]["var"].set(
                    f"{origen_apertura.get('segunda_caja', 0):.2f}"
                )

            if datos["diferencia_caja_inicial"]:
                self._arq_diferencia_inicial = datos["diferencia_caja_inicial"]
                self.arq_diferencia_inicial_lbl.config(text=f"€{datos['diferencia_caja_inicial']:,.2f}")
            if datos.get("notas"):
                self.arq_notas_text.insert("1.0", datos["notas"])

            if cerrado:
                # CIERRE DEFINITIVO: todos los valores quedaron congelados, se
                # cargan exactamente como se guardaron.
                self.arq_teorico_vars["z"].set(f"{datos['teorico_z']:.2f}")
                self.arq_teorico_vars["banco"].set(f"{datos['teorico_banco']:.2f}")
                self.arq_real_vars["z"].set(f"{datos['real_z']:.2f}")
                self.arq_real_vars["banco"].set(f"{datos['real_banco']:.2f}")
                self.arq_estado_lbl.config(text="🔒 Caja cerrada (definitivo)", fg=C["success"])
                self.set_status(f"Arqueo del {fecha} cargado (cerrado) ✓", C["success"])
            else:
                # GUARDADO PARCIAL (solo conteo): el teórico se RELEE del cierre
                # diario actual, para reflejar cualquier corrección posterior al
                # Cierre Diario. El real se conserva si se guardó, pero el teórico
                # manda y permite cuadrar de nuevo dinámicamente.
                teorico = None
                try:
                    teorico = get_teorico_efectivo(fecha)
                except Exception:
                    teorico = None
                if teorico is not None:
                    z, banco, _ = teorico
                    self.arq_teorico_vars["z"].set(f"{z:.2f}")
                    self.arq_teorico_vars["banco"].set(f"{banco:.2f}")
                else:
                    # Si no hay cierre diario, caer a lo último guardado
                    self.arq_teorico_vars["z"].set(f"{datos['teorico_z']:.2f}")
                    self.arq_teorico_vars["banco"].set(f"{datos['teorico_banco']:.2f}")
                self.arq_real_vars["z"].set(f"{datos['real_z']:.2f}")
                self.arq_real_vars["banco"].set(f"{datos['real_banco']:.2f}")
                self.arq_estado_lbl.config(
                    text="📝 Conteo guardado — teórico dinámico (sin cerrar)", fg=C["accent2"])
                self.set_status(
                    f"Conteo del {fecha} cargado — apertura propia respetada (o heredada si no la tocaste)",
                    C["success"])
        else:
            # No hay arqueo aún para este día: autocompletar la apertura con el
            # cierre del último arqueo guardado anterior (punto de partida editable)
            self.arq_estado_lbl.config(text="", fg=C["muted"])
            try:
                cierre_anterior = get_cierre_arqueo_anterior(fecha)
            except Exception:
                cierre_anterior = None
            if cierre_anterior:
                for clave, _, _ in DENOMINACIONES:
                    self.arq_entries["apertura"][clave]["var"].set(cierre_anterior.get(clave, 0))
                self.arq_entries["apertura"]["segunda_caja"]["var"].set(
                    f"{cierre_anterior.get('segunda_caja', 0):.2f}"
                )

            # Teórico autocompletado desde cierres_caja
            try:
                teorico = get_teorico_efectivo(fecha)
            except Exception:
                teorico = None
            if teorico is not None:
                z, banco, _ = teorico
                self.arq_teorico_vars["z"].set(f"{z:.2f}")
                self.arq_teorico_vars["banco"].set(f"{banco:.2f}")

            if cierre_anterior:
                self.set_status(
                    f"Apertura autocompletada con el cierre anterior — edítala si has metido/sacado efectivo",
                    C["success"]
                )
            else:
                self.set_status(f"No hay arqueo guardado para {fecha} — formulario en blanco", C["muted"])

        self._recalcular_arqueo()

    def _save_conteo_arqueo(self):
        """Guarda SOLO el conteo físico (cerrado=False). El teórico/real quedan
        dinámicos: al recargar se releen del cierre diario actualizado."""
        self._guardar_arqueo(cerrado=False)

    def _save_arqueo(self):
        """Cierre definitivo de caja (cerrado=True): congela todos los valores."""
        try:
            fecha = date.fromisoformat(self.arq_fecha_var.get().strip())
        except:
            messagebox.showerror("Fecha inválida", "Usa el formato AAAA-MM-DD")
            return
        if not messagebox.askyesno(
                "Cerrar caja definitivamente",
                f"¿Cerrar la caja del {fecha} de forma definitiva?\n\n"
                "Se congelarán el teórico y el real tal como están ahora. "
                "Hazlo solo cuando la caja ya esté cuadrada."):
            return
        self._guardar_arqueo(cerrado=True)

    def _guardar_arqueo(self, cerrado: bool):
        try:
            fecha = date.fromisoformat(self.arq_fecha_var.get().strip())
        except:
            messagebox.showerror("Fecha inválida", "Usa el formato AAAA-MM-DD")
            return

        if not self._pedir_password_fecha(fecha):
            return

        self._recalcular_arqueo()
        apertura = dict(self.arq_entries["apertura"]["_cantidades"])
        apertura["segunda_caja"] = self.arq_entries["apertura"]["_segunda_caja"]
        cierre = dict(self.arq_entries["cierre"]["_cantidades"])
        cierre["segunda_caja"] = self.arq_entries["cierre"]["_segunda_caja"]
        notas = self.arq_notas_text.get("1.0", "end").strip()

        diferencia_inicial = getattr(self, "_arq_diferencia_inicial", None) or 0.0

        try:
            z_teo = float(self.arq_teorico_vars["z"].get().replace(",", ".") or 0)
            banco_teo = float(self.arq_teorico_vars["banco"].get().replace(",", ".") or 0)
            z_real = float(self.arq_real_vars["z"].get().replace(",", ".") or 0)
            banco_real = float(self.arq_real_vars["banco"].get().replace(",", ".") or 0)
        except ValueError:
            messagebox.showerror("Error", "Revisa los importes de teórico/real, contienen valores no numéricos.")
            return

        try:
            save_arqueo(fecha, apertura, cierre, diferencia_inicial,
                        z_teo, banco_teo, z_real, banco_real, notas, cerrado=cerrado)
            if cerrado:
                self.arq_estado_lbl.config(text="🔒 Caja cerrada (definitivo)", fg=C["success"])
                self.set_status(f"✅ Caja del {fecha} cerrada definitivamente", C["success"])
                self._sincronizar_nube_async()
                messagebox.showinfo("Caja cerrada", f"Arqueo del {fecha} cerrado definitivamente.")
            else:
                self.arq_estado_lbl.config(
                    text="📝 Conteo guardado — teórico dinámico (sin cerrar)", fg=C["accent2"])
                self.set_status(
                    f"💾 Conteo del {fecha} guardado — el resto se recalculará al recargar",
                    C["success"])
                messagebox.showinfo(
                    "Conteo guardado",
                    f"Se ha guardado solo el conteo físico del {fecha}.\n\n"
                    "Si corriges el Cierre Diario, al volver a Arqueo se recuperará "
                    "este conteo y el teórico se recalculará automáticamente.")
        except Exception as e:
            self.set_status(f"Error al guardar arqueo: {e}", C["danger"])
            messagebox.showerror("Error", f"No se pudo guardar:\n{e}")

    # ── TAB: HISTÓRICO ────────────────────────
    MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

    def _build_tab_historico(self):
        parent = self.tab_historico

        # ── Barra superior: selector de mes y de año ──
        top = tk.Frame(parent, bg=C["bg"])
        top.pack(fill="x", padx=20, pady=(14,8))

        tk.Label(top, text="Periodo:", font=FONT_LABEL,
                 bg=C["bg"], fg=C["muted"]).pack(side="left", padx=(0,8))

        self.mes_var = tk.StringVar(value="Todo el histórico")
        self._meses_map = {"Todo el histórico": (None, None)}
        self.mes_combo = ttk.Combobox(top, textvariable=self.mes_var,
                                       state="readonly", width=28,
                                       font=FONT_LABEL)
        self.mes_combo.pack(side="left")
        self.mes_combo.bind("<<ComboboxSelected>>", self._on_mes_seleccionado)

        tk.Label(top, text="  ó año:", font=FONT_LABEL,
                 bg=C["bg"], fg=C["muted"]).pack(side="left", padx=(12,8))

        self.anio_var = tk.StringVar(value="—")
        self._anios_map = {"—": None}
        self.anio_combo = ttk.Combobox(top, textvariable=self.anio_var,
                                        state="readonly", width=10,
                                        font=FONT_LABEL)
        self.anio_combo.pack(side="left")
        self.anio_combo.bind("<<ComboboxSelected>>", self._on_anio_seleccionado)

        self.make_button(top, "🔄 Actualizar", self._refresh_historico,
                          bg=C["accent"], font=FONT_SMALL, side="right")
        self.make_button(top, "📥 Exportar a Excel", self._exportar_historico_excel,
                          bg=C["panel"], fg=C["text"], font=FONT_SMALL, side="right", padx_pack=(0,8))

        # ── Tarjetas KPI ──
        self.kpi_frame = tk.Frame(parent, bg=C["bg"])
        self.kpi_frame.pack(fill="x", padx=20, pady=(0,10))
        self.kpi_labels = {}
        self.kpi_pct_labels = {}
        self.kpi_var_labels = {}

        kpi_defs = [
            ("facturacion", "💰 Facturación"),
            ("gastos",      "💸 Gastos"),
            ("visa",        "💳 Visa"),
            ("plataformas", "🛵 Glovo+JustEat+Uber"),
            ("banco",       "🏦 Ingreso Banco"),
            ("z",           "🎫 Z"),
            ("total",       "🧾 TOTAL"),
        ]
        for key, label in kpi_defs:
            card = tk.Frame(self.kpi_frame, bg=C["card"], padx=12, pady=10)
            card.pack(side="left", fill="x", expand=True, padx=3)
            tk.Label(card, text=label, font=(FONT_SANS, 9),
                     bg=C["card"], fg=C["muted"]).pack(anchor="w")
            val_lbl = tk.Label(card, text="€0.00", font=(FONT_MONO, 14, "bold"),
                                bg=C["card"], fg=C["accent2"])
            val_lbl.pack(anchor="w", pady=(2,0))

            sub_row = tk.Frame(card, bg=C["card"])
            sub_row.pack(anchor="w", fill="x")
            pct_lbl = tk.Label(sub_row, text="0.0%", font=(FONT_SANS, 9, "bold"),
                                bg=C["card"], fg=C["accent"])
            pct_lbl.pack(side="left")
            var_lbl = tk.Label(sub_row, text="", font=(FONT_SANS, 9, "bold"),
                                bg=C["card"], fg=C["muted"])
            var_lbl.pack(side="left", padx=(6,0))

            self.kpi_labels[key] = val_lbl
            self.kpi_pct_labels[key] = pct_lbl
            self.kpi_var_labels[key] = var_lbl

        # ── Tabla completa ──
        cols = ("fecha","facturacion","gastos","visa","internet","justeat",
                "glovo","uber","ticket","banco","z","total")
        headers = ("Fecha","Facturación","Gastos","Visa","Internet","JustEat",
                   "Glovo","Uber","T.Rest.","Ing.Banco","Z","Total")
        widths  = [95, 95, 75, 75, 80, 75, 75, 75, 75, 90, 75, 90]

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dark.Treeview",
            background=C["card"], foreground=C["text"],
            rowheight=28, fieldbackground=C["card"],
            font=(FONT_SANS, 10))
        style.configure("Dark.Treeview.Heading",
            background=C["panel"], foreground=C["accent"],
            font=(FONT_SANS, 9, "bold"), relief="flat")
        style.map("Dark.Treeview", background=[("selected", C["accent"])])

        frame = tk.Frame(parent, bg=C["bg"])
        frame.pack(fill="both", expand=True, padx=20, pady=(0,16))

        self.tree = ttk.Treeview(frame, columns=cols, show="headings",
                                 style="Dark.Treeview")
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="center")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        self.tree.tag_configure("odd",  background=C["card"])
        self.tree.tag_configure("even", background=C["panel"])

        # Cargar opciones del combo (se hace tras init_db, ver __init__)

    def _populate_meses_combo(self):
        try:
            meses = get_meses_disponibles()
        except Exception as e:
            self.set_status(f"Error al cargar meses: {e}", C["danger"])
            return
        opciones = ["Todo el histórico"]
        self._meses_map = {"Todo el histórico": (None, None)}
        anios_vistos = []
        for anio, mes in meses:
            label = f"{self.MESES_ES[mes-1]} {anio}"
            opciones.append(label)
            self._meses_map[label] = (anio, mes)
            if anio not in anios_vistos:
                anios_vistos.append(anio)
        self.mes_combo["values"] = opciones
        # Por defecto, selecciona el mes en curso si existe en los datos; si no, "Todo el histórico"
        hoy = date.today()
        label_mes_actual = f"{self.MESES_ES[hoy.month-1]} {hoy.year}"
        if label_mes_actual in self._meses_map:
            self.mes_var.set(label_mes_actual)
        elif opciones:
            self.mes_var.set(opciones[0])

        # Combo de años (orden descendente, más reciente primero)
        anios_vistos.sort(reverse=True)
        opciones_anio = ["—"] + [str(a) for a in anios_vistos]
        self._anios_map = {"—": None}
        for a in anios_vistos:
            self._anios_map[str(a)] = a
        self.anio_combo["values"] = opciones_anio
        self.anio_var.set("—")

    def _on_mes_seleccionado(self, event=None):
        """Al elegir un mes concreto, resetea el selector de año."""
        self.anio_var.set("—")
        self._refresh_historico()

    def _on_anio_seleccionado(self, event=None):
        """Al elegir un año, resetea el selector de mes."""
        self.mes_var.set("Todo el histórico")
        self._refresh_historico()

    # ── TAB: PREVISIÓN ─────────────────────────
    EVENTO_ICONOS = {"futbol": "⚽", "motogp": "🏍️", "f1": "🏎️", "otro": "⚠️"}

    def _build_tab_prevision(self):
        parent = self.tab_prevision

        # Selector de fecha a prever
        top = tk.Frame(parent, bg=C["bg"])
        top.pack(fill="x", padx=20, pady=(14,8))
        tk.Label(top, text="Fecha a prever:", font=FONT_LABEL,
                 bg=C["bg"], fg=C["muted"]).pack(side="left")
        self.prev_fecha_var = tk.StringVar(value=date.today().isoformat())
        prev_fecha_entry = tk.Entry(top, textvariable=self.prev_fecha_var, font=FONT_ENTRY, width=12,
                 bg=C["entry_bg"], fg=C["accent2"], insertbackground=C["accent2"],
                 relief="flat", bd=4)
        prev_fecha_entry.pack(side="left", padx=8)
        self._bind_undo_redo(prev_fecha_entry, self.prev_fecha_var)
        self.make_button(top, "Calcular previsión", self._calcular_prevision,
                          bg=C["accent"], font=FONT_SMALL, padx=12, side="left", padx_pack=4)

        # Resultado principal
        self.prev_resultado_frame = tk.Frame(parent, bg=C["card"], padx=20, pady=16)
        self.prev_resultado_frame.pack(fill="x", padx=20, pady=(4,8))

        self.prev_fecha_lbl = tk.Label(self.prev_resultado_frame, text="—",
                                        font=FONT_LABEL, bg=C["card"], fg=C["muted"])
        self.prev_fecha_lbl.pack(anchor="w")

        self.prev_tipo_lbl = tk.Label(self.prev_resultado_frame, text="—",
                                       font=(FONT_SANS, 30, "bold"),
                                       bg=C["card"], fg=C["muted"])
        self.prev_tipo_lbl.pack(anchor="w", pady=(2,8))

        self.prev_valor_lbl = tk.Label(self.prev_resultado_frame, text="€0.00",
                                        font=(FONT_MONO, 30, "bold"),
                                        bg=C["card"], fg=C["accent2"])
        self.prev_valor_lbl.pack(anchor="w")

        # Cifra ajustada por partido gordo entre semana (solo aparece si aplica)
        self.prev_ajuste_lbl = tk.Label(self.prev_resultado_frame, text="",
                                         font=(FONT_SANS, 12, "bold"),
                                         bg=C["card"], fg=C["success"])
        self.prev_ajuste_lbl.pack(anchor="w", pady=(2, 0))

        self.prev_intervalo_lbl = tk.Label(self.prev_resultado_frame, text="",
                                            font=FONT_LABEL, bg=C["card"], fg=C["muted"])
        self.prev_intervalo_lbl.pack(anchor="w", pady=(4,2))

        self.prev_media2_lbl = tk.Label(self.prev_resultado_frame, text="",
                                         font=FONT_LABEL, bg=C["card"], fg=C["muted"])
        self.prev_media2_lbl.pack(anchor="w", pady=(0,10))

        kpi_row = tk.Frame(self.prev_resultado_frame, bg=C["card"])
        kpi_row.pack(anchor="w", fill="x")

        pct_box = tk.Frame(kpi_row, bg=C["panel"], padx=14, pady=8)
        pct_box.pack(side="left", padx=(0,10))
        tk.Label(pct_box, text="Percentil", font=FONT_SMALL,
                 bg=C["panel"], fg=C["muted"]).pack(anchor="w")
        self.prev_percentil_lbl = tk.Label(pct_box, text="—",
                                            font=(FONT_MONO, 18, "bold"),
                                            bg=C["panel"], fg=C["text"])
        self.prev_percentil_lbl.pack(anchor="w")

        clas_box = tk.Frame(kpi_row, bg=C["panel"], padx=14, pady=8)
        clas_box.pack(side="left")
        tk.Label(clas_box, text="Se espera día", font=FONT_SMALL,
                 bg=C["panel"], fg=C["muted"]).pack(anchor="w")
        self.prev_clasificacion_lbl = tk.Label(clas_box, text="—",
                                                font=(FONT_SANS, 16, "bold"),
                                                bg=C["panel"], fg=C["text"])
        self.prev_clasificacion_lbl.pack(anchor="w")

        # Alertas de eventos
        self.prev_alertas_frame = tk.Frame(parent, bg=C["bg"])
        self.prev_alertas_frame.pack(fill="x", padx=20, pady=(0,8))

        # ── Gestión de festivos ──
        fest_section = tk.Frame(parent, bg=C["panel"])
        fest_section.pack(fill="x", padx=20, pady=(10,4))
        tk.Label(fest_section, text="📅  Gestión de Festivos / Vísperas / Fiesta Mayor",
                 font=(FONT_SANS, 10, "bold"), bg=C["panel"], fg=C["accent"],
                 padx=10, pady=6).pack(anchor="w")

        fest_form = tk.Frame(parent, bg=C["card"])
        fest_form.pack(fill="x", padx=20, pady=(0,4))
        tk.Label(fest_form, text="Fecha:", font=FONT_SMALL,
                 bg=C["card"], fg=C["muted"]).pack(side="left", padx=(10,4))
        self.fest_fecha_var = tk.StringVar(value=date.today().isoformat())
        fest_fecha_entry = tk.Entry(fest_form, textvariable=self.fest_fecha_var, font=FONT_SMALL, width=11,
                 bg=C["entry_bg"], fg=C["accent2"], insertbackground=C["accent2"],
                 relief="flat", bd=3)
        fest_fecha_entry.pack(side="left", padx=4)
        self._bind_undo_redo(fest_fecha_entry, self.fest_fecha_var)

        tk.Label(fest_form, text="Tipo:", font=FONT_SMALL,
                 bg=C["card"], fg=C["muted"]).pack(side="left", padx=(10,4))
        self.fest_tipo_var = tk.StringVar(value="festivo")
        ttk.Combobox(fest_form, textvariable=self.fest_tipo_var, state="readonly",
                     width=9, font=FONT_SMALL,
                     values=["festivo", "vispera", "fm"]).pack(side="left", padx=4)

        tk.Label(fest_form, text="Nombre:", font=FONT_SMALL,
                 bg=C["card"], fg=C["muted"]).pack(side="left", padx=(10,4))
        self.fest_nombre_var = tk.StringVar()
        fest_nombre_entry = tk.Entry(fest_form, textvariable=self.fest_nombre_var, font=FONT_SMALL, width=18,
                 bg=C["entry_bg"], fg=C["text"], insertbackground=C["text"],
                 relief="flat", bd=3)
        fest_nombre_entry.pack(side="left", padx=4)
        self._bind_undo_redo(fest_nombre_entry, self.fest_nombre_var)

        self.make_button(fest_form, "➕ Añadir", self._add_festivo,
                          bg=C["accent"], font=FONT_SMALL, side="left", padx_pack=8)
        self.make_button(fest_form, "🗑 Eliminar", self._del_festivo_selected,
                          bg=C["panel"], fg=C["muted"], font=FONT_SMALL, side="left", padx_pack=2)

        fest_list_frame = tk.Frame(parent, bg=C["bg"])
        fest_list_frame.pack(fill="both", expand=False, padx=20, pady=(4,10))
        cols = ("fecha","tipo","nombre")
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Fest.Treeview", background=C["card"], foreground=C["text"],
                         rowheight=24, fieldbackground=C["card"], font=(FONT_SANS, 10))
        self.fest_tree = ttk.Treeview(fest_list_frame, columns=cols, show="headings",
                                       style="Fest.Treeview", height=6)
        for c, h, w in zip(cols, ("Fecha","Tipo","Nombre"), (100,90,250)):
            self.fest_tree.heading(c, text=h)
            self.fest_tree.column(c, width=w, anchor="w")
        self.fest_tree.pack(fill="x")

        self._refresh_festivos_list()

    # ── TAB: CALENDARIO SEMANAL (MAPA DE CALOR) ──
    DIAS_SEMANA_CORTO = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

    # ── TAB: SEGUIMIENTO (previsión acumulada vs real + factor) ──
    def _build_tab_seguimiento(self):
        parent = self.tab_seguimiento

        # Canvas scrollable para que quepa todo (gráfico + calendario + heatmap)
        seg_canvas = tk.Canvas(parent, bg=C["bg"], highlightthickness=0)
        seg_scrollbar = tk.Scrollbar(parent, orient="vertical", command=seg_canvas.yview)
        sframe = tk.Frame(seg_canvas, bg=C["bg"])
        sframe.bind("<Configure>",
                    lambda e: seg_canvas.configure(scrollregion=seg_canvas.bbox("all")))
        self._seg_canvas_window = seg_canvas.create_window((0, 0), window=sframe, anchor="nw")
        # Hacer que el frame interior ocupe todo el ancho del canvas
        seg_canvas.bind("<Configure>",
                        lambda e: seg_canvas.itemconfig(self._seg_canvas_window, width=e.width))
        seg_canvas.configure(yscrollcommand=seg_scrollbar.set)
        seg_canvas.pack(side="left", fill="both", expand=True)
        seg_scrollbar.pack(side="right", fill="y", pady=8)
        self._bind_canvas_scroll(seg_canvas)

        # A partir de aquí, todo cuelga de 'sframe' (el frame scrollable)
        # Cabecera
        top = tk.Frame(sframe, bg=C["bg"])
        top.pack(fill="x", padx=20, pady=(14, 8))
        tk.Label(top, text="Seguimiento: previsión vs facturación real",
                 font=(FONT_SANS, 12, "bold"), bg=C["bg"], fg=C["text"]).pack(side="left")
        self.make_button(top, "🔄 Actualizar", self._refresh_seguimiento,
                          bg=C["accent"], font=FONT_SMALL, padx=12, side="right")

        # Controles: año, mes (o año completo)
        ctrl = tk.Frame(sframe, bg=C["bg"])
        ctrl.pack(fill="x", padx=20, pady=(0, 6))

        hoy = date.today()
        anio_min, anio_max = get_rango_anios_disponibles()
        anios = [str(a) for a in range(anio_min, anio_max + 1)]

        tk.Label(ctrl, text="Año:", font=FONT_LABEL, bg=C["bg"], fg=C["muted"]).pack(side="left", padx=(0, 6))
        self.seg_anio_var = tk.StringVar(value=str(hoy.year))
        cb_anio = ttk.Combobox(ctrl, textvariable=self.seg_anio_var, state="readonly",
                               width=7, font=FONT_LABEL, values=anios)
        cb_anio.pack(side="left")
        cb_anio.bind("<<ComboboxSelected>>", lambda e: self._refresh_seguimiento())

        tk.Label(ctrl, text="Mes:", font=FONT_LABEL, bg=C["bg"], fg=C["muted"]).pack(side="left", padx=(12, 6))
        meses_opts = ["Año completo"] + self.MESES_ES
        self.seg_mes_var = tk.StringVar(value=self.MESES_ES[hoy.month - 1])
        cb_mes = ttk.Combobox(ctrl, textvariable=self.seg_mes_var, state="readonly",
                              width=14, font=FONT_LABEL, values=meses_opts)
        cb_mes.pack(side="left")
        cb_mes.bind("<<ComboboxSelected>>", lambda e: self._refresh_seguimiento())

        # Checkbox: aplicar factor de corrección
        self.seg_factor_var = tk.BooleanVar(value=True)
        chk = tk.Checkbutton(ctrl, text="Aplicar factor de corrección (últimas 8 semanas)",
                             variable=self.seg_factor_var, command=self._refresh_seguimiento,
                             bg=C["bg"], fg=C["text"], selectcolor=C["panel"],
                             activebackground=C["bg"], activeforeground=C["text"],
                             font=FONT_SMALL)
        chk.pack(side="left", padx=(16, 0))

        # Banda de resumen (números grandes)
        self.seg_resumen = tk.Label(sframe, text="", font=(FONT_SANS, 11, "bold"),
                                    bg=C["bg"], fg=C["muted"], anchor="w", justify="left")
        self.seg_resumen.pack(fill="x", padx=20, pady=(2, 6))

        # Contenedor del gráfico (altura fija para que el scroll funcione bien)
        self.seguimiento_container = tk.Frame(sframe, bg=C["bg"], height=420)
        self.seguimiento_container.pack(fill="x", padx=20, pady=(0, 6))
        self.seguimiento_container.pack_propagate(False)

        # Calendario mensual coloreado por facturación diaria
        tk.Label(sframe, text="Calendario del mes — facturación por día",
                 font=(FONT_SANS, 11, "bold"), bg=C["bg"], fg=C["text"]).pack(
                     anchor="w", padx=20, pady=(8, 2))
        self.calendario_container = tk.Frame(sframe, bg=C["bg"], height=340)
        self.calendario_container.pack(fill="x", padx=20, pady=(0, 6))
        self.calendario_container.pack_propagate(False)

        # Mapa de calor de facturación mensual (año × mes)
        tk.Label(sframe, text="Mapa de calor — facturación real por mes",
                 font=(FONT_SANS, 11, "bold"), bg=C["bg"], fg=C["text"]).pack(
                     anchor="w", padx=20, pady=(8, 2))
        self.heatmap_container = tk.Frame(sframe, bg=C["bg"], height=260)
        self.heatmap_container.pack(fill="x", padx=20, pady=(0, 8))
        self.heatmap_container.pack_propagate(False)

        self.seg_status_lbl = tk.Label(sframe, text="", font=FONT_SMALL,
                                       bg=C["bg"], fg=C["muted"], anchor="w")
        self.seg_status_lbl.pack(fill="x", padx=20, pady=(0, 8))

        self._refresh_seguimiento()

    def _refresh_seguimiento(self):
        # Limpiar gráfico anterior
        for w in self.seguimiento_container.winfo_children():
            w.destroy()

        anio = int(self.seg_anio_var.get())
        mes_label = self.seg_mes_var.get()
        usar_factor = self.seg_factor_var.get()

        # Factor de corrección
        fc = calcular_factor_correccion(semanas=8)
        factor = fc["factor"] if usar_factor else 1.0

        card_color = C["panel"]
        fig = plt.Figure(figsize=(11, 5.2), dpi=100, facecolor=C["bg"])
        ax = fig.add_subplot(111)
        ax.set_facecolor(card_color)

        if mes_label == "Año completo":
            self._plot_seguimiento_anio(ax, anio, factor)
        else:
            mes = self.MESES_ES.index(mes_label) + 1
            self._plot_seguimiento_mes(ax, anio, mes, factor)

        for spine in ax.spines.values():
            spine.set_color(C["border"])
        ax.tick_params(colors=C["muted"], labelsize=8)
        ax.grid(True, color=C["border"], linewidth=0.4, alpha=0.5)
        fig.tight_layout(pad=2.0)

        canvas = FigureCanvasTkAgg(fig, master=self.seguimiento_container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

        # Texto del factor en el status
        if fc["n"] >= 5:
            signo = "por debajo" if fc["desvio_pct"] < 0 else "por encima"
            self.seg_status_lbl.config(
                text=(f"Factor de corrección: ×{fc['factor']:.3f}  "
                      f"(en las últimas 8 semanas la venta real va {abs(fc['desvio_pct']):.1f}% "
                      f"{signo} de lo previsto, sobre {fc['n']} días)"),
                fg=C["muted"])
        else:
            self.seg_status_lbl.config(
                text="Factor de corrección: sin datos suficientes en las últimas 8 semanas (×1.000).",
                fg=C["muted"])

        # Redibujar el calendario del mes y el mapa de calor anual
        self._refresh_calendario()
        self._refresh_heatmap()

    def _refresh_calendario(self):
        for w in self.calendario_container.winfo_children():
            w.destroy()

        anio = int(self.seg_anio_var.get())
        mes_label = self.seg_mes_var.get()
        hoy = date.today()
        # Si está en "Año completo", el calendario muestra el mes actual
        if mes_label == "Año completo":
            mes = hoy.month
            anio_cal = hoy.year if anio == hoy.year else anio
        else:
            mes = self.MESES_ES.index(mes_label) + 1
            anio_cal = anio

        try:
            d = datos_calendario_mes(anio_cal, mes)
        except Exception:
            return

        import matplotlib.colors as mcolors
        try:
            cmap = plt.colormaps["YlOrRd"]
        except Exception:
            import matplotlib.cm as cm
            cmap = cm.get_cmap("YlOrRd")

        semanas = d["semanas"]
        n_filas = len(semanas)
        vmin, vmax = d["min"], d["max"]
        norm = mcolors.Normalize(vmin=vmin, vmax=vmax)

        fig = plt.Figure(figsize=(11, 0.7 + 0.85 * n_filas), dpi=100, facecolor=C["bg"])
        ax = fig.add_subplot(111)
        ax.set_facecolor(C["bg"])

        dias_sem = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
        for ci, nombre in enumerate(dias_sem):
            ax.text(ci + 0.5, n_filas + 0.18, nombre, ha="center", va="bottom",
                    fontsize=8, color=C["muted"], fontweight="bold")

        for fi, semana in enumerate(semanas):
            yrow = n_filas - 1 - fi  # primera semana arriba
            for ci, (dia, valor) in enumerate(semana):
                if dia is None:
                    continue
                if valor is None:
                    face = C["panel"]
                    txt_val = "—"
                    txt_color = C["muted"]
                else:
                    face = cmap(norm(valor))
                    lum = 0.299*face[0] + 0.587*face[1] + 0.114*face[2]
                    txt_color = "#000000" if lum > 0.6 else "#ffffff"
                    txt_val = f"{valor:,.0f}€"
                ax.add_patch(plt.Rectangle((ci, yrow), 1, 1, facecolor=face,
                             edgecolor=C["bg"], lw=2))
                # número de día (arriba-izquierda) e importe (centro)
                num_color = txt_color if valor is not None else C["muted"]
                ax.text(ci + 0.08, yrow + 0.92, str(dia), ha="left", va="top",
                        fontsize=8, color=num_color, fontweight="bold")
                ax.text(ci + 0.5, yrow + 0.42, txt_val, ha="center", va="center",
                        fontsize=8.5, color=txt_color)
                # marcar hoy con borde
                if (anio_cal == hoy.year and mes == hoy.month and dia == hoy.day):
                    ax.add_patch(plt.Rectangle((ci, yrow), 1, 1, fill=False,
                                 edgecolor=C["accent"], lw=2.5))

        ax.set_xlim(0, 7)
        ax.set_ylim(0, n_filas + 0.5)
        ax.set_xticks([]); ax.set_yticks([])
        for spine in ax.spines.values():
            spine.set_visible(False)
        ax.set_title(f"{self.MESES_ES[mes-1]} {anio_cal}  ·  total {d['total']:,.0f}€",
                     color=C["text"], fontsize=11, fontweight="bold", loc="left")
        fig.tight_layout(pad=1.2)

        canvas = FigureCanvasTkAgg(fig, master=self.calendario_container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    def _refresh_heatmap(self):
        for w in self.heatmap_container.winfo_children():
            w.destroy()
        try:
            d = datos_heatmap_mensual()
        except Exception:
            return
        if not d["anios"]:
            return

        import matplotlib.colors as mcolors

        anios = d["anios"]
        n_anios = len(anios)
        fig = plt.Figure(figsize=(11, 0.5 + 0.5 * n_anios), dpi=100, facecolor=C["bg"])
        ax = fig.add_subplot(111)
        ax.set_facecolor(C["bg"])

        vmin, vmax = d["min"], d["max"]
        norm = mcolors.Normalize(vmin=vmin, vmax=vmax)
        try:
            cmap = plt.colormaps["YlOrRd"]          # matplotlib moderno (>=3.7)
        except Exception:
            import matplotlib.cm as cm
            cmap = cm.get_cmap("YlOrRd")            # fallback versiones antiguas

        for yi, anio in enumerate(anios):
            for mes in range(1, 13):
                total = d["matriz"].get(anio, {}).get(mes)
                if total is None:
                    ax.add_patch(plt.Rectangle((mes - 1, yi), 1, 1,
                                 facecolor=C["panel"], edgecolor=C["bg"], lw=1.5))
                    continue
                color = cmap(norm(total))
                ax.add_patch(plt.Rectangle((mes - 1, yi), 1, 1,
                             facecolor=color, edgecolor=C["bg"], lw=1.5))
                # Texto del importe (en miles, legible)
                lum = 0.299*color[0] + 0.587*color[1] + 0.114*color[2]
                txt_color = "#000000" if lum > 0.6 else "#ffffff"
                ax.text(mes - 0.5, yi + 0.5, f"{total/1000:.1f}k",
                        ha="center", va="center", fontsize=7.5, color=txt_color)

        ax.set_xlim(0, 12)
        ax.set_ylim(0, n_anios)
        ax.set_xticks([m - 0.5 for m in range(1, 13)])
        ax.set_xticklabels([m[:3] for m in self.MESES_ES], fontsize=8, color=C["muted"])
        ax.set_yticks([yi + 0.5 for yi in range(n_anios)])
        ax.set_yticklabels([str(a) for a in anios], fontsize=8, color=C["muted"])
        ax.invert_yaxis()
        ax.tick_params(length=0)
        for spine in ax.spines.values():
            spine.set_visible(False)
        fig.tight_layout(pad=1.5)

        canvas = FigureCanvasTkAgg(fig, master=self.heatmap_container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    def _plot_seguimiento_mes(self, ax, anio, mes, factor):
        d = datos_mes_acumulado(anio, mes, factor=factor)
        dias = d["dias"]
        # Previsión original (difuminada)
        ax.plot(dias, d["prev_acum"], color="#2D9CDB", lw=1.8, ls="--", alpha=0.45,
                label="Previsión acumulada")
        # Previsión corregida (marcada) si factor != 1
        if abs(factor - 1.0) > 0.001:
            ax.plot(dias, d["prev_corr_acum"], color="#2D9CDB", lw=2.2,
                    label=f"Previsión corregida (×{factor:.3f})")
        # Real acumulada hasta hoy
        dias_r = [x for x, v in zip(dias, d["real_acum"]) if v is not None]
        real_v = [v for v in d["real_acum"] if v is not None]
        ax.plot(dias_r, real_v, color="#FF7F32", lw=2.6, label="Facturación real acumulada")
        # Banda real vs previsión ORIGINAL (base): verde si vas por encima
        if real_v:
            ref = d["prev_acum"][:len(real_v)]
            encima = real_v[-1] >= ref[-1]
            ax.fill_between(dias_r, ref, real_v,
                            color=(C["success"] if encima else C["danger"]), alpha=0.15)

        ax.set_title(f"{self.MESES_ES[mes-1]} {anio} — acumulado diario",
                     color=C["text"], fontsize=12, fontweight="bold", loc="left")
        ax.set_ylabel("€ acumulados", color=C["muted"], fontsize=9)
        ax.xaxis.set_major_locator(mdates.DayLocator(interval=2))
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d"))
        leg = ax.legend(loc="upper left", fontsize=8, facecolor=C["panel"], edgecolor=C["border"])
        for t in leg.get_texts():
            t.set_color(C["text"])

        # Resumen numérico: el veredicto (verde/rojo) compara la real contra la
        # previsión ORIGINAL hasta hoy — "¿voy mejor o peor de lo que tocaría?".
        # (El factor se usa para proyectar el futuro, no para juzgar el presente.)
        r = d["resumen"]
        if r["real_hoy"] is not None and r["prev_hoy"]:
            diff = r["real_hoy"] - r["prev_hoy"]
            pct = 100 * diff / r["prev_hoy"] if r["prev_hoy"] else 0
            estado = "POR ENCIMA" if diff >= 0 else "POR DEBAJO"
            col = C["success"] if diff >= 0 else C["danger"]
            self.seg_resumen.config(
                text=(f"{estado} de lo previsto: {diff:+,.0f}€ ({pct:+.1f}%)   ·   "
                      f"Real: {r['real_hoy']:,.0f}€   Previsto: {r['prev_hoy']:,.0f}€   "
                      f"(a fecha de hoy)"),
                fg=col)
        else:
            self.seg_resumen.config(text="Sin facturación registrada todavía este mes.", fg=C["muted"])

    def _plot_seguimiento_anio(self, ax, anio, factor):
        # Acumulado de todo el año: sumamos mes a mes
        meses_x, prev_y, prev_corr_y, real_y = [], [], [], []
        ap = acp = ar = 0.0
        hoy = date.today()
        hay_real = False
        for mes in range(1, 13):
            d = datos_mes_acumulado(anio, mes, factor=factor, completar_al_vuelo=False)
            # tomar el último valor acumulado de cada mes
            pv = next((v for v in reversed(d["prev_acum"]) if v is not None), 0.0)
            pcv = next((v for v in reversed(d["prev_corr_acum"]) if v is not None), 0.0)
            rv = next((v for v in reversed(d["real_acum"]) if v is not None), None)
            ap += pv; acp += pcv
            meses_x.append(date(anio, mes, 15))
            prev_y.append(ap)
            prev_corr_y.append(acp)
            if rv is not None and (anio < hoy.year or mes <= hoy.month):
                ar += rv
                real_y.append(ar)
                hay_real = True
            else:
                real_y.append(None)

        ax.plot(meses_x, prev_y, color="#2D9CDB", lw=1.8, ls="--", alpha=0.45,
                label="Previsión acumulada")
        if abs(factor - 1.0) > 0.001:
            ax.plot(meses_x, prev_corr_y, color="#2D9CDB", lw=2.2,
                    label=f"Previsión corregida (×{factor:.3f})")
        mx_r = [x for x, v in zip(meses_x, real_y) if v is not None]
        ry = [v for v in real_y if v is not None]
        ax.plot(mx_r, ry, color="#FF7F32", lw=2.6, label="Facturación real acumulada")
        if ry:
            ref = prev_y[:len(ry)]
            encima = ry[-1] >= ref[-1]
            ax.fill_between(mx_r, ref, ry,
                            color=(C["success"] if encima else C["danger"]), alpha=0.15)

        ax.set_title(f"Año {anio} — acumulado mensual",
                     color=C["text"], fontsize=12, fontweight="bold", loc="left")
        ax.set_ylabel("€ acumulados", color=C["muted"], fontsize=9)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b"))
        leg = ax.legend(loc="upper left", fontsize=8, facecolor=C["panel"], edgecolor=C["border"])
        for t in leg.get_texts():
            t.set_color(C["text"])

        if ry and prev_y[:len(ry)]:
            ref = prev_y[len(ry) - 1]
            diff = ry[-1] - ref
            pct = 100 * diff / ref if ref else 0
            estado = "POR ENCIMA" if diff >= 0 else "POR DEBAJO"
            col = C["success"] if diff >= 0 else C["danger"]
            self.seg_resumen.config(
                text=(f"Año {anio}: {estado} de lo previsto: {diff:+,.0f}€ ({pct:+.1f}%)   ·   "
                      f"Real: {ry[-1]:,.0f}€   Previsto: {ref:,.0f}€"),
                fg=col)
        else:
            self.seg_resumen.config(text=f"Sin facturación registrada en {anio}.", fg=C["muted"])

    def _build_tab_mapa(self):
        parent = self.tab_mapa

        top = tk.Frame(parent, bg=C["bg"])
        top.pack(fill="x", padx=20, pady=(14,8))
        tk.Label(top, text="Vista de las 3 próximas semanas — color = volumen de venta previsto",
                 font=FONT_LABEL, bg=C["bg"], fg=C["muted"]).pack(side="left")
        self.make_button(top, "🔄 Actualizar", self._refresh_mapa_calor,
                          bg=C["accent"], font=FONT_SMALL, padx=12, side="right")

        # Leyenda
        leyenda = tk.Frame(parent, bg=C["bg"])
        leyenda.pack(fill="x", padx=20, pady=(0,8))
        tk.Label(leyenda, text="Bajo", font=FONT_SMALL, bg=C["bg"], fg=C["muted"]).pack(side="left")
        for color in ["#1E3A5F", "#2D5F8A", "#4A8FA8", "#8FBC8F", "#E8C547", "#E8954A", "#D9534F"]:
            tk.Frame(leyenda, bg=color, width=24, height=14).pack(side="left", padx=1)
        tk.Label(leyenda, text="Alto", font=FONT_SMALL, bg=C["bg"], fg=C["muted"]).pack(side="left", padx=(4,0))

        # Cabecera de días de la semana
        header_row = tk.Frame(parent, bg=C["bg"])
        header_row.pack(fill="x", padx=20)
        tk.Label(header_row, text="", width=8, bg=C["bg"]).pack(side="left")
        for dia in self.DIAS_SEMANA_CORTO:
            tk.Label(header_row, text=dia, font=(FONT_SANS, 10, "bold"),
                     bg=C["bg"], fg=C["accent"], width=14).pack(side="left", padx=2)

        # Contenedor de las 3 semanas (se rellena dinámicamente)
        self.mapa_semanas_frame = tk.Frame(parent, bg=C["bg"])
        self.mapa_semanas_frame.pack(fill="both", expand=True, padx=20, pady=(4,16))

        self.mapa_status_lbl = tk.Label(parent, text="", font=FONT_SMALL,
                                         bg=C["bg"], fg=C["muted"])
        self.mapa_status_lbl.pack(anchor="w", padx=20, pady=(0,8))

        # ── Facturación por día de la semana (filtrable por mes/año/histórico) ──
        DIAS_SEMANA_LARGO = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
        dow_card = tk.Frame(parent, bg=C["card"], padx=20, pady=16)
        dow_card.pack(fill="x", padx=20, pady=(4,16))

        tk.Label(dow_card, text="📅  Facturación por día de la semana",
                 font=(FONT_SANS, 13, "bold"), bg=C["card"], fg=C["accent"]).pack(anchor="w")
        tk.Label(dow_card,
                 text="¿Qué días facturan más? Compara por mes, por año o con todo el histórico.",
                 font=FONT_SMALL, bg=C["card"], fg=C["muted"]).pack(anchor="w", pady=(0,10))

        dow_filtro = tk.Frame(dow_card, bg=C["card"])
        dow_filtro.pack(fill="x", pady=(0,10))

        tk.Label(dow_filtro, text="Periodo:", font=FONT_LABEL,
                 bg=C["card"], fg=C["muted"]).pack(side="left", padx=(0,8))

        self.dow_mes_var = tk.StringVar(value="Todo el histórico")
        self._dow_meses_map = {"Todo el histórico": (None, None)}
        self.dow_mes_combo = ttk.Combobox(dow_filtro, textvariable=self.dow_mes_var,
                                           state="readonly", width=26, font=FONT_LABEL)
        self.dow_mes_combo.pack(side="left")
        self.dow_mes_combo.bind("<<ComboboxSelected>>", self._on_dow_mes_seleccionado)

        tk.Label(dow_filtro, text="  ó año:", font=FONT_LABEL,
                 bg=C["card"], fg=C["muted"]).pack(side="left", padx=(12,8))

        self.dow_anio_var = tk.StringVar(value="—")
        self._dow_anios_map = {"—": None}
        self.dow_anio_combo = ttk.Combobox(dow_filtro, textvariable=self.dow_anio_var,
                                            state="readonly", width=10, font=FONT_LABEL)
        self.dow_anio_combo.pack(side="left")
        self.dow_anio_combo.bind("<<ComboboxSelected>>", self._on_dow_anio_seleccionado)

        self.dow_total_lbl = tk.Label(dow_filtro, text="Total: €0.00",
                                       font=(FONT_MONO, 13, "bold"), bg=C["card"], fg=C["accent2"])
        self.dow_total_lbl.pack(side="right")

        # Tabla: cabecera
        dow_tabla = tk.Frame(dow_card, bg=C["card"])
        dow_tabla.pack(fill="x")
        header = tk.Frame(dow_tabla, bg=C["panel"])
        header.pack(fill="x", pady=(0,2))
        tk.Label(header, text="Día", font=(FONT_SANS, 10, "bold"), bg=C["panel"], fg=C["muted"],
                 width=12, anchor="w").pack(side="left", padx=8, pady=6)
        tk.Label(header, text="Facturación", font=(FONT_SANS, 10, "bold"), bg=C["panel"], fg=C["muted"],
                 width=16, anchor="e").pack(side="left", padx=8, pady=6)
        tk.Label(header, text="% del periodo", font=(FONT_SANS, 10, "bold"), bg=C["panel"], fg=C["muted"],
                 width=13, anchor="e").pack(side="left", padx=8, pady=6)
        tk.Label(header, text="", bg=C["panel"]).pack(side="left", fill="x", expand=True, padx=(0,12))

        # Tabla: una fila por día, con barra de porcentaje visual
        self.dow_dia_widgets = []
        for i, dia in enumerate(DIAS_SEMANA_LARGO):
            bg_fila = C["bg"] if i % 2 == 0 else C["card"]
            fila = tk.Frame(dow_tabla, bg=bg_fila)
            fila.pack(fill="x")
            tk.Label(fila, text=dia, font=FONT_LABEL, bg=bg_fila, fg=C["text"],
                     width=12, anchor="w").pack(side="left", padx=8, pady=5)
            lbl_total = tk.Label(fila, text="€0.00", font=(FONT_MONO, 12), bg=bg_fila,
                                  fg=C["accent2"], width=16, anchor="e")
            lbl_total.pack(side="left", padx=8, pady=5)
            lbl_pct = tk.Label(fila, text="0.0%", font=(FONT_MONO, 12, "bold"), bg=bg_fila,
                                fg=C["text"], width=13, anchor="e")
            lbl_pct.pack(side="left", padx=8, pady=5)
            barra_bg = tk.Frame(fila, bg=bg_fila, height=16)
            barra_bg.pack(side="left", fill="x", expand=True, padx=(4,12))
            barra = tk.Frame(barra_bg, bg=C["accent"])
            barra.place(relx=0, rely=0, anchor="nw", relwidth=0.0, relheight=1.0)
            self.dow_dia_widgets.append({"total": lbl_total, "pct": lbl_pct,
                                          "barra": barra, "barra_bg": barra_bg})

        self._refresh_mapa_calor()

    def _color_calor(self, valor, vmin, vmax):
        """Interpola un color de frío (azul) a caliente (rojo) según valor en [vmin, vmax]."""
        paleta = ["#1E3A5F", "#2D5F8A", "#4A8FA8", "#8FBC8F", "#E8C547", "#E8954A", "#D9534F"]
        if vmax <= vmin:
            return paleta[len(paleta)//2]
        t = max(0.0, min(1.0, (valor - vmin) / (vmax - vmin)))
        idx = int(t * (len(paleta) - 1))
        return paleta[idx]

    def _refresh_mapa_calor(self):
        for w in self.mapa_semanas_frame.winfo_children():
            w.destroy()
        self.mapa_status_lbl.config(text="Calculando previsión de 21 días…", fg=C["muted"])
        self.update_idletasks()

        hoy = date.today()
        lunes_actual = hoy - timedelta(days=hoy.weekday())  # weekday(): lunes=0

        datos_dias = []  # lista de 21 dicts con fecha, resultado de prever_facturacion
        try:
            for i in range(21):
                fecha = lunes_actual + timedelta(days=i)
                r = prever_facturacion(fecha)
                datos_dias.append({"fecha": fecha, "r": r})
        except Exception as e:
            self.mapa_status_lbl.config(text=f"Error calculando previsión: {e}", fg=C["danger"])
            return

        valores_validos = [d["r"]["valor"] for d in datos_dias if d["r"].get("valor") is not None]
        vmin = min(valores_validos) if valores_validos else 0
        vmax = max(valores_validos) if valores_validos else 1

        # 3 filas (semanas), 7 columnas (lun-dom)
        for semana in range(3):
            week_row = tk.Frame(self.mapa_semanas_frame, bg=C["bg"])
            week_row.pack(fill="x", pady=2)

            lunes_semana = lunes_actual + timedelta(days=semana*7)
            label_semana = f"{lunes_semana.strftime('%d/%m')}"
            tk.Label(week_row, text=label_semana, font=FONT_SMALL, width=8,
                     bg=C["bg"], fg=C["muted"]).pack(side="left")

            for dia_idx in range(7):
                d = datos_dias[semana*7 + dia_idx]
                fecha = d["fecha"]
                r = d["r"]
                valor = r.get("valor")
                tipo = r.get("tipo_dia", "normal")
                eventos = r.get("eventos", [])

                color_fondo = self._color_calor(valor, vmin, vmax) if valor is not None else C["panel"]

                cell = tk.Frame(week_row, bg=color_fondo, width=130, height=70,
                                cursor="hand2", relief="flat", bd=1)
                cell.pack_propagate(False)
                cell.pack(side="left", padx=2)

                texto_fecha = fecha.strftime("%d/%m")
                texto_valor = f"€{valor:,.0f}" if valor is not None else "—"

                lbl_fecha = tk.Label(cell, text=texto_fecha, font=(FONT_SANS, 9, "bold"),
                                      bg=color_fondo, fg="white")
                lbl_fecha.pack(anchor="nw", padx=4, pady=(2,0))

                lbl_valor = tk.Label(cell, text=texto_valor, font=(FONT_MONO, 13, "bold"),
                                      bg=color_fondo, fg="white")
                lbl_valor.pack(anchor="w", padx=4)

                # Iconos de alerta: tipo de día especial + eventos
                iconos = []
                if tipo != "normal":
                    iconos.append(tipo[:4].upper())
                for ev_id, ev_tipo, ev_desc in eventos:
                    iconos.append(self.EVENTO_ICONOS.get(ev_tipo, "⚠️"))
                widgets_celda = [cell, lbl_fecha, lbl_valor]
                if iconos:
                    lbl_alerta = tk.Label(cell, text=" ".join(iconos), font=(FONT_SANS, 8, "bold"),
                                           bg=color_fondo, fg="#1C1917")
                    lbl_alerta.pack(anchor="w", padx=4, pady=(0,2))
                    widgets_celda.append(lbl_alerta)

                # Click en cualquier parte de la celda -> ir a Previsión con esa fecha
                def ir_a_prevision(ev, f=fecha):
                    self.prev_fecha_var.set(f.isoformat())
                    try:
                        self.nb.select(self.tab_prevision)
                    except tk.TclError:
                        pass  # pestaña oculta (p.ej. modo empleado) — no hay nada que seleccionar
                    self._calcular_prevision()

                for widget in widgets_celda:
                    widget.bind("<Button-1>", ir_a_prevision)

        self.mapa_status_lbl.config(text=f"Actualizado · {hoy.strftime('%d/%m/%Y %H:%M')}", fg=C["success"])

    # ── Facturación por día de la semana ──
    def _populate_dow_combo(self):
        try:
            meses = get_meses_disponibles()
        except Exception as e:
            self.set_status(f"Error al cargar meses: {e}", C["danger"])
            return
        opciones = ["Todo el histórico"]
        self._dow_meses_map = {"Todo el histórico": (None, None)}
        anios_vistos = []
        for anio, mes in meses:
            label = f"{self.MESES_ES[mes-1]} {anio}"
            opciones.append(label)
            self._dow_meses_map[label] = (anio, mes)
            if anio not in anios_vistos:
                anios_vistos.append(anio)
        self.dow_mes_combo["values"] = opciones
        self.dow_mes_var.set("Todo el histórico")

        anios_vistos.sort(reverse=True)
        opciones_anio = ["—"] + [str(a) for a in anios_vistos]
        self._dow_anios_map = {"—": None}
        for a in anios_vistos:
            self._dow_anios_map[str(a)] = a
        self.dow_anio_combo["values"] = opciones_anio
        self.dow_anio_var.set("—")

    def _on_dow_mes_seleccionado(self, event=None):
        self.dow_anio_var.set("—")
        self._refresh_dow_tabla()

    def _on_dow_anio_seleccionado(self, event=None):
        self.dow_mes_var.set("Todo el histórico")
        self._refresh_dow_tabla()

    def _restringir_dow_anio_actual(self):
        """En modo empleado, igual que en Histórico: solo el año en curso."""
        anio_actual = date.today().year
        meses_anio_actual = [lbl for lbl, (a, m) in self._dow_meses_map.items() if a == anio_actual]
        orden = [v for v in self.dow_mes_combo["values"] if v in meses_anio_actual]
        self.dow_mes_combo["values"] = orden
        self._dow_meses_map = {k: v for k, v in self._dow_meses_map.items() if k in orden}
        self.dow_anio_combo["values"] = [str(anio_actual)]
        self._dow_anios_map = {str(anio_actual): anio_actual}
        hoy = date.today()
        label_mes_actual = f"{self.MESES_ES[hoy.month-1]} {hoy.year}"
        if label_mes_actual in self._dow_meses_map:
            self.dow_mes_var.set(label_mes_actual)
        elif orden:
            self.dow_mes_var.set(orden[0])
        else:
            self.dow_mes_var.set(str(anio_actual))
            self.dow_anio_var.set(str(anio_actual))

    def _refresh_dow_tabla(self):
        try:
            anio_sel = self._dow_anios_map.get(self.dow_anio_var.get())
            if anio_sel is not None:
                anio, mes = anio_sel, None
            else:
                anio, mes = self._dow_meses_map.get(self.dow_mes_var.get(), (None, None))

            totales, total_general = get_facturacion_por_dia_semana(anio, mes)
            self.dow_total_lbl.config(text=f"Total: €{total_general:,.2f}")

            max_total = max(totales) if totales else 0
            for i, widgets in enumerate(self.dow_dia_widgets):
                total_dia = totales[i]
                pct = (total_dia / total_general * 100) if total_general else 0.0
                frac = (total_dia / max_total) if max_total else 0.0
                widgets["total"].config(text=f"€{total_dia:,.2f}")
                widgets["pct"].config(text=f"{pct:.1f}%")
                widgets["barra"].place(relx=0, rely=0, anchor="nw",
                                        relwidth=max(0.0, min(1.0, frac)), relheight=1.0)
        except Exception as e:
            self.set_status(f"Error al calcular facturación por día de la semana: {e}", C["danger"])

    def _on_tab_changed(self, event=None):
        """Al entrar en la pestaña Previsión, si la fecha sigue siendo 'hoy', recalcula sola."""
        try:
            tab_actual = self.nb.select()
            if tab_actual == str(self.tab_prevision):
                fecha_campo = self.prev_fecha_var.get().strip()
                if fecha_campo == date.today().isoformat():
                    self._calcular_prevision()
        except Exception:
            pass

    def _calcular_prevision(self):
        try:
            fecha = date.fromisoformat(self.prev_fecha_var.get().strip())
        except:
            messagebox.showerror("Fecha inválida", "Usa el formato AAAA-MM-DD")
            return
        try:
            r = prever_facturacion(fecha)
        except Exception as e:
            self.set_status(f"Error en previsión: {e}", C["danger"])
            return

        # Fecha pequeña arriba
        self.prev_fecha_lbl.config(
            text=fecha.strftime("%A %d/%m/%Y").capitalize()
        )

        # Tipo de día en grande: verde si normal, rojo si festivo/vispera/fm
        color_tipo = C["success"] if r["tipo_dia"] == "normal" else C["danger"]
        texto_tipo = r["tipo_dia"].upper()
        nombre_dia = r.get("nombre_dia")
        if nombre_dia:
            texto_tipo += f" — {nombre_dia}"
        self.prev_tipo_lbl.config(text=texto_tipo, fg=color_tipo)

        if r["valor"] is None:
            self.prev_valor_lbl.config(text="Sin datos suficientes")
            self.prev_ajuste_lbl.config(text="")
            self.prev_intervalo_lbl.config(text="")
            self.prev_media2_lbl.config(text="")
            self.prev_percentil_lbl.config(text="—", fg=C["text"])
            self.prev_clasificacion_lbl.config(text="—", fg=C["text"])
        else:
            self.prev_valor_lbl.config(text=f"€{r['valor']:,.2f}")

            # Si hay partidazo (nivel 4-5) entre semana, mostrar TAMBIÉN la cifra
            # ajustada (+55%), conservando siempre la normal arriba.
            if r.get("futbol_ajuste") and r.get("valor_ajustado"):
                self.prev_ajuste_lbl.config(
                    text=(f"⚽ Con partido nivel {r['futbol_nivel']} (entre semana): "
                          f"€{r['valor_ajustado']:,.2f}   (+55%)"),
                    fg=C["success"])
            else:
                self.prev_ajuste_lbl.config(text="")

            if r.get("intervalo_inferior") is not None:
                self.prev_intervalo_lbl.config(
                    text=f"Intervalo: €{r['intervalo_inferior']:,.2f} — €{r['intervalo_superior']:,.2f}"
                )
            else:
                self.prev_intervalo_lbl.config(text="")

            if r.get("media2") is not None:
                self.prev_media2_lbl.config(text=f"Media 2 (días cercanos año pasado): €{r['media2']:,.2f}")
            else:
                self.prev_media2_lbl.config(text="")

            if r.get("percentil") is not None:
                self.prev_percentil_lbl.config(text=f"{r['percentil']*100:.1f}%")
            else:
                self.prev_percentil_lbl.config(text="—")

            clasif = r.get("clasificacion")
            self.prev_clasificacion_lbl.config(text=clasif or "—")

            # Colores por tramo: Flojo=rojo, Normal=amarillo, Potente/Record=verde
            color_map = {
                "Flojo":   C["danger"],
                "Normal":  C["accent2"],
                "Potente": C["success"],
                "Record":  C["success"],
            }
            color_clasif = color_map.get(clasif, C["text"])
            self.prev_percentil_lbl.config(fg=color_clasif)
            self.prev_clasificacion_lbl.config(fg=color_clasif)

        # Alertas de eventos
        for w in self.prev_alertas_frame.winfo_children():
            w.destroy()
        eventos = r.get("eventos", [])
        if eventos:
            es_finde = fecha.isoweekday() in (5, 6, 7)
            for ev_id, tipo, desc in eventos:
                icono = self.EVENTO_ICONOS.get(tipo, "⚠️")
                # Para fútbol con nivel, mensaje según el caso
                if tipo == "futbol" and "nivel=" in (desc or ""):
                    try:
                        nivel = int(desc.split("nivel=")[1].split()[0].strip())
                    except Exception:
                        nivel = 0
                    # parte legible: equipos + hora (extraída de la descripción)
                    partes = [p.strip() for p in desc.split("|")]
                    titulo = partes[0] if partes else desc
                    # buscar el campo que sea una hora HH:MM
                    hora_txt = ""
                    for p in partes:
                        if ":" in p and len(p) <= 5 and p.replace(":", "").isdigit():
                            hora_txt = f" ({p}h)"
                            break
                    titulo = f"{titulo}{hora_txt}"
                    es_provisional = partes[0].strip().lower().startswith("por definir")
                    if es_provisional:
                        # fase importante sin equipos aún: solo aviso, sin nivel
                        comp = partes[1] if len(partes) > 1 else ""
                        stage_up = (partes[2] if len(partes) > 2 else "").upper()
                        fase = _nombre_fase(stage_up)
                        msg = f"⚽ {fase} de {comp}{hora_txt} — equipos por definir (posible partidazo, atento)"
                        color = C["accent2"]
                    elif nivel >= 4 and not es_finde:
                        msg = f"⚽ {titulo} — partido GORDO (nivel {nivel}) entre semana: previsión ajustada +55% ↑"
                        color = C["success"]
                    elif nivel >= 4:
                        msg = f"⚽ {titulo} — partido gordo (nivel {nivel}) en finde: alerta, sin ajuste (el finde ya vende)"
                        color = C["accent2"]
                    else:
                        msg = f"⚽ {titulo} — partido nivel {nivel}: alerta informativa, sin ajuste"
                        color = C["muted"]
                    tk.Label(self.prev_alertas_frame, text=msg, font=FONT_SMALL,
                             bg=C["bg"], fg=color, anchor="w").pack(fill="x", pady=2)
                else:
                    tk.Label(self.prev_alertas_frame,
                             text=f"{icono}  {desc} — la previsión no incluye este factor, ajusta a ojo",
                             font=FONT_SMALL, bg=C["bg"], fg=C["danger"],
                             anchor="w").pack(fill="x", pady=2)

        self.set_status(f"Previsión calculada para {fecha}", C["success"])

    def _add_festivo(self):
        try:
            fecha = date.fromisoformat(self.fest_fecha_var.get().strip())
        except:
            messagebox.showerror("Fecha inválida", "Usa el formato AAAA-MM-DD")
            return
        try:
            save_festivo(fecha, self.fest_tipo_var.get(), self.fest_nombre_var.get().strip())
            self._refresh_festivos_list()
            self.set_status(f"Festivo {fecha} guardado", C["success"])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar: {e}")

    def _del_festivo_selected(self):
        sel = self.fest_tree.selection()
        if not sel:
            return
        fecha_str = self.fest_tree.item(sel[0])["values"][0]
        try:
            delete_festivo(date.fromisoformat(str(fecha_str)))
            self._refresh_festivos_list()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar: {e}")

    def _refresh_festivos_list(self):
        for item in self.fest_tree.get_children():
            self.fest_tree.delete(item)
        try:
            for fecha, tipo, nombre in load_festivos():
                self.fest_tree.insert("", "end", values=(str(fecha), tipo, nombre or ""))
        except Exception as e:
            self.set_status(f"Error al cargar festivos: {e}", C["danger"])

    # ── TAB: CALENDARIO DE EVENTOS ─────────────
    def _build_tab_calendario(self):
        parent = self.tab_calendario

        # Botón scraper de fútbol
        top = tk.Frame(parent, bg=C["bg"])
        top.pack(fill="x", padx=20, pady=(14,8))
        tk.Label(top, text="⚽  Partidos de fútbol", font=(FONT_SANS, 12, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(side="left")
        self.cal_scraper_btn = self.make_button(top, "🔄 Buscar partidos (próximos 21 días)",
                          self._run_scraper_futbol,
                          bg=C["accent"], font=FONT_SMALL, padx=12, side="right")

        self.cal_scraper_status = tk.Label(parent, text="", font=FONT_SMALL,
                                            bg=C["bg"], fg=C["muted"])
        self.cal_scraper_status.pack(anchor="w", padx=20, pady=(0,8))

        # ── Formulario para añadir evento manual (MotoGP, F1, etc.) ──
        ev_section = tk.Frame(parent, bg=C["panel"])
        ev_section.pack(fill="x", padx=20, pady=(6,4))
        tk.Label(ev_section, text="➕  Añadir evento manual (MotoGP, F1, otros)",
                 font=(FONT_SANS, 10, "bold"), bg=C["panel"], fg=C["accent"],
                 padx=10, pady=6).pack(anchor="w")

        ev_form = tk.Frame(parent, bg=C["card"])
        ev_form.pack(fill="x", padx=20, pady=(0,4))
        tk.Label(ev_form, text="Fecha:", font=FONT_SMALL,
                 bg=C["card"], fg=C["muted"]).pack(side="left", padx=(10,4))
        self.ev_fecha_var = tk.StringVar(value=date.today().isoformat())
        ev_fecha_entry = tk.Entry(ev_form, textvariable=self.ev_fecha_var, font=FONT_SMALL, width=11,
                 bg=C["entry_bg"], fg=C["accent2"], insertbackground=C["accent2"],
                 relief="flat", bd=3)
        ev_fecha_entry.pack(side="left", padx=4)
        self._bind_undo_redo(ev_fecha_entry, self.ev_fecha_var)

        tk.Label(ev_form, text="Tipo:", font=FONT_SMALL,
                 bg=C["card"], fg=C["muted"]).pack(side="left", padx=(10,4))
        self.ev_tipo_var = tk.StringVar(value="motogp")
        ttk.Combobox(ev_form, textvariable=self.ev_tipo_var, state="readonly",
                     width=9, font=FONT_SMALL,
                     values=["futbol", "motogp", "f1", "otro"]).pack(side="left", padx=4)

        tk.Label(ev_form, text="Descripción:", font=FONT_SMALL,
                 bg=C["card"], fg=C["muted"]).pack(side="left", padx=(10,4))
        self.ev_desc_var = tk.StringVar()
        ev_desc_entry = tk.Entry(ev_form, textvariable=self.ev_desc_var, font=FONT_SMALL, width=28,
                 bg=C["entry_bg"], fg=C["text"], insertbackground=C["text"],
                 relief="flat", bd=3)
        ev_desc_entry.pack(side="left", padx=4)
        self._bind_undo_redo(ev_desc_entry, self.ev_desc_var)

        self.make_button(ev_form, "➕ Añadir", self._add_evento_manual,
                          bg=C["accent"], font=FONT_SMALL, side="left", padx_pack=8)
        self.make_button(ev_form, "🗑 Eliminar", self._del_evento_selected,
                          bg=C["panel"], fg=C["muted"], font=FONT_SMALL, side="left", padx_pack=2)

        # ── Lista de eventos próximos ──
        ev_list_frame = tk.Frame(parent, bg=C["bg"])
        ev_list_frame.pack(fill="both", expand=True, padx=20, pady=(8,16))
        cols = ("fecha", "tipo", "descripcion")
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Eventos.Treeview", background=C["card"], foreground=C["text"],
                         rowheight=26, fieldbackground=C["card"], font=(FONT_SANS, 10))
        self.ev_tree = ttk.Treeview(ev_list_frame, columns=cols, show="headings",
                                     style="Eventos.Treeview")
        for c, h, w in zip(cols, ("Fecha", "Tipo", "Descripción"), (100, 80, 400)):
            self.ev_tree.heading(c, text=h)
            self.ev_tree.column(c, width=w, anchor="w")
        vsb = ttk.Scrollbar(ev_list_frame, orient="vertical", command=self.ev_tree.yview)
        self.ev_tree.configure(yscrollcommand=vsb.set)
        self.ev_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self._refresh_eventos_list()

    def _run_scraper_futbol(self):
        if getattr(self, "_scraper_en_curso", False):
            return  # evita doble click mientras ya está corriendo
        self._scraper_en_curso = True
        self.cal_scraper_btn.config(bg=C["panel"], fg=C["muted"])
        self.cal_scraper_status.config(text="Iniciando búsqueda…", fg=C["muted"])

        def progress(msg):
            self.after(0, lambda: self.cal_scraper_status.config(text=msg, fg=C["muted"]))

        def trabajo():
            try:
                scrape_partidos_futbol(progress_callback=progress)
                self.after(0, self._on_scraper_done, None)
            except Exception as e:
                self.after(0, self._on_scraper_done, e)

        threading.Thread(target=trabajo, daemon=True).start()

    def _on_scraper_done(self, error):
        self._scraper_en_curso = False
        self.cal_scraper_btn.config(bg=C["accent"], fg="white")
        if error:
            self.cal_scraper_status.config(text=f"Error: {error}", fg=C["danger"])
            self.set_status(f"Error en scraper de fútbol: {error}", C["danger"])
        else:
            self.cal_scraper_status.config(text="✅ Búsqueda completada", fg=C["success"])
            self.set_status("Partidos de fútbol actualizados", C["success"])
        self._refresh_eventos_list()

    def _add_evento_manual(self):
        try:
            fecha = date.fromisoformat(self.ev_fecha_var.get().strip())
        except:
            messagebox.showerror("Fecha inválida", "Usa el formato AAAA-MM-DD")
            return
        desc = self.ev_desc_var.get().strip()
        if not desc:
            messagebox.showerror("Falta descripción", "Escribe una descripción del evento")
            return
        try:
            save_evento(fecha, self.ev_tipo_var.get(), desc)
            self.ev_desc_var.set("")
            self._refresh_eventos_list()
            self.set_status(f"Evento del {fecha} guardado", C["success"])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar: {e}")

    def _del_evento_selected(self):
        sel = self.ev_tree.selection()
        if not sel:
            return
        ev_id = self.ev_tree.item(sel[0])["tags"][0] if self.ev_tree.item(sel[0])["tags"] else None
        if ev_id is None:
            return
        try:
            delete_evento(int(ev_id))
            self._refresh_eventos_list()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar: {e}")

    def _refresh_eventos_list(self):
        for item in self.ev_tree.get_children():
            self.ev_tree.delete(item)
        try:
            hoy = date.today()
            hasta = hoy + timedelta(days=60)
            for ev_id, fecha, tipo, desc in load_eventos_rango(hoy, hasta):
                icono = self.EVENTO_ICONOS.get(tipo, "⚠️")
                self.ev_tree.insert("", "end", tags=(str(ev_id),),
                                     values=(str(fecha), f"{icono} {tipo}", desc))
        except Exception as e:
            self.set_status(f"Error al cargar eventos: {e}", C["danger"])

    # ── TAB: GRÁFICOS ──────────────────────────
    def _build_tab_graficos(self):
        parent = self.tab_graficos

        top = tk.Frame(parent, bg=C["bg"])
        top.pack(fill="x", padx=20, pady=(14,8))
        tk.Label(top, text="Evolución mensual de ventas", font=(FONT_SANS, 12, "bold"),
                 bg=C["bg"], fg=C["text"]).pack(side="left")
        self.make_button(top, "🔄 Actualizar", self._refresh_graficos,
                          bg=C["accent"], font=FONT_SMALL, padx=12, side="right")

        # ── Selector de rango de años ──
        rango_row = tk.Frame(parent, bg=C["bg"])
        rango_row.pack(fill="x", padx=20, pady=(0,8))
        tk.Label(rango_row, text="Rango de años:", font=FONT_LABEL,
                 bg=C["bg"], fg=C["muted"]).pack(side="left", padx=(0,8))

        anio_min, anio_max = get_rango_anios_disponibles()
        anios_opciones = [str(a) for a in range(anio_min, anio_max + 1)]

        self.graf_anio_desde_var = tk.StringVar(value=str(anio_min))
        self.graf_anio_desde_combo = ttk.Combobox(rango_row, textvariable=self.graf_anio_desde_var,
                                                   state="readonly", width=8, font=FONT_LABEL,
                                                   values=anios_opciones)
        self.graf_anio_desde_combo.pack(side="left")

        tk.Label(rango_row, text="—", font=FONT_LABEL,
                 bg=C["bg"], fg=C["muted"]).pack(side="left", padx=6)

        self.graf_anio_hasta_var = tk.StringVar(value=str(anio_max))
        self.graf_anio_hasta_combo = ttk.Combobox(rango_row, textvariable=self.graf_anio_hasta_var,
                                                   state="readonly", width=8, font=FONT_LABEL,
                                                   values=anios_opciones)
        self.graf_anio_hasta_combo.pack(side="left")

        self.graf_anio_desde_combo.bind("<<ComboboxSelected>>", lambda e: self._refresh_graficos())
        self.graf_anio_hasta_combo.bind("<<ComboboxSelected>>", lambda e: self._refresh_graficos())

        self.make_button(rango_row, "Todo el histórico", self._reset_rango_graficos,
                          bg=C["panel"], fg=C["muted"], font=FONT_SMALL, side="left", padx_pack=(12,0))

        self.graficos_tendencia_frame = tk.Frame(parent, bg=C["bg"])
        self.graficos_tendencia_frame.pack(fill="x", padx=20, pady=(0,8))

        self.lbl_tendencia_3meses = tk.Label(self.graficos_tendencia_frame, text="",
                                              font=(FONT_SANS, 10, "bold"),
                                              bg=C["bg"], fg=C["muted"], anchor="w", justify="left")
        self.lbl_tendencia_3meses.pack(anchor="w")

        self.lbl_tendencia_mes_simple = tk.Label(self.graficos_tendencia_frame, text="",
                                                  font=(FONT_SANS, 10, "bold"),
                                                  bg=C["bg"], fg=C["muted"], anchor="w", justify="left")
        self.lbl_tendencia_mes_simple.pack(anchor="w", pady=(2,0))

        self.lbl_impacto_media = tk.Label(self.graficos_tendencia_frame, text="",
                                           font=(FONT_SANS, 10, "bold"),
                                           bg=C["bg"], fg=C["muted"], anchor="w", justify="left")
        self.lbl_impacto_media.pack(anchor="w", pady=(2,0))

        self.graficos_container = tk.Frame(parent, bg=C["bg"])
        self.graficos_container.pack(fill="both", expand=True, padx=12, pady=(0,12))

        self.graficos_status_lbl = tk.Label(parent, text="", font=FONT_SMALL,
                                             bg=C["bg"], fg=C["muted"])
        self.graficos_status_lbl.pack(anchor="w", padx=20, pady=(0,8))

        self._refresh_graficos()

    def _refresh_graficos(self):
        for w in self.graficos_container.winfo_children():
            w.destroy()

        try:
            anio_desde = int(self.graf_anio_desde_var.get())
            anio_hasta = int(self.graf_anio_hasta_var.get())
            if anio_desde > anio_hasta:
                anio_desde, anio_hasta = anio_hasta, anio_desde
            datos = load_evolucion_mensual(anio_desde, anio_hasta)
        except Exception as e:
            self.graficos_status_lbl.config(text=f"Error al cargar datos: {e}", fg=C["danger"])
            return

        if not datos:
            self.graficos_status_lbl.config(text="No hay datos suficientes todavía.", fg=C["muted"])
            return

        meses   = [d[0] for d in datos]
        totales = [float(d[1] or 0) for d in datos]
        justeat = [float(d[2] or 0) for d in datos]
        glovo   = [float(d[3] or 0) for d in datos]
        uber    = [float(d[4] or 0) for d in datos]

        import numpy as np

        # ── Comparativa 1: pendiente de la tendencia global, sin vs con el último mes ──
        if len(totales) >= 3:
            totales_sin_ultimo = totales[:-1]
            x_sin = np.arange(len(totales_sin_ultimo))
            pend_sin_ultimo, _ = np.polyfit(x_sin, totales_sin_ultimo, 1)

            x_con = np.arange(len(totales))
            pend_con_ultimo, _ = np.polyfit(x_con, totales, 1)

            acelerando = pend_con_ultimo > pend_sin_ultimo
            color_tend = C["success"] if acelerando else C["danger"]
            flecha_tend = "▲" if acelerando else "▼"
            mes_nuevo_label = meses[-1].strftime("%b %y")
            self.lbl_tendencia_3meses.config(
                text=f"Pendiente tendencia antes de {mes_nuevo_label}: {pend_sin_ultimo:+,.0f}€/mes  →  "
                     f"con {mes_nuevo_label} incluido: {pend_con_ultimo:+,.0f}€/mes  "
                     f"({flecha_tend} {'Acelerando' if acelerando else 'Frenando'})",
                fg=color_tend
            )
        else:
            self.lbl_tendencia_3meses.config(text="", fg=C["muted"])

        # ── Comparativa 2: último mes cerrado vs el mes anterior a ese ──
        if len(totales) >= 2:
            ultimo_mes = totales[-1]
            mes_anterior = totales[-2]
            if mes_anterior:
                diff_pct = ((ultimo_mes - mes_anterior) / mes_anterior) * 100
                subiendo = ultimo_mes >= mes_anterior
                color_simple = C["success"] if subiendo else C["danger"]
                flecha_simple = "▲" if subiendo else "▼"
                mes_actual_label = meses[-1].strftime("%b %y")
                mes_prev_label = meses[-2].strftime("%b %y")
                self.lbl_tendencia_mes_simple.config(
                    text=f"Último mes ({mes_actual_label}) vs. anterior ({mes_prev_label}): "
                         f"{flecha_simple} {diff_pct:+.1f}%",
                    fg=color_simple
                )
            else:
                self.lbl_tendencia_mes_simple.config(text="", fg=C["muted"])
        else:
            self.lbl_tendencia_mes_simple.config(text="", fg=C["muted"])

        # ── Comparativa 3: impacto del último mes cerrado sobre la media histórica ──
        if len(totales) >= 2:
            media_sin_ultimo = sum(totales[:-1]) / len(totales[:-1])
            media_con_ultimo = sum(totales) / len(totales)
            diff_media = media_con_ultimo - media_sin_ultimo
            sube_media = diff_media >= 0
            color_media = C["success"] if sube_media else C["danger"]
            flecha_media = "▲" if sube_media else "▼"
            mes_nuevo_label = meses[-1].strftime("%b %y")
            self.lbl_impacto_media.config(
                text=f"Media histórica antes de {mes_nuevo_label}: €{media_sin_ultimo:,.0f}  →  "
                     f"con {mes_nuevo_label} incluido: €{media_con_ultimo:,.0f}  "
                     f"({flecha_media} {diff_media:+,.0f}€)",
                fg=color_media
            )
        else:
            self.lbl_impacto_media.config(text="", fg=C["muted"])

        # Estilo oscuro consistente con la app
        plt.style.use("dark_background")
        bg_color = C["bg"]
        card_color = C["card"]

        fig = plt.Figure(figsize=(11, 7.5), dpi=100, facecolor=bg_color)

        # ── Gráfico 1: Evolución de ventas totales ──
        ax1 = fig.add_subplot(211)
        ax1.set_facecolor(card_color)
        ax1.plot(meses, totales, color=C["accent"], linewidth=2.5, marker="o",
                  markersize=5, markerfacecolor=C["accent2"], label="Facturación total", zorder=3)

        # Línea de media (horizontal, de fondo)
        media_total = sum(totales) / len(totales)
        ax1.axhline(media_total, color=C["muted"], linewidth=1.3, linestyle="--",
                     alpha=0.7, label=f"Media (€{media_total:,.0f})", zorder=1)

        # Línea de tendencia lineal (regresión simple sobre el índice de mes)
        import numpy as np
        x_idx = np.arange(len(meses))
        if len(x_idx) >= 2:
            pendiente, intercepto = np.polyfit(x_idx, totales, 1)
            tendencia = pendiente * x_idx + intercepto
            ax1.plot(meses, tendencia, color="#E53935", linewidth=2, linestyle="-",
                      alpha=0.85, label="Tendencia lineal", zorder=2)

        ax1.set_title("Facturación total por mes", color=C["text"], fontsize=12, fontweight="bold", loc="left")
        ax1.tick_params(colors=C["muted"], labelsize=8)
        ax1.xaxis.set_major_formatter(mdates.DateFormatter("%b %y"))
        ax1.grid(True, color=C["border"], linewidth=0.4, alpha=0.5)
        leg1 = ax1.legend(loc="upper left", fontsize=8, facecolor=card_color, edgecolor=C["border"])
        for text in leg1.get_texts():
            text.set_color(C["text"])
        for spine in ax1.spines.values():
            spine.set_color(C["border"])

        fig.autofmt_xdate()

        # ── Gráfico 2: Plataformas de delivery (colores de marca) ──
        ax2 = fig.add_subplot(212)
        ax2.set_facecolor(card_color)
        ax2.plot(meses, justeat, color="#FF8000", linewidth=2, marker="o", markersize=4, label="Just Eat")
        ax2.plot(meses, glovo, color="#FFC244", linewidth=2, marker="D", markersize=4,
                  linestyle="--", label="Glovo")
        ax2.plot(meses, uber, color="#06C167", linewidth=2, marker="^", markersize=4,
                  linestyle=":", label="Uber")
        suma_plataformas = [j+g+u for j, g, u in zip(justeat, glovo, uber)]
        ax2.plot(meses, suma_plataformas, color=C["muted"], linewidth=2, label="Total plataformas")
        ax2.set_title("Evolución por plataforma (Just Eat, Glovo, Uber)", color=C["text"],
                       fontsize=12, fontweight="bold", loc="left")
        ax2.tick_params(colors=C["muted"], labelsize=8)
        ax2.xaxis.set_major_formatter(mdates.DateFormatter("%b %y"))
        ax2.grid(True, color=C["border"], linewidth=0.4, alpha=0.5)
        legend = ax2.legend(loc="upper left", fontsize=8, facecolor=card_color, edgecolor=C["border"])
        for text in legend.get_texts():
            text.set_color(C["text"])
        for spine in ax2.spines.values():
            spine.set_color(C["border"])

        fig.tight_layout(pad=2.5)

        canvas = FigureCanvasTkAgg(fig, master=self.graficos_container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

        self.graficos_status_lbl.config(
            text=f"Actualizado · {len(datos)} meses de datos · {date.today().strftime('%d/%m/%Y %H:%M')}",
            fg=C["success"]
        )

    def _reset_rango_graficos(self):
        anio_min, anio_max = get_rango_anios_disponibles()
        self.graf_anio_desde_var.set(str(anio_min))
        self.graf_anio_hasta_var.set(str(anio_max))
        self._refresh_graficos()

    # ── TAB: KPIs GENERALES ────────────────────
    def _build_tab_kpis(self):
        parent = self.tab_kpis

        top = tk.Frame(parent, bg=C["bg"])
        top.pack(fill="x", padx=20, pady=(14,8))
        tk.Label(top, text="Comparativas clave vs. periodos anteriores",
                 font=(FONT_SANS, 12, "bold"), bg=C["bg"], fg=C["text"]).pack(side="left")
        self.make_button(top, "🔄 Actualizar", self._refresh_kpis,
                          bg=C["accent"], font=FONT_SMALL, padx=12, side="right")

        self.kpis_status_lbl = tk.Label(parent, text="", font=FONT_SMALL,
                                         bg=C["bg"], fg=C["muted"])
        self.kpis_status_lbl.pack(side="bottom", anchor="w", padx=20, pady=(0,8))

        # Canvas scrollable para que quepan todas las tarjetas
        kpi_canvas = tk.Canvas(parent, bg=C["bg"], highlightthickness=0)
        kpi_scrollbar = tk.Scrollbar(parent, orient="vertical", command=kpi_canvas.yview)
        kframe = tk.Frame(kpi_canvas, bg=C["bg"])
        kframe.bind("<Configure>",
                    lambda e: kpi_canvas.configure(scrollregion=kpi_canvas.bbox("all")))
        self._kpi_canvas_window = kpi_canvas.create_window((0, 0), window=kframe, anchor="nw")
        kpi_canvas.bind("<Configure>",
                        lambda e: kpi_canvas.itemconfig(self._kpi_canvas_window, width=e.width))
        kpi_canvas.configure(yscrollcommand=kpi_scrollbar.set)
        kpi_canvas.pack(side="left", fill="both", expand=True, padx=(20, 0), pady=(4, 0))
        kpi_scrollbar.pack(side="right", fill="y", pady=8)
        self._bind_canvas_scroll(kpi_canvas)

        self.kpis_container = tk.Frame(kframe, bg=C["bg"])
        self.kpis_container.pack(fill="both", expand=True, padx=(0, 20), pady=(0, 16))

        self._refresh_kpis()

    def _refresh_kpis(self):
        for w in self.kpis_container.winfo_children():
            w.destroy()

        try:
            kpis = calcular_kpis_generales()
        except Exception as e:
            self.kpis_status_lbl.config(text=f"Error al calcular KPIs: {e}", fg=C["danger"])
            return

        orden = [
            "hoy_vs_semana_pasada",
            "hoy_vs_anio_pasado",
            "semana_vs_semana_pasada",
            "mes_vs_mes_anterior",
            "mes_vs_mismo_mes_anio_pasado",
            "mes_anterior_completo_vs_mismo_mes_anio_pasado",
            "media_diaria_mes",
            "anio_vs_anio_pasado",
        ]

        for key in orden:
            info = kpis.get(key)
            if not info:
                continue
            self._crear_tarjeta_kpi(self.kpis_container, info)

        if kpis.get("mejor_peor_dia_mes"):
            self._crear_tarjeta_mejor_peor_dia(self.kpis_container, kpis["mejor_peor_dia_mes"])

        if kpis.get("cuota_plataformas"):
            self._crear_tarjeta_cuota_plataformas(self.kpis_container, kpis["cuota_plataformas"])

        self.kpis_status_lbl.config(
            text=f"Actualizado · {date.today().strftime('%d/%m/%Y %H:%M')}",
            fg=C["success"]
        )

    def _crear_tarjeta_mejor_peor_dia(self, parent, info):
        card = tk.Frame(parent, bg=C["card"], padx=18, pady=14)
        card.pack(fill="x", pady=5)
        tk.Label(card, text="Mejor y peor día del mes en curso", font=FONT_LABEL,
                 bg=C["card"], fg=C["muted"]).pack(anchor="w")

        row = tk.Frame(card, bg=C["card"])
        row.pack(fill="x", pady=(8,0))

        mejor = info.get("mejor")
        peor = info.get("peor")

        if mejor:
            mejor_box = tk.Frame(row, bg=C["panel"], padx=14, pady=8)
            mejor_box.pack(side="left", fill="x", expand=True, padx=(0,6))
            tk.Label(mejor_box, text="🏆 Mejor", font=FONT_SMALL,
                     bg=C["panel"], fg=C["success"]).pack(anchor="w")
            tk.Label(mejor_box, text=f"€{mejor[1]:,.2f}", font=(FONT_MONO, 16, "bold"),
                     bg=C["panel"], fg=C["text"]).pack(anchor="w")
            tk.Label(mejor_box, text=mejor[0].strftime("%d/%m/%Y"), font=FONT_SMALL,
                     bg=C["panel"], fg=C["muted"]).pack(anchor="w")

        if peor:
            peor_box = tk.Frame(row, bg=C["panel"], padx=14, pady=8)
            peor_box.pack(side="left", fill="x", expand=True, padx=(6,0))
            tk.Label(peor_box, text="📉 Peor", font=FONT_SMALL,
                     bg=C["panel"], fg=C["danger"]).pack(anchor="w")
            tk.Label(peor_box, text=f"€{peor[1]:,.2f}", font=(FONT_MONO, 16, "bold"),
                     bg=C["panel"], fg=C["text"]).pack(anchor="w")
            tk.Label(peor_box, text=peor[0].strftime("%d/%m/%Y"), font=FONT_SMALL,
                     bg=C["panel"], fg=C["muted"]).pack(anchor="w")

    def _crear_tarjeta_cuota_plataformas(self, parent, info):
        card = tk.Frame(parent, bg=C["card"], padx=18, pady=14)
        card.pack(fill="x", pady=5)
        tk.Label(card, text="Cuota por plataforma (mes en curso) y evolución",
                 font=FONT_LABEL, bg=C["card"], fg=C["muted"]).pack(anchor="w")

        nombres = {"justeat": "Just Eat", "glovo": "Glovo", "uber": "Uber"}
        row = tk.Frame(card, bg=C["card"])
        row.pack(fill="x", pady=(8,0))

        for p, datos in info["plataformas"].items():
            box = tk.Frame(row, bg=C["panel"], padx=12, pady=8)
            box.pack(side="left", fill="x", expand=True, padx=3)
            tk.Label(box, text=nombres.get(p, p), font=FONT_SMALL,
                     bg=C["panel"], fg=C["muted"]).pack(anchor="w")
            tk.Label(box, text=f"{datos['cuota']:.1f}%", font=(FONT_MONO, 18, "bold"),
                     bg=C["panel"], fg=C["accent2"]).pack(anchor="w")
            tk.Label(box, text=f"€{datos['actual']:,.2f}", font=FONT_SMALL,
                     bg=C["panel"], fg=C["text"]).pack(anchor="w")
            if datos["pct"] is not None:
                color = C["success"] if datos["pct"] >= 0 else C["danger"]
                flecha = "▲" if datos["pct"] >= 0 else "▼"
                signo = "+" if datos["pct"] >= 0 else ""
                tk.Label(box, text=f"{flecha} {signo}{datos['pct']:.1f}%", font=(FONT_SANS, 11, "bold"),
                         bg=C["panel"], fg=color).pack(anchor="w")
            else:
                tk.Label(box, text="—", font=FONT_SMALL,
                         bg=C["panel"], fg=C["muted"]).pack(anchor="w")

        mejor = info.get("mejor_crecimiento")
        if mejor:
            nombre_mejor = nombres.get(mejor[0], mejor[0])
            signo = "+" if mejor[1] >= 0 else ""
            color = C["success"] if mejor[1] >= 0 else C["danger"]
            tk.Label(card, text=f"📈 La que más crece: {nombre_mejor} ({signo}{mejor[1]:.1f}%)",
                     font=FONT_SMALL, bg=C["card"], fg=color).pack(anchor="w", pady=(8,0))

    def _crear_tarjeta_kpi(self, parent, info):
        card = tk.Frame(parent, bg=C["card"], padx=18, pady=14)
        card.pack(fill="x", pady=5)

        tk.Label(card, text=info["label"], font=FONT_LABEL,
                 bg=C["card"], fg=C["muted"]).pack(anchor="w")

        row = tk.Frame(card, bg=C["card"])
        row.pack(fill="x", pady=(6,0))

        actual = info["actual"]
        anterior = info["anterior"]
        diff = info["diff"]
        pct = info["pct"]

        texto_actual = f"€{actual:,.2f}" if actual is not None else "—"
        tk.Label(row, text=texto_actual, font=(FONT_MONO, 22, "bold"),
                 bg=C["card"], fg=C["accent2"]).pack(side="left")

        if pct is not None:
            color_pct = C["success"] if pct >= 0 else C["danger"]
            signo = "+" if pct >= 0 else ""
            flecha = "▲" if pct >= 0 else "▼"
            tk.Label(row, text=f"  {flecha} {signo}{pct:.1f}%", font=(FONT_SANS, 16, "bold"),
                     bg=C["card"], fg=color_pct).pack(side="left", padx=(10,0))
        else:
            tk.Label(row, text="  Sin datos comparables", font=FONT_SMALL,
                     bg=C["card"], fg=C["muted"]).pack(side="left", padx=(10,0))

        texto_anterior = f"€{anterior:,.2f}" if anterior is not None else "—"
        detalle = f"vs. {texto_anterior}"
        if diff is not None:
            signo_diff = "+" if diff >= 0 else ""
            detalle += f"  ({signo_diff}€{diff:,.2f})"
        tk.Label(card, text=detalle, font=FONT_SMALL,
                 bg=C["card"], fg=C["muted"]).pack(anchor="w", pady=(4,0))

    # ── TAB: RANKING DE MESES ──────────────────
    MEDALLAS = ["🥇", "🥈", "🥉"]

    def _build_tab_ranking(self):
        parent = self.tab_ranking

        top = tk.Frame(parent, bg=C["bg"])
        top.pack(fill="x", padx=20, pady=(14,8))
        self.ranking_titulo_lbl = tk.Label(top, text="Los 10 mejores meses de la historia (por facturación total)",
                 font=(FONT_SANS, 12, "bold"), bg=C["bg"], fg=C["text"])
        self.ranking_titulo_lbl.pack(side="left")
        self.make_button(top, "🔄 Actualizar", self._refresh_ranking,
                          bg=C["accent"], font=FONT_SMALL, padx=12, side="right")

        self.ranking_container = tk.Frame(parent, bg=C["bg"])
        self.ranking_container.pack(fill="both", expand=True, padx=20, pady=(4,16))

        self.ranking_status_lbl = tk.Label(parent, text="", font=FONT_SMALL,
                                            bg=C["bg"], fg=C["muted"])
        self.ranking_status_lbl.pack(anchor="w", padx=20, pady=(0,8))

        self._refresh_ranking()

    def _refresh_ranking(self):
        for w in self.ranking_container.winfo_children():
            w.destroy()

        try:
            total_meses = get_count_meses_disponibles()
            self.ranking_titulo_lbl.config(
                text=f"Los 10 mejores meses de la historia (por facturación total) — de {total_meses} meses disponibles"
            )
            ranking = get_ranking_meses(10)
        except Exception as e:
            self.ranking_status_lbl.config(text=f"Error al calcular ranking: {e}", fg=C["danger"])
            return

        if not ranking:
            self.ranking_status_lbl.config(text="No hay datos suficientes todavía.", fg=C["muted"])
            return

        valor_max = float(ranking[0][1])

        for i, (mes, total) in enumerate(ranking):
            total = float(total)
            row = tk.Frame(self.ranking_container, bg=C["card"], padx=16, pady=10)
            row.pack(fill="x", pady=3)

            posicion = self.MEDALLAS[i] if i < 3 else f"#{i+1}"
            tk.Label(row, text=posicion, font=(FONT_SANS, 16, "bold"),
                     bg=C["card"], fg=C["accent"], width=4).pack(side="left")

            mes_label = f"{self.MESES_ES[mes.month-1]} {mes.year}"
            tk.Label(row, text=mes_label, font=(FONT_SANS, 13, "bold"),
                     bg=C["card"], fg=C["text"], width=18, anchor="w").pack(side="left")

            tk.Label(row, text=f"€{total:,.2f}", font=(FONT_MONO, 14, "bold"),
                     bg=C["card"], fg=C["accent2"], width=14, anchor="w").pack(side="left")

            # Barra visual proporcional al máximo
            barra_frame = tk.Frame(row, bg=C["panel"], height=18)
            barra_frame.pack(side="left", fill="x", expand=True, padx=(10,0))
            ancho_pct = (total / valor_max) if valor_max else 0
            barra = tk.Frame(barra_frame, bg=C["accent"], height=18)
            barra.place(relx=0, rely=0, relwidth=max(ancho_pct, 0.02), relheight=1)

        self.ranking_status_lbl.config(
            text=f"Actualizado · {date.today().strftime('%d/%m/%Y %H:%M')}",
            fg=C["success"]
        )

    # ── STATUS BAR ────────────────────────────
    def _build_status(self):
        self.status_var = tk.StringVar(value="Iniciando…")
        bar = tk.Frame(self, bg=C["panel"], pady=4)
        bar.pack(fill="x", side="bottom")
        self.status_lbl = tk.Label(bar, textvariable=self.status_var,
                                   font=FONT_SMALL, bg=C["panel"], fg=C["muted"],
                                   padx=14)
        self.status_lbl.pack(side="left")

        if SYNC_NUBE_ACTIVADO:
            self.make_button(bar, "☁️ Sincronizar ahora", self._sincronizar_nube_async,
                              bg=C["panel"], fg=C["accent2"], font=FONT_SMALL, side="right", padx_pack=(0,14))
            self.make_button(bar, "☁️⟳ Sincronizar TODO", self._sincronizar_nube_completo_async,
                              bg=C["panel"], fg=C["muted"], font=FONT_SMALL, side="right", padx_pack=(0,6))
            self.sync_nube_lbl = tk.Label(bar, text="☁️ Aún sin sincronizar",
                                           font=FONT_SMALL, bg=C["panel"], fg=C["muted"], padx=8)
            self.sync_nube_lbl.pack(side="right")

    def set_status(self, msg, color=None):
        self.status_var.set(msg)
        if color:
            self.status_lbl.config(fg=color)

    # ── LOGIC ─────────────────────────────────
    def _parse(self, key) -> float:
        try:
            v = self.entries[key].get().strip().replace(",",".")
            return float(v) if v else 0.0
        except:
            return 0.0

    def _set(self, key, val):
        self.entries[key].set(f"{val:.2f}")

    def _recalculate(self):
        fact  = self._parse("facturacion")
        z     = self._parse("z_caja")
        total = fact + z
        self._set("facturacion_total", total)

        visa  = self._parse("visa")
        inet  = self._parse("internet")
        je    = self._parse("justeat")
        gl    = self._parse("glovo")
        ub    = self._parse("uber")
        tr    = self._parse("ticket_restaurant")
        gast  = self._parse("gastos")

        # ingreso banco = facturacion - (todos los pagos no-efectivo) - gastos
        # (Z no entra aquí: es efectivo del TPV, no facturación a depositar)
        banco = fact - visa - inet - je - gl - ub - tr - gast
        self._set("ingreso_banco", banco)

    def _marcar_sin_guardar(self):
        self.cierre_guardado = False
        self._actualizar_estado_cierre()

    def _actualizar_estado_cierre(self, hora_guardado=None):
        """Actualiza el indicador 'CIERRE PENDIENTE' / 'CIERRE COMPLETADO (hh:mm)'."""
        if not hasattr(self, "estado_cierre_lbl"):
            return
        if self.cierre_guardado:
            texto = f"CIERRE COMPLETADO ({hora_guardado})" if hora_guardado else "CIERRE COMPLETADO"
            self.estado_cierre_lbl.config(text=texto, fg=C["success"])
        else:
            self.estado_cierre_lbl.config(text="CIERRE PENDIENTE", fg=C["danger"])

    def _on_close_app(self):
        """
        Al pulsar la X: si el cierre del día no está guardado, avisa. Luego ofrece
        realizar el CIERRE DIARIO, que dispara en cadena:
          1. Guarda el contexto de hoy (estación, cole, lluvia)
          2. Congela la previsión que había para hoy
          3. Sincroniza con la nube (solo los últimos días)
        y finalmente cierra la app.
        """
        if not self.cierre_guardado:
            respuesta = messagebox.askyesno(
                "Cierre sin guardar",
                "El cierre de hoy todavía no se ha guardado.\n\n¿Seguro que quieres salir sin guardar?"
            )
            if not respuesta:
                return  # cancelar cierre de la app

        # Ofrecer el cierre diario (contexto + previsión + sincronización)
        hacer_cierre = messagebox.askyesno(
            "Cierre diario",
            "¿Desea realizar el cierre diario?\n\n"
            "Se guardará el contexto del día (lluvia, estación, cole), se "
            "registrará la previsión, se buscarán los próximos partidos de "
            "fútbol y se sincronizará todo con la nube."
        )
        if hacer_cierre:
            self._ejecutar_cierre_diario()
        else:
            self.destroy()

    def _ejecutar_cierre_diario(self):
        """Ejecuta el cierre diario en un hilo para no congelar la ventana, y al
        terminar cierra la app. Ningún fallo impide el cierre."""
        hoy = date.today()

        # Ventana de espera simple
        try:
            self.set_status("⏳ Realizando cierre diario…", C["muted"])
        except Exception:
            pass

        def trabajo():
            resultados = []
            try:
                ok_ctx = guardar_contexto_dia(hoy)
                resultados.append("contexto ✓" if ok_ctx else "contexto ⚠️")
            except Exception:
                resultados.append("contexto ⚠️")
            try:
                ok_prev = registrar_prevision_dia(hoy)
                resultados.append("previsión ✓" if ok_prev else "previsión ⚠️")
            except Exception:
                resultados.append("previsión ⚠️")
            try:
                partidos = scrape_partidos_futbol()
                resultados.append(f"fútbol ✓ ({len(partidos)} partidos)")
            except Exception:
                resultados.append("fútbol ⚠️")
            try:
                ok_sync, msg_sync = sincronizar_con_nube()
                resultados.append("nube ✓" if ok_sync else f"nube ⚠️ ({msg_sync})")
            except Exception as e:
                resultados.append(f"nube ⚠️ ({e})")
            try:
                ok_mail, msg_mail = enviar_correo_cierre(hoy)
                if ok_mail:
                    resultados.append("correo ✓")
                elif msg_mail == "correo no configurado":
                    pass  # no molestar si no está configurado
                else:
                    resultados.append(f"correo ⚠️ ({msg_mail})")
            except Exception as e:
                resultados.append(f"correo ⚠️ ({e})")
            # Volver al hilo principal para cerrar
            self.after(0, lambda: self._fin_cierre_diario(resultados))

        threading.Thread(target=trabajo, daemon=True).start()

    def _fin_cierre_diario(self, resultados):
        """Muestra un resumen breve y cierra la app."""
        try:
            messagebox.showinfo("Cierre diario", "Cierre diario completado:\n\n" +
                                "\n".join(resultados))
        except Exception:
            pass
        self.destroy()

    # ── SINCRONIZACIÓN CON LA NUBE (no bloqueante) ──
    def _sincronizar_nube_async(self):
        """
        Lanza sincronizar_con_nube() en un hilo aparte para no congelar la
        ventana mientras se espera la red. Si SYNC_NUBE_ACTIVADO es False, no
        hace nada (ni siquiera intenta conectar).
        """
        if not SYNC_NUBE_ACTIVADO:
            return

        def trabajo():
            ok, msg = sincronizar_con_nube()
            self.after(0, lambda: self._on_sync_nube_terminado(ok, msg))

        self.set_status("☁️ Sincronizando con la nube…", C["muted"])
        threading.Thread(target=trabajo, daemon=True).start()

    def _sincronizar_nube_completo_async(self):
        """Sincronización COMPLETA (todo el histórico) — botón de respaldo manual.
        Útil si se sospecha algún hueco por días cerrados sin internet."""
        if not SYNC_NUBE_ACTIVADO:
            return
        if not messagebox.askyesno(
            "Sincronizar todo",
            "Esto sube TODO el histórico a la nube (más lento).\n\n"
            "Úsalo de vez en cuando o si sospechas que falta algún día.\n\n¿Continuar?"
        ):
            return

        def trabajo():
            ok, msg = sincronizar_con_nube(completo=True)
            self.after(0, lambda: self._on_sync_nube_terminado(ok, msg))

        self.set_status("☁️ Sincronizando TODO el histórico…", C["muted"])
        threading.Thread(target=trabajo, daemon=True).start()

    def _on_sync_nube_terminado(self, ok: bool, msg: str):
        color = C["success"] if ok else C["danger"]
        self.set_status(f"☁️ {msg}", color)
        if hasattr(self, "sync_nube_lbl"):
            hora = datetime.now().strftime("%H:%M")
            icono = "✅" if ok else "⚠️"
            self.sync_nube_lbl.config(
                text=f"{icono} Última sincronización: {hora} — {msg}",
                fg=color
            )

    def _pedir_password_fecha(self, fecha) -> bool:
        """
        Pide contraseña para editar una fecha distinta a hoy.
        Devuelve True si se autoriza (o ya estaba autorizada esa fecha), False si no.
        """
        if fecha == date.today():
            return True
        if self.fecha_desbloqueada == fecha:
            return True

        pwd = simpledialog.askstring(
            "Fecha protegida",
            f"La fecha {fecha.strftime('%d/%m/%Y')} no es hoy.\n"
            "Introduce la contraseña para editar cierres pasados o futuros:",
            show="*"
        )
        if pwd == PASSWORD_MASTER:
            self.fecha_desbloqueada = fecha
            return True
        if pwd is not None:  # se escribió algo pero estaba mal
            messagebox.showerror("Contraseña incorrecta", "La contraseña no es correcta.")
        return False

    def _clear(self):
        for key, var in self.entries.items():
            var.set("0.00")
        self.notas_text.delete("1.0", "end")
        self.fecha_var.set(date.today().isoformat())
        self.tipo_dia_var.set("normal")

    def _load_today(self):
        self.fecha_var.set(date.today().isoformat())
        self._load_by_date()

    def _load_by_date(self):
        try:
            fecha = date.fromisoformat(self.fecha_var.get().strip())
        except:
            messagebox.showerror("Fecha inválida", "Usa el formato AAAA-MM-DD")
            return

        if not self._pedir_password_fecha(fecha):
            self.fecha_var.set(date.today().isoformat())
            return

        try:
            row = load_cierre(fecha)
        except Exception as e:
            self.set_status(f"Error al cargar: {e}", C["danger"])
            return
        if row:
            keys = ["facturacion","z_caja","facturacion_total",
                    "visa","internet","justeat","glovo","uber",
                    "ticket_restaurant","gastos","ingreso_banco"]
            for k, v in zip(keys, row[:11]):
                self.entries[k].set(f"{float(v or 0):.2f}")
            self.notas_text.delete("1.0","end")
            if row[11]:
                self.notas_text.insert("1.0", row[11])
            self.tipo_dia_var.set(row[12] if len(row) > 12 and row[12] else "normal")
            self.set_status(f"Cierre del {fecha} cargado ✓", C["success"])
            self.cierre_guardado = True
            self._actualizar_estado_cierre()
        else:
            self._clear()
            self.fecha_var.set(fecha.isoformat())
            try:
                self.tipo_dia_var.set(get_tipo_dia(fecha))
            except Exception:
                pass

            mensaje_estado = f"No hay cierre para {fecha} — formulario en blanco"
            try:
                efectivo_tpv = get_efectivo_tpv(fecha)
                if efectivo_tpv is not None:
                    self.entries["z_caja"].set(f"{efectivo_tpv:.2f}")
                    self._recalculate()
                    mensaje_estado = f"Z autocompletado desde el TPV: €{efectivo_tpv:.2f} (puedes editarlo)"
            except Exception:
                pass  # si pagaments no existe o falla, simplemente no autocompleta

            self.set_status(mensaje_estado, C["success"] if "autocompletado" in mensaje_estado else C["muted"])
            self.cierre_guardado = False
            self._actualizar_estado_cierre()

    def _save(self):
        try:
            fecha = date.fromisoformat(self.fecha_var.get().strip())
        except:
            messagebox.showerror("Fecha inválida", "Usa el formato AAAA-MM-DD")
            return

        if not self._pedir_password_fecha(fecha):
            return

        self._recalculate()
        data = {
            "id":                 fecha.strftime("%Y%m%d"),
            "fecha":              fecha,
            "facturacion":        self._parse("facturacion"),
            "gastos":             self._parse("gastos"),
            "visa":               self._parse("visa"),
            "internet":           self._parse("internet"),
            "justeat":            self._parse("justeat"),
            "glovo":              self._parse("glovo"),
            "uber":               self._parse("uber"),
            "ticket_restaurant":  self._parse("ticket_restaurant"),
            "ingreso_banco":      self._parse("ingreso_banco"),
            "z_caja":             self._parse("z_caja"),
            "facturacion_total":  self._parse("facturacion_total"),
            "notas":              self.notas_text.get("1.0","end").strip(),
            "tipo_dia":           self.tipo_dia_var.get(),
        }
        try:
            save_cierre(data)
            hora_actual = datetime.now().strftime("%H:%M")
            self.set_status(f"✅ Cierre del {fecha} guardado correctamente", C["success"])
            self.cierre_guardado = True
            self._actualizar_estado_cierre(hora_guardado=hora_actual)
            self._refresh_historico()
            self._sincronizar_nube_async()
            messagebox.showinfo("Guardado", f"Cierre del {fecha} guardado en la base de datos.")
        except Exception as e:
            self.set_status(f"Error al guardar: {e}", C["danger"])
            messagebox.showerror("Error", f"No se pudo guardar:\n{e}")

    def _refresh_historico(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        try:
            anio_sel = self._anios_map.get(self.anio_var.get())
            if anio_sel is not None:
                anio, mes = anio_sel, None
            else:
                anio, mes = self._meses_map.get(self.mes_var.get(), (None, None))

            rows = load_historico(anio, mes)

            tot_fact = tot_gastos = tot_visa = tot_inet = 0.0
            tot_je = tot_glovo = tot_uber = tot_tr = 0.0
            tot_banco = tot_z = tot_total = 0.0

            for i, row in enumerate(rows):
                (fecha, fact, gastos, visa, inet, je, glovo, uber,
                 tr, banco, z, total) = row

                fact   = float(fact or 0);   gastos = float(gastos or 0)
                visa   = float(visa or 0);   inet   = float(inet or 0)
                je     = float(je or 0);     glovo  = float(glovo or 0)
                uber   = float(uber or 0);   tr     = float(tr or 0)
                banco  = float(banco or 0);  z      = float(z or 0)
                total  = float(total or 0)

                tot_fact += fact; tot_gastos += gastos; tot_visa += visa
                tot_inet += inet; tot_je += je; tot_glovo += glovo
                tot_uber += uber; tot_tr += tr; tot_banco += banco
                tot_z += z; tot_total += total

                tag = "odd" if i % 2 == 0 else "even"
                self.tree.insert("", "end", tag=tag, values=(
                    str(fecha), f"€{fact:.2f}", f"€{gastos:.2f}", f"€{visa:.2f}",
                    f"€{inet:.2f}", f"€{je:.2f}", f"€{glovo:.2f}", f"€{uber:.2f}",
                    f"€{tr:.2f}", f"€{banco:.2f}", f"€{z:.2f}", f"€{total:.2f}",
                ))

            tot_plataformas = tot_glovo + tot_je + tot_uber
            base = tot_total if tot_total else 1  # evita división por cero

            valores = {
                "facturacion": tot_fact,
                "gastos":      tot_gastos,
                "visa":        tot_visa,
                "plataformas": tot_plataformas,
                "banco":       tot_banco,
                "z":           tot_z,
                "total":       tot_total,
            }

            # Variación respecto al periodo anterior:
            # - Si hay mes concreto: compara con el mes anterior
            # - Si hay año completo: compara con el año anterior completo
            valores_periodo_anterior = None
            if anio is not None and mes is not None:
                if mes == 1:
                    anio_ant, mes_ant = anio - 1, 12
                else:
                    anio_ant, mes_ant = anio, mes - 1
                valores_periodo_anterior = self._sumar_historico(anio_ant, mes_ant)
            elif anio is not None and mes is None:
                valores_periodo_anterior = self._sumar_historico(anio - 1, None)

            for key, val in valores.items():
                self.kpi_labels[key].config(text=f"€{val:,.2f}")
                pct = (val / base) * 100
                self.kpi_pct_labels[key].config(text=f"{pct:.1f}%")

                var_lbl = self.kpi_var_labels[key]
                if valores_periodo_anterior is not None:
                    anterior = valores_periodo_anterior.get(key, 0.0)
                    if anterior:
                        var_pct = ((val - anterior) / anterior) * 100
                        color = C["success"] if var_pct >= 0 else C["danger"]
                        flecha = "▲" if var_pct >= 0 else "▼"
                        signo = "+" if var_pct >= 0 else ""
                        var_lbl.config(text=f"{flecha} {signo}{var_pct:.1f}%", fg=color)
                    else:
                        var_lbl.config(text="—", fg=C["muted"])
                else:
                    var_lbl.config(text="", fg=C["muted"])

            self.set_status(f"Histórico actualizado — {len(rows)} días", C["success"])
        except Exception as e:
            self.set_status(f"Error al cargar histórico: {e}", C["danger"])

    def _sumar_historico(self, anio, mes=None) -> dict:
        """Suma los KPIs de un periodo (mes concreto o año completo si mes es None)."""
        rows = load_historico(anio, mes)
        tot_fact = tot_gastos = tot_visa = 0.0
        tot_je = tot_glovo = tot_uber = 0.0
        tot_banco = tot_z = tot_total = 0.0
        for row in rows:
            (fecha, fact, gastos, visa, inet, je, glovo, uber,
             tr, banco, z, total) = row
            tot_fact   += float(fact or 0)
            tot_gastos += float(gastos or 0)
            tot_visa   += float(visa or 0)
            tot_je     += float(je or 0)
            tot_glovo  += float(glovo or 0)
            tot_uber   += float(uber or 0)
            tot_banco  += float(banco or 0)
            tot_z      += float(z or 0)
            tot_total  += float(total or 0)
        return {
            "facturacion": tot_fact,
            "gastos":      tot_gastos,
            "visa":        tot_visa,
            "plataformas": tot_glovo + tot_je + tot_uber,
            "banco":       tot_banco,
            "z":           tot_z,
            "total":       tot_total,
        }

    def _exportar_historico_excel(self):
        """Exporta a Excel exactamente lo que se ve en la tabla del Histórico (según el filtro activo)."""
        anio_sel = self._anios_map.get(self.anio_var.get())
        if anio_sel is not None:
            anio, mes = anio_sel, None
            nombre_periodo = str(anio)
        else:
            anio, mes = self._meses_map.get(self.mes_var.get(), (None, None))
            nombre_periodo = self.mes_var.get().replace(" ", "_") if mes else "Todo_el_historico"

        try:
            rows = load_historico(anio, mes)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el histórico: {e}")
            return

        if not rows:
            messagebox.showinfo("Sin datos", "No hay filas que exportar para el periodo seleccionado.")
            return

        ruta_sugerida = f"cierre_caja_{nombre_periodo}.xlsx"
        ruta = filedialog.asksaveasfilename(
            title="Guardar Excel",
            defaultextension=".xlsx",
            initialfile=ruta_sugerida,
            filetypes=[("Excel", "*.xlsx")]
        )
        if not ruta:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cierre de caja"

            headers = ["Fecha", "Facturación", "Gastos", "Visa", "Internet", "Just Eat",
                       "Glovo", "Uber", "Ticket Rest.", "Ingreso Banco", "Z", "Total"]
            ws.append(headers)
            header_fill = PatternFill(start_color="E85D04", end_color="E85D04", fill_type="solid")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Las filas vienen ordenadas DESC en load_historico; las invertimos para Excel cronológico
            for row in reversed(rows):
                fecha, fact, gastos, visa, inet, je, glovo, uber, tr, banco, z, total = row
                ws.append([
                    fecha, float(fact or 0), float(gastos or 0), float(visa or 0),
                    float(inet or 0), float(je or 0), float(glovo or 0), float(uber or 0),
                    float(tr or 0), float(banco or 0), float(z or 0), float(total or 0),
                ])

            # Formato de columna fecha y columnas numéricas
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
                for c in range(2, 13):
                    ws.cell(row=r, column=c).number_format = '€#,##0.00'

            # Fila de totales
            fila_total = ws.max_row + 1
            ws.cell(row=fila_total, column=1, value="TOTAL").font = Font(bold=True)
            for c in range(2, 13):
                col_letra = openpyxl.utils.get_column_letter(c)
                ws.cell(row=fila_total, column=c,
                        value=f"=SUM({col_letra}2:{col_letra}{fila_total-1})").font = Font(bold=True)
                ws.cell(row=fila_total, column=c).number_format = '€#,##0.00'

            # Ancho de columnas
            ws.column_dimensions["A"].width = 12
            for c in range(2, 13):
                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 13

            ws.freeze_panes = "A2"

            wb.save(ruta)
            self.set_status(f"Exportado a {ruta}", C["success"])
            messagebox.showinfo("Exportado", f"Se han exportado {len(rows)} días a:\n{ruta}")
        except Exception as e:
            messagebox.showerror("Error al exportar", str(e))

# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    app = PizzeriaApp()
    app.mainloop()
